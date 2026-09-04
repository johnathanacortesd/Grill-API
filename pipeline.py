# ======================================
# Pipeline de limpieza y análisis (pipeline.py)
# ======================================
import datetime
import gc
import io
import logging
import math
import os
import re
import time
from typing import Callable, Dict, List, Optional, Tuple
from zipfile import ZipFile
from xml.etree import ElementTree as ET

import numpy as np
import pandas as pd
import xlsxwriter
from openpyxl import load_workbook
from openpyxl.utils.cell import column_index_from_string, coordinate_from_string, range_boundaries
from unidecode import unidecode

from ai_analyzer import enrich_rows_with_ai
from pkl_classifier import (
    apply_pkl_classifiers,
    fill_classification_context,
    load_sklearn_estimator,
)

logger = logging.getLogger("limpieza_grill")

ProgressCb = Optional[Callable[[int, str], None]]
TIPOS_AV = frozenset({"Radio", "Televisión"})
TIPO_MEDIO_MAP = {
    "online": "Internet", "internet": "Internet",
    "diario": "Prensa",
    "am": "Radio", "fm": "Radio", "radio": "Radio",
    "aire": "Televisión", "cable": "Televisión", "tv": "Televisión",
    "television": "Televisión", "televisión": "Televisión",
    "revista": "Revistas", "revistas": "Revistas",
}

# Columnas base (sin las 11 manuales obsoletas)
BASE_OUTPUT_COLUMNS = [
    "ID Noticia", "Fecha", "Hora", "Medio", "Tipo de Medio",
    "Sección - Programa", "Región", "Título", "Autor - Conductor",
    "Nro. Pagina", "Dimensión", "Duración - Nro. Caracteres",
    "CPE", "Tier", "Audiencia",
    "revalorización", "resumen corto",
    "Link Nota", "Resumen - Aclaracion", "Link (Streaming - Imagen)", "Menciones - Empresa",
    "ID duplicada",
]

KEY_MAP = {
    "idnoticia": "ID Noticia",
    "fecha": "Fecha",
    "hora": "Hora",
    "medio": "Medio",
    "tipodemedio": "Tipo de Medio",
    "seccion_programa": "Sección - Programa",
    "region": "Región",
    "titulo": "Título",
    "autor_conductor": "Autor - Conductor",
    "nro_pagina": "Nro. Pagina",
    "dimension": "Dimensión",
    "duracion_caracteres": "Duración - Nro. Caracteres",
    "cpe": "CPE",
    "tier": "Tier",
    "audiencia": "Audiencia",
    "revalorizacion": "revalorización",
    "resumen_corto": "resumen corto",
    "link_nota": "Link Nota",
    "resumen": "Resumen - Aclaracion",
    "link_streaming": "Link (Streaming - Imagen)",
    "menciones": "Menciones - Empresa",
    "idduplicada": "ID duplicada",
}

THOUSANDS_COLS = {"Nro. Pagina", "Dimensión", "Duración - Nro. Caracteres", "Tier", "Audiencia"}
CURRENCY_COLS = {"CPE", "revalorización"}
NUMERIC_COLS = {"ID Noticia", "ID duplicada"} | THOUSANDS_COLS | CURRENCY_COLS
# Display "Link" as black, non-underlined text while keeping the hyperlink.
PLAIN_HYPERLINK_COLUMNS = frozenset({"Link Nota", "Link (Streaming - Imagen)"})

REL_NS = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}"
HYPERLINK_TAIL_BYTES = 4 * 1024 * 1024


def emit_progress(progress: ProgressCb, pct: int, msg: str) -> None:
    pct = max(0, min(100, int(pct)))
    logger.info("%s%% %s", pct, msg)
    if progress:
        progress(pct, msg)


def file_to_bytes(file_obj) -> bytes:
    if isinstance(file_obj, (bytes, bytearray)):
        return bytes(file_obj)
    if hasattr(file_obj, "getvalue"):
        return file_obj.getvalue()
    if hasattr(file_obj, "read"):
        pos = file_obj.tell() if hasattr(file_obj, "tell") else None
        data = file_obj.read()
        if pos is not None and hasattr(file_obj, "seek"):
            try:
                file_obj.seek(pos)
            except Exception:
                pass
        if isinstance(data, bytes):
            return data
        return str(data).encode("utf-8")
    with open(file_obj, "rb") as fh:
        return fh.read()


def _local_tag(tag: str) -> str:
    return tag.rsplit("}", 1)[-1]


def _workbook_rel_path(target: str) -> str:
    t = (target or "").lstrip("/")
    if t.startswith("xl/"):
        return t
    if t.startswith("worksheets/") or t.startswith("theme/") or t == "styles.xml":
        return "xl/" + t
    return t


def extract_hyperlinks_from_xlsx(xlsx_bytes: bytes, sheet_title: str) -> Dict[Tuple[int, int], str]:
    result: Dict[Tuple[int, int], str] = {}
    try:
        with ZipFile(io.BytesIO(xlsx_bytes)) as zf:
            wb_tree = ET.fromstring(zf.read("xl/workbook.xml"))
            rels_tree = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
            rid_to_target = {
                rel.attrib.get("Id"): rel.attrib.get("Target")
                for rel in rels_tree
                if rel.attrib.get("Id") and rel.attrib.get("Target")
            }

            sheet_path = None
            for el in wb_tree.iter():
                if _local_tag(el.tag) != "sheet":
                    continue
                if el.attrib.get("name") != sheet_title:
                    continue
                rid = el.attrib.get(REL_NS + "id") or el.attrib.get("r:id")
                target = rid_to_target.get(rid, "")
                sheet_path = _workbook_rel_path(target)
                break

            if not sheet_path:
                return result
            if sheet_path not in zf.namelist():
                alt = sheet_path[3:] if sheet_path.startswith("xl/") else "xl/" + sheet_path
                if alt in zf.namelist():
                    sheet_path = alt
                else:
                    return result

            dirname, fname = sheet_path.rsplit("/", 1)
            rels_path = f"{dirname}/_rels/{fname}.rels"
            url_by_rid = {}
            if rels_path in zf.namelist():
                st_rels = ET.fromstring(zf.read(rels_path))
                for rel in st_rels:
                    rel_type = rel.attrib.get("Type", "")
                    if rel.attrib.get("TargetMode") == "External" or "hyperlink" in rel_type:
                        rid = rel.attrib.get("Id")
                        target = rel.attrib.get("Target")
                        if rid and target:
                            url_by_rid[rid] = target

            with zf.open(sheet_path) as sheet_fh:
                xml_tail = _read_xml_tail(sheet_fh, HYPERLINK_TAIL_BYTES)

            idx = xml_tail.rfind(b"<hyperlinks")
            if idx < 0:
                return result
            end = xml_tail.find(b"</hyperlinks>", idx)
            chunk = xml_tail[idx:] if end < 0 else xml_tail[idx:end + len(b"</hyperlinks>")]
            try:
                hyper_tree = ET.fromstring(chunk)
            except ET.ParseError:
                logger.warning("No se pudo parsear el bloque de hipervínculos; se continúa sin ellos.")
                return result

            for el in hyper_tree.iter():
                if _local_tag(el.tag) != "hyperlink":
                    continue
                ref = el.attrib.get("ref")
                rid = el.attrib.get(REL_NS + "id") or el.attrib.get("r:id")
                target = url_by_rid.get(rid) if rid else None
                if not target:
                    target = el.attrib.get("location")
                if not ref or not target:
                    continue
                _assign_hyperlink_ref(result, ref, target)
    except Exception:
        logger.exception("Fallo al extraer hipervínculos; se continúa con los valores de celda.")
    return result


def _read_xml_tail(sheet_fh, tail_bytes: int) -> bytes:
    try:
        sheet_fh.seek(0, os.SEEK_END)
        size = sheet_fh.tell()
        start = max(0, size - tail_bytes)
        sheet_fh.seek(start)
        return sheet_fh.read()
    except (OSError, AttributeError):
        return sheet_fh.read()


def _assign_hyperlink_ref(result: dict, ref: str, target: str) -> None:
    if ":" in ref:
        min_col, min_row, max_col, max_row = range_boundaries(ref)
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                result[(r, c)] = target
        return
    col_letter, row = coordinate_from_string(ref)
    result[(row, column_index_from_string(col_letter))] = target


# ======================================
# Utilidades de Limpieza de Texto
# ======================================
def norm_key(text):
    if text is None:
        return ""
    return re.sub(r"[^a-z0-9]+", "", unidecode(str(text).strip().lower()))


def get_column_robust(df, name):
    name_norm = norm_key(name)
    for col in df.columns:
        if norm_key(col) == name_norm:
            return df[col]
    return pd.Series([np.nan] * len(df))


def clean_text(text):
    if not isinstance(text, str):
        return text
    return re.sub(r"\s+", " ", text).strip()


def clean_cuerpo(text):
    if not isinstance(text, str) or text.strip() in ("", "nan", "None"):
        return ""
    text = re.sub(r"<br\s*/?>", "\n", text, flags=re.IGNORECASE)
    text = re.sub(r"<[^>]+>", "", text)
    return text.strip()


def clean_title_for_output(title):
    if not isinstance(title, str):
        return ""
    return re.sub(r"\s+", " ", str(title)).strip()


def corregir_texto(text):
    if not isinstance(text, str) or text.strip() in ("", "nan", "None"):
        return ""
    text = re.sub(r"(<br>|\[\.\.\.\]|\s+)", " ", text).strip()
    m = re.search(r"[A-ZÁÉÍÓÚÑ]", text)
    if m:
        text = text[m.start():]
    if text and not text.endswith("..."):
        text = text.rstrip(".") + "..."
    return text


def normalizar_tipo_medio(tipo_raw):
    if not isinstance(tipo_raw, str):
        return str(tipo_raw)
    t = unidecode(tipo_raw.strip().lower())
    return TIPO_MEDIO_MAP.get(t, str(tipo_raw).strip().title() or "Otro")


def parse_numeric(val):
    if val is None:
        return None
    if isinstance(val, (int, float)):
        if isinstance(val, float) and val.is_integer():
            return int(val)
        return val
    s = str(val).strip()
    if not s:
        return None

    s = re.sub(r"[^\d.,\-eE]", "", s)
    if not s:
        return None

    num_dots = s.count(".")
    num_commas = s.count(",")

    if num_dots > 1 and num_commas == 0:
        s = s.replace(".", "")
    elif num_commas > 1 and num_dots == 0:
        s = s.replace(",", "")
    elif num_dots > 0 and num_commas > 0:
        dot_idx = s.rfind(".")
        comma_idx = s.rfind(",")
        if dot_idx > comma_idx:
            s = s.replace(",", "")
        else:
            s = s.replace(".", "").replace(",", ".")
    elif num_dots == 1:
        parts = s.split(".")
        if len(parts[1]) == 3:
            s = s.replace(".", "")
    elif num_commas == 1:
        parts = s.split(",")
        if len(parts[1]) == 3:
            s = s.replace(",", "")
        else:
            s = s.replace(",", ".")

    try:
        f_val = float(s)
        if f_val.is_integer():
            return int(f_val)
        return f_val
    except ValueError:
        return None


# ======================================
# Algoritmo de Duplicados
# ======================================
def _normalizar_url(url: str) -> str:
    if not url:
        return ""
    url = url.strip().lower()
    url = re.sub(r"^https?://", "", url)
    url = re.sub(r"^www\.", "", url)
    url = url.rstrip("/")
    return url


def _extract_url(val) -> str:
    if val is None:
        return ""
    if isinstance(val, dict):
        return str(val.get("url") or "").strip()
    if isinstance(val, float) and pd.isna(val):
        return ""
    s = str(val).strip()
    if s.lower() in ("", "nan", "none", "link"):
        return ""
    return s


def _normalizar_hora(val) -> str:
    if val is None:
        return ""
    if isinstance(val, float) and pd.isna(val):
        return ""
    if isinstance(val, pd.Timestamp):
        if pd.isna(val):
            return ""
        return val.strftime("%H:%M:%S")
    if isinstance(val, datetime.datetime):
        return val.strftime("%H:%M:%S")
    if isinstance(val, datetime.time):
        return val.strftime("%H:%M:%S")
    s = str(val).strip()
    if s.lower() in ("", "nan", "nat", "none"):
        return ""
    m = re.match(r"^(\d{1,2}):(\d{2})(?::(\d{2}))?", s)
    if m:
        h, mi, se = int(m.group(1)), int(m.group(2)), int(m.group(3) or 0)
        if 0 <= h < 24 and 0 <= mi < 60 and 0 <= se < 60:
            return f"{h:02d}:{mi:02d}:{se:02d}"
    return s


def detectar_duplicados_avanzado(rows, km):
    processed = rows
    seen_url = {}
    seen_bcast = {}

    for i, row in enumerate(processed):
        if row.get("is_duplicate"):
            continue

        tipo = normalizar_tipo_medio(str(row.get(km["tipodemedio"], "")))
        mencion = norm_key(row.get(km["menciones"], ""))
        if not mencion:
            continue

        if tipo in TIPOS_AV:
            medio = norm_key(row.get(km["medio"], ""))
            hora = _normalizar_hora(row.get(km["hora"]))
            if not medio or not hora:
                continue
            k = (mencion, medio, hora)
            if k in seen_bcast:
                row["is_duplicate"] = True
                row[km["idduplicada"]] = processed[seen_bcast[k]].get(km["idnoticia"], "")
            else:
                seen_bcast[k] = i
            continue

        url = _extract_url(row.get("URL Nota")) or _extract_url(row.get(km["link_streaming"]))
        if not url:
            continue
        url_norm = _normalizar_url(url)
        if not url_norm:
            continue
        k = (url_norm, mencion)
        if k in seen_url:
            row["is_duplicate"] = True
            row[km["idduplicada"]] = processed[seen_url[k]].get(km["idnoticia"], "")
        else:
            seen_url[k] = i

    return processed


# ======================================
# Lectura y Estructuración de Datos
# ======================================
def load_dossier_dataframe(file_bytes: bytes, progress: ProgressCb = None) -> pd.DataFrame:
    emit_progress(progress, 8, "Abriendo archivo Excel…")
    try:
        return _load_dossier_calamine(file_bytes, progress)
    except Exception:
        logger.exception("Calamine no pudo leer el xlsx; se usa openpyxl.")
        return _load_dossier_openpyxl(file_bytes, progress)


def _rows_to_dataframe(raw_headers, data_rows, hyperlinks, progress: ProgressCb = None) -> pd.DataFrame:
    rows = []
    empty_streak = 0
    for ridx, values in enumerate(data_rows, start=2):
        if all(v is None or v == "" for v in values):
            empty_streak += 1
            if empty_streak >= 50:
                break
            continue
        empty_streak = 0
        row_data = {}
        for i, h in enumerate(raw_headers):
            if not h:
                continue
            val = values[i] if i < len(values) else None
            url = hyperlinks.get((ridx, i + 1))
            if url:
                row_data[h] = {"value": val or "Link", "url": url}
            else:
                row_data[h] = val
        rows.append(row_data)
        if progress and (len(rows) % 800 == 0):
            emit_progress(
                progress,
                min(39, 16 + len(rows) // 400),
                f"Leyendo filas del dossier… {len(rows)}",
            )
    emit_progress(progress, 40, f"Leídas {len(rows)} filas. Normalizando columnas…")
    return pd.DataFrame(rows)


def _load_dossier_calamine(file_bytes: bytes, progress: ProgressCb = None) -> pd.DataFrame:
    from python_calamine import CalamineWorkbook

    wb = CalamineWorkbook.from_filelike(io.BytesIO(file_bytes))
    sheet_title = wb.sheet_names[0]
    emit_progress(progress, 12, f"Extrayendo hipervínculos de «{sheet_title}»…")
    hyperlinks = extract_hyperlinks_from_xlsx(file_bytes, sheet_title)
    emit_progress(progress, 16, "Leyendo celdas del Excel…")
    sheet = wb.get_sheet_by_name(sheet_title)
    data = sheet.to_python(skip_empty_area=False)
    emit_progress(progress, 22, f"Celdas leídas ({len(data)} filas). Armando tabla…")
    if not data:
        return pd.DataFrame()
    raw_headers = list(data[0])
    return _rows_to_dataframe(raw_headers, data[1:], hyperlinks, progress)


def _load_dossier_openpyxl(file_bytes: bytes, progress: ProgressCb = None) -> pd.DataFrame:
    emit_progress(progress, 8, "Abriendo archivo Excel (openpyxl)…")
    bio = io.BytesIO(file_bytes)
    wb = load_workbook(bio, read_only=True, data_only=True)
    try:
        sheet = wb.active
        sheet_title = sheet.title or "Sheet"
        emit_progress(progress, 12, f"Extrayendo hipervínculos de «{sheet_title}»…")
        hyperlinks = extract_hyperlinks_from_xlsx(file_bytes, sheet_title)
        emit_progress(progress, 16, "Leyendo filas del dossier…")
        it = sheet.iter_rows()
        header_row = next(it, None)
        if header_row is None:
            return pd.DataFrame()
        raw_headers = [c.value for c in header_row]
        data_rows = ([c.value for c in row] for row in it)
        return _rows_to_dataframe(raw_headers, data_rows, hyperlinks, progress)
    finally:
        wb.close()
        bio.close()


def normalize_dossier_dataframe(df, region_map, internet_map, progress: ProgressCb = None):
    if df is None or df.empty:
        return pd.DataFrame()

    emit_progress(progress, 42, "Normalizando tipo de medio, región y columnas…")

    if "Tipo de Medio" in df.columns:
        df["Tipo de Medio"] = (
            df["Tipo de Medio"].astype(str).str.lower().str.strip()
            .map(TIPO_MEDIO_MAP)
            .fillna(df["Tipo de Medio"].astype(str).str.strip())
        )
    else:
        df["Tipo de Medio"] = "Otro"

    is_av = df["Tipo de Medio"].isin(list(TIPOS_AV))
    is_grafica = df["Tipo de Medio"].isin(["Prensa", "Internet", "Revistas"])
    is_internet = df["Tipo de Medio"] == "Internet"

    # BÚSQUEDA ROBUSTA DEL RESUMEN
    cuerpo_series = get_column_robust(df, "CuerpoEs")
    if cuerpo_series.dropna().empty:
        cuerpo_series = get_column_robust(df, "Resumen - Aclaracion")
    if cuerpo_series.dropna().empty:
        cuerpo_series = get_column_robust(df, "Resumen")
    if cuerpo_series.dropna().empty:
        cuerpo_series = get_column_robust(df, "Cuerpo")

    raw_resumen_orig = cuerpo_series

    if "Medio" in df.columns:
        raw_medios_clean = df["Medio"].astype(str).str.lower().str.strip()
        df["Región"] = raw_medios_clean.map(region_map).fillna("N/A")
    else:
        df["Medio"] = "N/A"
        df["Región"] = "N/A"

    if "Medio" in df.columns:
        df.loc[is_internet, "Medio"] = (
            df.loc[is_internet, "Medio"]
            .astype(str).str.lower().str.strip()
            .map(internet_map)
            .fillna(df.loc[is_internet, "Medio"])
        )

    df["ID Noticia"] = df.get("NoticiaId", df.get("ID Noticia", pd.Series(dtype=str)))
    df["Fecha"] = pd.to_datetime(df.get("Fecha", pd.Series(dtype=str)), dayfirst=True, errors="coerce").dt.normalize()
    df["Hora"] = df.get("Hora", pd.Series(dtype=str))
    df["Sección - Programa"] = df.get("Sección - Programa", pd.Series(dtype=str)).astype(str).apply(clean_text)

    titulo_col = "Título" if "Título" in df.columns else "Titulo"
    df["Título"] = df.get(titulo_col, pd.Series(dtype=str)).astype(str).apply(clean_text)
    df["Autor - Conductor"] = df.get("Autor - Conductor", pd.Series(dtype=str)).astype(str).apply(clean_text)
    df["Nro. Pagina"] = df.get("Nro. Pagina", pd.Series(dtype=str))

    dim_col = "Dimensioncm2" if "Dimensioncm2" in df.columns else "Dimensión"
    df["Dimensión"] = df.get(dim_col, pd.Series(dtype=str))
    df["Duración - Nro. Caracteres"] = df.get("Duración - Nro. Caracteres", pd.Series(dtype=str))

    df.loc[is_av, "Dimensión"] = df.loc[is_av, "Duración - Nro. Caracteres"]
    df.loc[is_av, "Duración - Nro. Caracteres"] = 0

    cpe_input = get_column_robust(df, "CPE")
    valor_nota_input = get_column_robust(df, "Valor de Nota")

    df["CPE"] = np.where(is_av, cpe_input, np.where(is_grafica, valor_nota_input, np.nan))
    df["revalorización"] = np.where(is_grafica, cpe_input, np.nan)

    df["Tier"] = df.get("Tier", pd.Series(dtype=str))
    df["Audiencia"] = df.get("Audiencia", pd.Series(dtype=str))

    df["resumen corto"] = raw_resumen_orig.fillna("").astype(str).str.strip()

    emit_progress(progress, 48, "Limpiando cuerpos y enlaces…")

    cuerpo_cleaned = raw_resumen_orig.astype(str).apply(clean_cuerpo)

    def fmt_grafica(text):
        if not isinstance(text, str) or not text.strip():
            return text
        parrafos = [p.strip() for p in text.split("\n") if p.strip()]
        return "\n\n".join(parrafos) if len(parrafos) > 1 else text

    df["Resumen - Aclaracion"] = cuerpo_cleaned
    grafica_mask = is_grafica.fillna(False)
    if grafica_mask.any():
        df.loc[grafica_mask, "Resumen - Aclaracion"] = (
            df.loc[grafica_mask, "Resumen - Aclaracion"].apply(fmt_grafica)
        )

    url_nota_av = df.get("URL Nota AV", df.get("Link Nota AV", pd.Series([""] * len(df))))
    url_streaming = df.get("URL (Streaming - Imagen)", pd.Series([""] * len(df)))

    link_nota_final = []
    for val_av, val_str, is_av_row in zip(url_nota_av, url_streaming, is_av):
        if is_av_row:
            if isinstance(val_av, dict):
                url_t = val_av.get("url", "")
                link_nota_final.append({"value": "Link", "url": url_t.replace(".com.ar", ".com.co") if url_t else None})
            else:
                url_t = str(val_av or "")
                link_nota_final.append({"value": "Link", "url": url_t.replace(".com.ar", ".com.co") if url_t else None})
        else:
            if isinstance(val_str, dict):
                link_nota_final.append(val_str)
            else:
                link_nota_final.append({"value": "Link", "url": val_str if val_str else None})

    df["Link Nota"] = link_nota_final

    url_nota_raw = df.get("URL Nota", pd.Series([""] * len(df)))
    link_stream_final = []
    for val_url, is_int in zip(url_nota_raw, is_internet):
        if is_int:
            if isinstance(val_url, dict):
                link_stream_final.append(val_url)
            else:
                link_stream_final.append({"value": "Link", "url": val_url if val_url else None})
        else:
            link_stream_final.append(None)

    df["Link (Streaming - Imagen)"] = link_stream_final

    menciones_av = df.get("Menciones - Empresa", pd.Series([""] * len(df))).fillna("").astype(str).apply(clean_text)
    menciones_grafica = df.get("Empresa rel.", pd.Series([""] * len(df))).fillna("").astype(str).apply(clean_text)
    df["Menciones - Empresa"] = np.where(is_av, menciones_av, np.where(is_grafica, menciones_grafica, menciones_av))

    emit_progress(progress, 52, "Columnas normalizadas.")
    return df


def expand_menciones(df) -> List[dict]:
    records = df.to_dict("records")
    rows_expanded = []
    for idx, rec in enumerate(records):
        menciones = [m.strip() for m in str(rec.get("Menciones - Empresa", "")).split(";") if m.strip()]
        if not menciones:
            row_dict = dict(rec)
            row_dict["Menciones - Empresa"] = ""
            row_dict["original_index"] = idx
            row_dict["is_duplicate"] = False
            rows_expanded.append(row_dict)
            continue
        for m in menciones:
            row_dict = dict(rec)
            row_dict["Menciones - Empresa"] = m
            row_dict["original_index"] = idx
            row_dict["is_duplicate"] = False
            rows_expanded.append(row_dict)
    return rows_expanded


# ======================================
# Exportar a Excel (XlsxWriter, streaming)
# ======================================
def generate_output_excel(rows, km, progress: ProgressCb = None, columns_to_use: List[str] = None):
    cols = columns_to_use or BASE_OUTPUT_COLUMNS
    buf = io.BytesIO()
    wb = xlsxwriter.Workbook(
        buf,
        {
            "constant_memory": True,
            "strings_to_urls": False,
            "nan_inf_to_errors": False,
        },
    )
    ws = wb.add_worksheet("Resultado")
    fmt_header = wb.add_format({"bold": True})
    fmt_link = wb.add_format({"font_color": "#0563C1", "underline": 1, "align": "left"})
    fmt_plain_hlink = wb.add_format({"font_color": "#000000", "underline": False, "align": "left"})
    fmt_date = wb.add_format({"num_format": "DD/MM/YYYY"})
    fmt_currency = wb.add_format({"num_format": "$#,##0"})
    fmt_thousands = wb.add_format({"num_format": "#,##0"})
    # Formato entero plano sin puntos de miles ni decimales para IDs
    fmt_plain_id = wb.add_format({"num_format": "0"})

    for i, col_name in enumerate(cols):
        if col_name in ["Título", "Resumen - Aclaracion", "resumen corto", "Contexto analizado"]:
            ws.set_column(i, i, 55)
        elif col_name in ["Link Nota", "Link (Streaming - Imagen)"]:
            ws.set_column(i, i, 15)
        elif col_name in ["Subtema_IA", "Tema_IA"]:
            ws.set_column(i, i, 28)
        else:
            ws.set_column(i, i, 20)
        ws.write(0, i, col_name, fmt_header)

    n = len(rows)
    step = max(1, n // 50) if n else 1
    emit_progress(progress, 0, f"Generando archivo de resultado… 0/{n} filas")

    try:
        _write_xlsx_rows(
            ws, rows, km, n, step, progress,
            fmt_link, fmt_plain_hlink, fmt_date, fmt_currency, fmt_thousands, fmt_plain_id, cols,
        )
        emit_progress(progress, 100, "Guardando archivo Excel…")
    finally:
        wb.close()
    return buf.getvalue()


def _write_xlsx_rows(ws, rows, km, n, step, progress, fmt_link, fmt_plain_hlink, fmt_date, fmt_currency, fmt_thousands, fmt_plain_id, cols):
    for i, row in enumerate(rows):
        tk = km.get("titulo")
        if tk and tk in row:
            row[tk] = clean_title_for_output(row.get(tk))
        rk = km.get("resumen")
        if rk and rk in row:
            row[rk] = corregir_texto(row.get(rk))

        excel_row = i + 1
        for cidx, h in enumerate(cols):
            val = row.get(h)
            cv = None
            url = None

            if h == "Fecha" and val is not None and not isinstance(val, dict) and pd.notna(val):
                if isinstance(val, pd.Timestamp):
                    cv = val.to_pydatetime()
                elif isinstance(val, (datetime.datetime, datetime.date)):
                    cv = val
                else:
                    cv = str(val)
            # ID Noticia e ID duplicada se procesan como enteros puros sin separadores
            elif h in ("ID Noticia", "ID duplicada"):
                if val is not None and str(val).strip() not in ("", "nan", "None", "-"):
                    clean_id = re.sub(r"[^\d.]", "", str(val)).strip()
                    if clean_id:
                        try:
                            cv = int(float(clean_s if (clean_s := clean_id) else 0))
                        except ValueError:
                            cv = str(val)
                else:
                    cv = None
            elif h in NUMERIC_COLS:
                cv = parse_numeric(val)
            elif isinstance(val, dict) and "url" in val:
                cv = val.get("value", "Link")
                if val.get("url"):
                    url = val["url"]
            elif val is not None:
                if isinstance(val, str) and val.startswith("http"):
                    cv = "Link"
                    url = val
                else:
                    cv = str(val)

            if url:
                display = str(cv or "Link")
                url_fmt = fmt_plain_hlink if h in PLAIN_HYPERLINK_COLUMNS else fmt_link
                try:
                    ws.write_url(excel_row, cidx, str(url), url_fmt, string=display)
                except Exception:
                    ws.write(excel_row, cidx, display, url_fmt)
            elif h in ("ID Noticia", "ID duplicada") and isinstance(cv, int):
                ws.write_number(excel_row, cidx, cv, fmt_plain_id)
            elif h == "Fecha" and isinstance(cv, datetime.datetime):
                ws.write_datetime(excel_row, cidx, cv, fmt_date)
            elif h == "Fecha" and isinstance(cv, datetime.date):
                ws.write_datetime(
                    excel_row,
                    cidx,
                    datetime.datetime(cv.year, cv.month, cv.day),
                    fmt_date,
                )
            elif h in CURRENCY_COLS and isinstance(cv, (int, float)) and math.isfinite(cv):
                ws.write_number(excel_row, cidx, cv, fmt_currency)
            elif h in THOUSANDS_COLS and isinstance(cv, (int, float)) and math.isfinite(cv):
                ws.write_number(excel_row, cidx, cv, fmt_thousands)
            elif cv is None or cv == "":
                ws.write_blank(excel_row, cidx, None)
            elif isinstance(cv, float) and not math.isfinite(cv):
                ws.write_blank(excel_row, cidx, None)
            else:
                ws.write(excel_row, cidx, cv)

        if progress and (i % step == 0 or i == n - 1):
            emit_progress(
                progress,
                int((i + 1) / n * 100) if n else 100,
                f"Generando archivo de resultado… {i + 1}/{n} filas",
            )


# ======================================
# Modelos PKL opcionales
# ======================================
def _load_optional_pkl_models(ai_config: Optional[dict]):
    if not ai_config:
        return None, None
    tone_model = None
    theme_model = None
    tone_bytes = ai_config.get("tone_pkl_bytes")
    theme_bytes = ai_config.get("theme_pkl_bytes")
    if tone_bytes:
        tone_model, _ = load_sklearn_estimator(tone_bytes, "tono")
    if theme_bytes:
        theme_model, _ = load_sklearn_estimator(theme_bytes, "tema")
    return tone_model, theme_model


# ======================================
# Proceso Principal
# ======================================
def process_dossier(
    file_obj,
    region_map,
    internet_map,
    progress: ProgressCb = None,
    ai_config: Optional[dict] = None
) -> dict:
    t0 = time.time()
    emit_progress(progress, 2, "Cargando archivo…")
    file_bytes = file_to_bytes(file_obj)

    df_normalized = load_dossier_dataframe(file_bytes, progress=progress)
    del file_bytes
    df_normalized = normalize_dossier_dataframe(df_normalized, region_map, internet_map, progress=progress)

    medios_sin_region = []
    if not df_normalized.empty and "Región" in df_normalized.columns and "Medio" in df_normalized.columns:
        medios_sin_region = sorted(set(
            df_normalized.loc[df_normalized["Región"] == "N/A", "Medio"]
            .astype(str).str.strip()
        ) - {"", "nan", "None"})

    emit_progress(progress, 55, "Expandiendo menciones…")
    rows_expanded = expand_menciones(df_normalized)
    del df_normalized
    gc.collect()

    emit_progress(progress, 62, "Detectando duplicados…")
    rows = detectar_duplicados_avanzado(rows_expanded, KEY_MAP)

    # Orden de columnas: Ubicar Contexto analizado, Tono_IA, Tema_IA, Subtema_IA
    # DESPUÉS de 'revalorización' y ANTES de 'resumen corto'
    has_ai = bool(ai_config and ai_config.get("enabled"))
    tone_model, theme_model = _load_optional_pkl_models(ai_config)
    has_pkl = tone_model is not None or theme_model is not None

    if has_ai:
        emit_progress(progress, 70, "Iniciando análisis reputacional con IA…")
        rows = enrich_rows_with_ai(
            rows=rows,
            km=KEY_MAP,
            brand=ai_config["brand"],
            aliases=ai_config.get("aliases", []),
            api_key=ai_config["api_key"],
            model=ai_config.get("model", "gpt-4.1-nano-2025-04-14"),
            progress_callback=progress,
            tone_model=tone_model,
            theme_model=theme_model,
        )
    elif has_pkl:
        emit_progress(progress, 70, "Preparando textos para clasificadores PKL…")
        rows = fill_classification_context(
            rows,
            KEY_MAP,
            brand=(ai_config or {}).get("brand", ""),
            aliases=(ai_config or {}).get("aliases", []),
        )
        emit_progress(progress, 88, "Aplicando modelos PKL del cliente (tono/tema)…")
        rows = apply_pkl_classifiers(
            rows,
            KEY_MAP,
            tone_model=tone_model,
            theme_model=theme_model,
            progress_callback=progress,
            brand=(ai_config or {}).get("brand", ""),
            aliases=(ai_config or {}).get("aliases", []),
        )

    if has_ai or has_pkl:
        rev_idx = BASE_OUTPUT_COLUMNS.index("revalorización")
        ai_cols = ["Contexto analizado", "Tono_IA", "Tema_IA", "Subtema_IA"]
        cols_to_export = BASE_OUTPUT_COLUMNS[:rev_idx + 1] + ai_cols + BASE_OUTPUT_COLUMNS[rev_idx + 1:]
    else:
        cols_to_export = list(BASE_OUTPUT_COLUMNS)

    emit_progress(progress, 94, "✓ Estructuración finalizada. Generando archivo Excel…")

    unique_rows = sum(1 for r in rows if not r.get("is_duplicate"))
    total_rows = len(rows)

    def export_progress(pct, msg):
        overall = 94 + int(pct * 0.06)
        emit_progress(progress, overall, msg)

    output_data = generate_output_excel(rows, KEY_MAP, progress=export_progress, columns_to_use=cols_to_export)
    del rows, rows_expanded
    gc.collect()
    duration = time.time() - t0
    emit_progress(progress, 100, "Limpieza y análisis completados")

    # Nombre del archivo con la primera marca buscada para orden
    if ai_config and ai_config.get("brand"):
        brand_raw = ai_config.get("brand", "")
        clean_tag = re.sub(r"[^\w\s-]", "", unidecode(brand_raw)).strip()
        brand_tag = re.sub(r"[-\s]+", "_", clean_tag)
        filename_prefix = f"Dossier_{brand_tag}" if brand_tag else "Dossier_Limpio"
    else:
        filename_prefix = "Dossier_Limpio"

    timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M')
    final_filename = f"{filename_prefix}_{timestamp}.xlsx"

    return {
        "output_data": output_data,
        "output_filename": final_filename,
        "total_rows": total_rows,
        "unique_rows": unique_rows,
        "duplicates": total_rows - unique_rows,
        "process_duration": f"{duration:.2f}s",
        "medios_sin_mapear": medios_sin_region,
    }
