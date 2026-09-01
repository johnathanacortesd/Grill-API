# ======================================
# Importaciones
# ======================================
import streamlit as st
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, NamedStyle, Alignment
from collections import defaultdict, Counter
from difflib import SequenceMatcher
from copy import deepcopy
import datetime
import io
import openai
import re
import time
from unidecode import unidecode
import numpy as np
from sklearn.metrics.pairwise import cosine_similarity
from sklearn.cluster import AgglomerativeClustering
import json
import asyncio
import hashlib
from typing import List, Dict, Tuple, Optional, Any
import joblib
import gc
import requests
import os
import zipfile
import xml.etree.ElementTree as ET
import html
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor, as_completed

# ======================================
# Configuración general
# ======================================
st.set_page_config(
    page_title="Análisis de Noticias · API - Realizado por Johnathan Cortés",
    page_icon="◈",
    layout="wide",
    initial_sidebar_state="collapsed"
)

def _resolver_modelo_clasificacion():
    """Modelo de clasificación configurable sin redeploy.
    Orden: env OPENAI_CLASIF_MODEL → secret OPENAI_CLASIF_MODEL → gpt-4.1-nano por defecto.
    (gpt-5-nano-2025-08-07 NO es fiable con esta pila: subtemas 'X de Y' y tono todo Neutro.)"""
    env = os.environ.get("OPENAI_CLASIF_MODEL")
    if env:
        return env
    try:
        s = st.secrets.get("OPENAI_CLASIF_MODEL")
        if s:
            return s
    except Exception:
        pass
    return "gpt-4.1-nano-2025-04-14"

OPENAI_MODEL_EMBEDDING     = "text-embedding-3-small"
OPENAI_MODEL_CLASIFICACION = _resolver_modelo_clasificacion()

CONCURRENT_REQUESTS          = 50
SIMILARITY_THRESHOLD_TONO    = 0.94
SIMILARITY_THRESHOLD_TITULOS = 0.92
MAX_PALABRAS_SUBTEMA         = 5

# ── Umbrales base (corpus grande ≥ 20 noticias) ──────────────────────────────
UMBRAL_SUBTEMA = 0.78
UMBRAL_TEMA    = 0.72
NUM_TEMAS_MAX  = 15

UMBRAL_DEDUP_LABEL           = 0.86
UMBRAL_FUSION_SUBTEMAS       = 0.88
UMBRAL_FUSION_INTERGRUPO     = 0.90
MAX_ITER_FUSION              = 3

UMBRAL_MIN_PERTENENCIA_SUBTEMA = 0.60
UMBRAL_MIN_PERTENENCIA_TEMA    = 0.52

UMBRAL_COHERENCIA_ETIQUETA   = 0.35

MAX_GRUPO_ETIQUETA           = 40

# ── Umbrales mínimos de similitud REAL para agrupar ──────────────────────────
SIM_MINIMA_AGRUPACION_SUBTEMA = 0.90
SIM_MINIMA_KEYWORDS_RARAS     = 0.86   
SIM_MINIMA_FUSION_INTER       = 0.90   

PRICE_INPUT_1M     = 0.10
PRICE_OUTPUT_1M    = 0.40
PRICE_EMBEDDING_1M = 0.02

if 'tokens_input' not in st.session_state: st.session_state['tokens_input']     = 0
if 'tokens_output' not in st.session_state: st.session_state['tokens_output']    = 0
if 'tokens_embedding' not in st.session_state: st.session_state['tokens_embedding'] = 0

STOPWORDS_ES = set("""
a ante bajo cabe con contra de desde durante en entre hacia hasta mediante
para por segun sin so sobre tras y o u e la el los las un una unos unas lo
al del se su sus le les mi mis tu tus nuestro nuestros vuestra vuestras este
esta estos estas ese esa esos esas aquel aquella aquellos aquellas que cual
cuales quien quienes cuyo cuya cuyos cuyas como cuando donde cual es son fue
fueron era eran sera seran seria serian he ha han habia han hay hubo habra
habria estoy esta estan estaba estaban estamos estan estar estare estaria
estuvieron estarian estuvo asi ya mas menos tan tanto cada muy todo toda todos
todas ser haber hacer tener poder deber ir dar ver saber querer llegar pasar
encontrar creer decir poner salir volver seguir llevar sentir cambiar
""".split())

_TRAILING_INCOMPLETE = {
    "de","del","la","el","los","las","un","una","unos","unas","al","su","sus",
    "en","con","sin","por","para","sobre","ante","bajo","contra","desde",
    "entre","hacia","hasta","mediante","tras","y","o","u","e","lo","que","se",
    "como","donde","cuando","cual","cuyo","cuya","cuyos","cuyas",
    "este","esta","estos","estas","ese","esa","esos","esas",
    "aquel","aquella","aquellos","aquellas","cada","todo","toda","todos","todas",
    "otro","otra","otros","otras","nuevo","nueva","nuevos","nuevas",
    "gran","grandes","mayor","mayores","menor","menores","mejor","mejores",
    "peor","peores","primer","primera","segundo","segunda","tercer","tercera",
    "más","mas","muy","tan","tanto","tanta","tantos","tantas",
    "mi","mis","tu","tus","nuestro","nuestra","nuestros","nuestras",
    "a","ha","he","ser","estar","haber","hacer","tener","poder","deber",
    "ir","dar","ver","saber","querer","llegar","pasar","decir","poner",
}

_VERBOS_LEAD_SUBTEMA = {
    "levanta", "levantan", "levantaron", "levanto", "impacta", "impactan", "impacto",
    "encarece", "encarecen", "encarecio", "sube", "suben", "subio", "baja", "bajan",
    "bajaron", "bajo", "aumenta", "aumentan", "aumento", "aumentaron", "crece", "crecen",
    "crecio", "crecieron", "gana", "ganan", "gano", "ganaron", "pierde", "pierden",
    "perdio", "pierden", "logra", "logran", "logro", "busca", "buscan", "busco",
    "ofrece", "ofrecen", "ofrecio", "entrega", "entregan", "entrego", "abre", "abren",
    "abrio", "vende", "venden", "vendio", "anuncia", "anuncian", "presenta", "presentan",
    "inaugura", "inauguran", "lanza", "lanzan", "firma", "firman", "solicita", "solicitan",
    "reconoce", "reconocen", "conquista", "conquistan", "inicia", "inician", "llega",
    "llegan", "supera", "superan", "alcanza", "alcanzan", "derrumba", "derrumban",
    "colapsa", "colapsan", "recupera", "recuperan", "avanza", "avanzan", "consolida",
    "consolidan", "espera", "esperan", "planea", "planean", "prepara", "preparan",
    "plantea", "plantean", "planteo", "renuncia", "renuncian", "renuncio", "renuncio",
    "asume", "asumen", "asumio", "asumieron", "posesiona", "posesionan", "posesiono",
    "nombra", "nombran", "nombro", "nombramiento", "designa", "designan", "designo",
    "designacion", "representante", "representa", "representan", "dimite", "dimitio",
    # Verbos inequívocos que faltaban y encabezaban etiquetas ('Investiga a ...').
    "investiga", "investigan", "investigo", "investigaron", "indaga", "indagan",
    "sanciona", "sancionan", "sanciono", "cuestiona", "cuestionan", "cuestiono",
    "critica", "critican", "rechaza", "rechazan", "rechazo", "aprueba", "aprueban",
    "advierte", "advierten", "confirma", "confirman", "confirmo", "revela", "revelan",
    "asegura", "aseguran", "explica", "explican", "afirma", "afirman", "sostiene",
    "sostienen", "responde", "responden", "niega", "niegan", "acusa", "acusan",
    "exige", "exigen", "pide", "piden", "denuncian", "demandan", "multan",
    "suspende", "suspenden", "cancela", "cancelan", "retira", "retiran",
    "amplia", "amplian", "reduce", "reducen", "cierra", "cierran", "reabre",
}

_RE_VERBO_SUBTEMA = re.compile(
    r'\b(presenta|presentan|anuncia|anuncian|lanza|lanzan|inaugura|inauguran|'
    r'realiza|realizan|desarrolla|desarrollan|ejecuta|ejecutan|gestiona|gestionan|'
    r'impulsa|impulsan|promueve|promueven|lidera|lideran|encabeza|encabezan|'
    r'aprueba|aprueban|firma|firman|suscribe|suscriben|invierte|invierten|'
    r'construye|construyen|instala|instalan|entrega|entregan|recibe|reciben|'
    r'solicita|solicitan|visita|visitan|atiende|atienden|destaca|destacan|'
    r'señala|señalan|indica|indican|expresa|expresan|afirma|afirman|'
    r'propone|proponen|pide|piden|exige|exigen|apoya|apoyan|'
    r'informa|informan|reporta|reportan|advierte|advierten|'
    r'levanta|levantan|levantaron|levanto|impacta|impactan|encarece|encarecen|'
    r'encarecio|sube|suben|subio|baja|bajan|bajaron|gano|ganan|ganaron|'
    r'pierde|pierden|perdio|logra|logran|busca|buscan|crece|crecen|'
    r'aumenta|aumentan|conquista|conquistan|derrumba|derrumban|recupera|recuperan|'
    r'plantea|plantean|planteo|renuncia|renuncian|renuncio|asume|asumen|asumio|'
    r'posesiona|posesionan|posesiono|nombra|nombran|nombro|designa|designan|designo|'
    r'representa|representan|dimite|dimitio)\b',
    re.IGNORECASE)

# ======================================
# Subtema: reglas de apuntamiento (grounding) y rechazo de nombres propios/cargos
# ======================================
_CARGOS_SUBTEMA = {
    "alcalde", "alcaldesa", "gobernador", "gobernadora", "ministro", "ministra",
    "viceministro", "viceministra", "presidente", "presidenta", "vicepresidente",
    "vicepresidenta", "director", "directora", "subdirector", "subdirectora",
    "gerente", "rector", "rectora", "vicerrector", "vicerrectora", "decano", "decana",
    "superintendente", "secretario", "secretaria", "procurador", "procuradora",
    "contralor", "contralora", "defensor", "defensora", "magistrado", "magistrada",
    "juez", "jueza", "fiscal", "concejal", "concejala", "senador", "senadora",
    "representante", "diputado", "diputada", "comisionado", "comisionada",
    "embajador", "embajadora", "consul", "vocero", "vocera", "portavoz",
    "comandante", "coronel", "capitan", "obispo", "arzobispo", "cardenal",
    "profesor", "profesora", "medico", "medica",
}

# Sustantivos abstractos/de evento que pueden encabezar un subtema válido.
# (Se usa como "lista verde" para no marcar como nombre propio una cabeza legítima.)
_CABEZAS_SUBTEMA_VALIDAS = {
    "lanzamiento", "apertura", "inauguracion", "estreno", "presentacion", "anuncio",
    "convenio", "acuerdo", "alianza", "pacto", "firma", "colaboracion", "cooperacion",
    "partenariado", "inversion", "proyecto", "programa", "plan", "campana", "cruzada",
    "iniciativa", "propuesta", "estrategia", "politica", "foro", "congreso", "cumbre",
    "encuentro", "feria", "festival", "evento", "ceremonia", "jornada", "debate",
    "conversatorio", "seminario", "taller", "capacitacion", "formacion", "educacion",
    "intercambio", "visita", "gira", "premio", "premiacion", "reconocimiento",
    "distincion", "condecoracion", "homenaje", "nombramiento", "designacion", "posesion",
    "renuncia", "contratacion", "licitacion", "adjudicacion", "convocatoria",
    "investigacion", "indagacion", "denuncia", "demanda", "sancion", "multa",
    "sentencia", "fallo", "auditoria", "fiscalizacion", "veeduria", "regulacion",
    "normativa", "reforma", "aprobacion", "expansion", "fusion", "adquisicion",
    "compra", "venta", "exportacion", "importacion", "comercializacion", "negociacion",
    "construccion", "infraestructura", "modernizacion", "renovacion", "restauracion",
    "rehabilitacion", "ampliacion", "remodelacion", "instalacion", "operacion",
    "intervencion", "prevencion", "atencion", "respuesta", "asistencia", "ayuda",
    "apoyo", "solidaridad", "acompanamiento", "participacion", "adhesion",
    "vinculacion", "integracion", "publicacion", "libro", "informe", "estudio",
    "encuesta", "articulo", "documental", "serie", "podcast", "balance", "resultado",
    "ranking", "clasificacion", "transicion", "transformacion", "digitalizacion",
    "automatizacion", "optimizacion", "mejora", "avance", "progreso", "logro",
    "exito", "triunfo", "crecimiento", "aumento", "reduccion", "crisis", "emergencia",
    "catastrofe", "sismo", "terremoto", "desastre", "gestion", "administracion",
    "coordinacion", "diagnostico", "evaluacion", "explotacion", "trata", "violencia",
    "abuso", "acoso", "discriminacion", "inclusion", "equidad", "movilidad",
    "transporte", "seguridad", "salud", "vivienda", "infancia", "juventud", "cultura",
    "deporte", "turismo", "medioambiente", "energia", "tecnologia", "innovacion",
    "emprendimiento", "empleo", "competitividad", "sostenibilidad", "reputacion",
    "imagen", "comunicacion", "periodismo", "oratoria", "beca", "matricula",
    "graduacion", "egresados", "curriculo", "oferta", "presupuesto", "financiacion",
    "credito", "subsidio", "impuesto", "vacunacion", "salubridad", "cierre",
    "suspension", "clausura", "reactivacion", "reapertura", "defensa", "queja",
    "reclamo", "dialogo", "tregua", "conflicto", "victoria", "derrota", "record",
    "ruta", "hoja", "financiamiento", "fortalecimiento", "cobertura", "puerto",
    "estacion", "memoria", "patrimonio", "identidad", "grieta", "duelo", "salud",
    "cobertura", "plataforma", "bootcamp", "open", "house", "sede", "revision",
}

_CONECTORES_ETIQUETA = {
    "de", "del", "la", "el", "los", "las", "un", "una", "unos", "unas", "al", "lo",
    "y", "e", "o", "u", "en", "sobre", "para", "por", "con", "sin", "ante", "bajo",
    "hacia", "hasta", "entre", "tras", "contra", "desde", "segun", "mediante", "a",
    "que", "como", "cuando", "donde", "cuyo", "cuya", "su", "sus", "mi", "mis",
    "nuestro", "nuestra", "nuestros", "nuestras", "este", "esta", "estos", "estas",
    "ese", "esa", "esos", "esas", "aquel", "aquella", "aquellos", "aquellas",
}

_ARTICULOS_SUBTEMA = {"el", "la", "los", "las", "un", "una", "unos", "unas", "lo"}


def _normaliza_token(w):
    return re.sub(r"[^a-z0-9]", "", unidecode(w.lower()))


def _stem_es(w):
    """Stem liviano en español: quita plurales/género y sufijos flexivos comunes."""
    if len(w) <= 4:
        return w
    for suf in ("aciones", "amientos", "imientos", "dores", "ciones", "siones",
                "idades", "mente", "es", "os", "as", "s", "a", "o", "e"):
        if w.endswith(suf) and len(w) - len(suf) >= 4:
            return w[:len(w) - len(suf)]
    return w


def _nombres_propios_iniciales_titulo(titulo):
    """Toma el 'run' inicial de palabras en mayúscula del titular (sin artículos/nexos)
    => candidatos de NOMBRE PROPIO (persona, lugar, marca) que no deben encabezar
    el subtema."""
    s = str(titulo or "").strip()
    if not s:
        return set()
    toks = re.findall(r"[A-ZÁÉÍÓÚÑÜa-záéíóúñü]+", s)
    res = set()
    for w in toks:
        if not w[:1].isupper():
            break  # terminó el run de mayúsculas iniciales
        norm = unidecode(w.lower())
        if norm in _ARTICULOS_SUBTEMA or norm in {"de", "del", "a", "y", "e", "o", "u"} or len(norm) < 3:
            continue  # artículo/nexo: lo saltamos pero seguimos mirando
        res.add(norm)
        # si tras un nombre propio viene una palabra en minúscula (verbo), terminamos el run
        # (mantenemos el nombre capturado). Proseguimos sólo si el siguiente sigue en mayúscula.
    return res


def _empieza_por_nombre_propio(etiqueta, titulos_fuente=None):
    """True si el subtema arranca por un cargo o por un nombre propio extraído del titular."""
    s = (etiqueta or "").strip().strip('"\'')
    if not s:
        return False
    toks = s.split()
    if not toks:
        return False
    head = _normaliza_token(toks[0])
    if head in _CARGOS_SUBTEMA:
        return True
    if titulos_fuente:
        nombres = set()
        for t in titulos_fuente:
            nombres |= _nombres_propios_iniciales_titulo(t)
        if head in nombres and head not in _CABEZAS_SUBTEMA_VALIDAS:
            return True
    return False


def _contiene_numero_o_acronimo(etiqueta):
    s = unidecode(etiqueta or "")
    if re.search(r"\d", s):
        return True
    for w in s.split():
        w = w.strip(".,;:!?")
        if 1 <= len(w) <= 3 and w.isalpha() and w.isupper() and w not in {"LA", "EL", "LOS", "LAS", "DEL", "UN", "UNA"}:
            return True
    return False


def _subtema_grounded(etiqueta, fuentes):
    """True si TODA palabra de contenido (≥4 letras, no conector) del subtema aparece
    (o deriva de forma evidente por _stem_es) en el TEXTO FUENTE. Evita 'inventos'."""
    if not etiqueta or not fuentes:
        return False
    fuente_tokens = set()
    fuente_stems = set()
    for f in fuentes:
        for w in re.findall(r"[a-z0-9]{4,}", unidecode(str(f).lower())):
            fuente_tokens.add(w)
            fuente_stems.add(_stem_es(w))
    contenido = []
    for w in etiqueta.split():
        wt = _normaliza_token(w)
        if not wt or len(wt) < 4 or wt in _CONECTORES_ETIQUETA:
            continue
        contenido.append(wt)
    if not contenido:
        return True  # sólo conectores / palabras cortas: nada que "inventar"
    no_coinciden = 0
    for w in contenido:
        ws = _stem_es(w)
        if w in fuente_tokens or ws in fuente_tokens or ws in fuente_stems:
            continue
        # coincidencia por raíz/prefijo (mínimo 4 letras en común)
        if any((fs.startswith(w) or w.startswith(fs)) and min(len(fs), len(w)) >= 4
               for fs in fuente_stems):
            continue
        no_coinciden += 1
    # Antes rechazaba con 1 sola palabra no anclada, lo que tiraba a fallback etiquetas
    # legítimas (el tema no siempre aparece literal en un contexto corto). Ahora solo rechaza
    # si MÁS DE LA MITAD de las palabras de contenido no tiene anclaje (label casi entera inventada).
    return no_coinciden * 2 <= len(contenido)


# Núcleos de HECHO de contexto (eventos del cliente) que no están en la lista de
# clase `_NUCLEOS_HECHO`: cubren los casos de una clínica/hospital que ATIENDE a
# alguien (no lo reconoce). Usados por `_derivar_desde_texto_nominal`.
_NUCLEOS_HECHO_EVENTO = [
    (r"\b(atendio|atendida|atendido|atenci[oó]n)\w*", "Atención"),
    (r"\b(trasladad[oa]|remitid[oa]|traslado)\w*", "Traslado"),
    (r"\b(cirug[íi]a|operad[oa]|quir[úu]rgic[oa])\w*", "Cirugía"),
    (r"\b(rehabilitaci[oó]n|terapia|terapeutica)\w*", "Rehabilitación"),
    (r"\b(tratamiento|tratad[oa])\w*", "Tratamiento"),
    (r"\b(diagn[oó]stic[oa]|diagnostica)\w*", "Diagnóstico"),
    (r"\b(hospitalizado|internad[oa]|ingres[oa] al hospital)\w*", "Hospitalización"),
    (r"\b(rescate|rescatad[oa]) \w*", "Rescate"),
    (r"\b(v[íi]ctima|victim[oa])\w*", "Atención a víctima"),
    (r"\b(paciente)\w*", "Atención a paciente"),
    (r"\b(urgencia|emergencias?)\w*", "Atención de urgencia"),
    (r"\b(operac[íi]on de la marca)\w*", "Uso"),
]

# Adjetivos de contexto que acompañan a un hecho (alta complejidad, cirugía compleja)
# y se pegan al sustantivo, no tras 'de'.
_ADJ_POSIBLE = {}


# ── Regla de CABEZA ANCLADA (anti-invención) ──────────────────────────────────
# El núcleo del subtema (su primera palabra de hecho: "Reconocimiento", "Alianza",
# "Premio", "Inversión"...) DEBE estar respaldado en el texto del cliente. Detiene
# el caso real "Reconocimiento a fundación santa fe": 'reconocimiento' no aparece ni
# como palabra ni como verbo derivado ('reconocer') en un texto que solo dice que
# atendieron a un paciente — el LLM lo inventó del titular. Un subtema cuyo NÚCLEO
# no está en el texto es semánticamente falso.
def _head_anclada(etiqueta, fuentes):
    if not etiqueta or not fuentes:
        return False
    contenido = []
    for w in etiqueta.split():
        wt = _normaliza_token(w)
        if not wt or len(wt) < 4 or wt in _CONECTORES_ETIQUETA:
            continue
        contenido.append(wt)
    if not contenido:
        return False
    cabeza = contenido[0]
    fuente_tokens, fuente_stems = set(), set()
    for f in fuentes:
        for w in re.findall(r"[a-z0-9]{4,}", unidecode(str(f).lower())):
            fuente_tokens.add(w)
            fuente_stems.add(_stem_es(w))
    # 1) aparece literal o por stem
    if cabeza in fuente_tokens or _stem_es(cabeza) in fuente_stems:
        return True
    # 2) por prefijo largo compartido (reconocimiento vs reconocer del texto)
    if any((fs.startswith(cabeza) or cabeza.startswith(fs)) and min(len(fs), len(cabeza)) >= 5
           for fs in fuente_stems):
        return True
    # 3) Cabeza derivada de verbo del texto: reconocimiento<-reconocer, alianza<-aliar,
    #    inversión<-invertir, premio<-premiar, campaña<-campaña. Busca la raíz común.
    raiz_cabeza = _stem_es_nominal(cabeza)
    if raiz_cabeza and len(raiz_cabeza) >= 4:
        if any(fs == raiz_cabeza or (fs.startswith(raiz_cabeza) or raiz_cabeza.startswith(fs))
               for fs in fuente_stems):
            return True
    return False


def _stem_es_nominal(w):
    """Raíz verbal/nominal de un sustantivo de hecho: reconocimiento->reconoc,
    alianza->ali, inversión->invers, construcción->construc."""
    w = _normaliza_token(str(w or ""))
    for suf in ("amientos", "imiento", "imientos", "aciones", "acion", "ación",
                "miento", "amient", "siones", "sion", "ción", "cion",
                "encia", "idad", "anzas", "anza", "aje", "tura", "uracion",
                "acion", "cciones", "ccion", "acion", "ores", "oras", "dor"):
        if w.endswith(suf) and len(w) - len(suf) >= 4:
            return w[:len(w) - len(suf)]
    return w


_PATRON_TITULAR = re.compile(
    r"^(nuevo|nueva|anuncia|lanza|presenta|inaugura|llega|abre|inicia|"
    r"logra|alcanza|supera|confirma|destaca|revela|señala|advierte|"
    r"lanzamiento|anuncio|apertura|inicio|presentacion|presentación)\b",
    re.IGNORECASE
)
_PATRON_ESTADO = re.compile(
    r"\b(calma|caos|urgente|hoy|ya|ahora|yesterday|mañana|nuevo|nueva|"
    r"gran|grande|importante|especial|exclusivo)\s*$",
    re.IGNORECASE
)

_TILDE_MAP = {
    "regulacion":"regulación","regulaciones":"regulaciones","innovacion":"innovación",
    "innovaciones":"innovaciones","tecnologia":"tecnología","tecnologias":"tecnologías",
    "tecnologica":"tecnológica","tecnologico":"tecnológico","educacion":"educación",
    "gestion":"gestión","administracion":"administración","informacion":"información",
    "comunicacion":"comunicación","comunicaciones":"comunicaciones","operacion":"operación",
    "operaciones":"operaciones","inversion":"inversión","inversiones":"inversiones",
    "expansion":"expansión","adquisicion":"adquisición","adquisiciones":"adquisiciones",
    "fusion":"fusión","fusiones":"fusiones","transicion":"transición",
    "transformacion":"transformación","digitalizacion":"digitalización",
    "automatizacion":"automatización","modernizacion":"modernización",
    "optimizacion":"optimización","implementacion":"implementación","evaluacion":"evaluación",
    "planificacion":"planificación","organizacion":"organización","atencion":"atención",
    "produccion":"producción","construccion":"construcción","distribucion":"distribución",
    "exportacion":"exportación","importacion":"importación","comercializacion":"comercialización",
    "negociacion":"negociación","negociaciones":"negociaciones","participacion":"participación",
    "colaboracion":"colaboración","asociacion":"asociación","integracion":"integración",
    "relacion":"relación","relaciones":"relaciones","situacion":"situación",
    "condicion":"condición","condiciones":"condiciones","solucion":"solución",
    "soluciones":"soluciones","prevencion":"prevención","proteccion":"protección",
    "fiscalizacion":"fiscalización","sancion":"sanción","sanciones":"sanciones",
    "investigacion":"investigación","investigaciones":"investigaciones","accion":"acción",
    "acciones":"acciones","direccion":"dirección","decision":"decisión",
    "decisiones":"decisiones","eleccion":"elección","elecciones":"elecciones",
    "votacion":"votación","aprobacion":"aprobación","legislacion":"legislación",
    "reclamacion":"reclamación","reclamaciones":"reclamaciones","obligacion":"obligación",
    "obligaciones":"obligaciones","inflacion":"inflación","tributacion":"tributación",
    "financiera":"financiera","financiero":"financiero","economica":"económica",
    "economico":"económico","economia":"economía","credito":"crédito",
    "creditos":"créditos","prestamo":"préstamo","prestamos":"préstamos",
    "interes":"interés","comision":"comisión","comisiones":"comisiones",
    "politica":"política","politicas":"políticas","politico":"político",
    "publica":"pública","publico":"público","estrategia":"estrategia",
    "estrategica":"estratégica","estrategico":"estratégico","logistica":"logística",
    "analisis":"análisis","diagnostico":"diagnóstico","indice":"índice",
    "vehiculo":"vehículo","vehiculos":"vehículos","electrico":"eléctrico",
    "electrica":"eléctrica","energia":"energía","energetica":"energética",
    "petroleo":"petróleo","mineria":"minería","agricola":"agrícola",
    "biologica":"biológica","ecologica":"ecológica","inclusion":"inclusión",
    "exclusion":"exclusión","pension":"pensión","pensiones":"pensiones",
    "jubilacion":"jubilación","compensacion":"compensación","remuneracion":"remuneración",
    "contratacion":"contratación","capacitacion":"capacitación","formacion":"formación",
    "certificacion":"certificación","habilitacion":"habilitación","autorizacion":"autorización",
    "concesion":"concesión","licitacion":"licitación","migracion":"migración",
    "poblacion":"población","recaudacion":"recaudación","asignacion":"asignación",
    "corporacion":"corporación","fundacion":"fundación","institucion":"institución",
    "instituciones":"instituciones","region":"región","unico":"único","unica":"única",
    "ultimo":"último","ultima":"última","proximo":"próximo","basico":"básico",
    "basica":"básica","historico":"histórico","historica":"histórica",
    "medico":"médico","medica":"médica","farmaceutica":"farmacéutica",
    "clinica":"clínica","numero":"número","telefono":"teléfono","telefonia":"telefonía",
    "movil":"móvil","moviles":"móviles","codigo":"código","informatica":"informática",
    "electronica":"electrónica","robotica":"robótica","ciberseguridad":"ciberseguridad",
    "trafico":"tráfico","transito":"tránsito","aereo":"aéreo","maritimo":"marítimo",
    "turistica":"turística","turistico":"turístico","gastronomia":"gastrónomía",
    "academica":"académica","academico":"académico","pedagogica":"pedagógica",
    "cientifica":"científica","cientifico":"científico","juridica":"jurídica",
    "juridico":"jurídico","constitucion":"constitución","resolucion":"resolución",
    "notificacion":"notificación","programacion":"programación","actualizacion":"actualización",
    "verificacion":"verificación","validacion":"validación","liquidacion":"liquidación",
    "facturacion":"facturación","evasion":"evasión","corrupcion":"corrupción",
    "deforestacion":"deforestación","contaminacion":"contaminación","conservacion":"conservación",
    "restauracion":"restauración","rehabilitacion":"rehabilitación","renovacion":"renovación",
    "ampliacion":"ampliación","inauguracion":"inauguración","celebracion":"celebración",
    "clasificacion":"clasificación","eliminacion":"eliminación","motivacion":"motivación",
    "satisfaccion":"satisfacción","reputacion":"reputación","disposicion":"disposición",
}

_ENIE_MAP = {
    "desempeno":"desempeño","desempenos":"desempeños","empeno":"empeño","empenos":"empeños",
    "ensenanza":"enseñanza","ensenanzas":"enseñanzas","diseno":"diseño","disenos":"diseños",
    "disenador":"diseñador","disenadora":"diseñadora","disenadores":"diseñadores",
    "nino":"niño","nina":"niña","ninos":"niños","ninas":"niñas","ninez":"niñez",
    "ano":"año","anos":"años","danio":"daño","danios":"daños","dano":"daño","danos":"daños",
    "danino":"dañino","danina":"dañina","montana":"montaña","montanas":"montañas",
    "espana":"España","espanol":"español","espanola":"española","espanoles":"españoles",
    "companero":"compañero","companera":"compañera","companeros":"compañeros","companeras":"compañeras",
    "compania":"compañía","companias":"compañías","acompanamiento":"acompanamiento",
    "cana":"caña","canas":"cañas","banio":"baño","banios":"baños","bano":"baño","banos":"baños",
    "pena":"peña","penas":"peñas","penon":"peñón","senor":"señor","senora":"señora",
    "senores":"señores","senoras":"señoras","senal":"señal","senales":"señales",
    "senalizacion":"señalización","pequeno":"pequeño","pequena":"pequeña",
    "pequenos":"pequeños","pequenas":"peñas","sueno":"sueño","suenos":"sueños",
    "dueno":"dueño","duena":"dueña","duenos":"dueños","duenas":"dueñas",
    "otono":"otoño","punio":"puño","punios":"puños","puno":"puño",
    "canon":"cañón","canones":"cañones","manana":"mañana","mananas":"mañanas",
    "cabana":"cabaña","cabanas":"cabañas","banera":"bañera","vinedo":"viñedo",
    "vinedos":"viñedos","rebano":"rebaño","rebanos":"rebaños","extrano":"extraño",
    "extrana":"extraña","extranos":"extraños","extranas":"extrañas",
    "enganio":"engaño","engano":"engaño","enganos":"engaños","tamanio":"tamaño",
    "tamano":"tamaño","tamanos":"tamaños","muneca":"muñeca","munecas":"muñecas",
    "cunado":"cuñado","cunada":"cuñada","cunados":"cuñados","albanil":"albañil",
    "albaniles":"albañiles","narino":"Nariño","quindio":"Quindío",
    "ibanez":"Ibáñez","nunez":"Núñez","munoz":"Muñoz","ordonez":"Ordóñez",
    "yanez":"Yáñez","castaneda":"Castañeda","penalosa":"Peñalosa",
    "vineta":"viñeta","vinetas":"viñetas","banado":"bañado","banada":"bañada",
    "rinon":"riñón","rinones":"riñones","panial":"pañal","paniales":"pañales",
    "panal":"pañal","panales":"pañales","arana":"araña","aranas":"arañas",
    "pestana":"pestaña","pestanas":"pestañas","guino":"guiño","guinos":"guiños",
    "munequera":"muñequera","lenador":"leñador","lenadores":"leñadores",
    "resena":"reseña","resenas":"reseñas","panuelo":"pañuelo","panuelos":"pañuelos",
    "companerismo":"compañerismo","desengano":"desengaño","lenio":"leño","leno":"leño",
}

def corregir_tildes(texto: str) -> str:
    if not texto: return texto
    palabras = texto.split()
    resultado = []
    for p in palabras:
        low = p.lower()
        if low in _TILDE_MAP:
            c = _TILDE_MAP[low]
            if p[0].isupper() and not c[0].isupper(): c = c[0].upper() + c[1:]
            resultado.append(c)
        elif low in _ENIE_MAP:
            c = _ENIE_MAP[low]
            if p[0].isupper() and not c[0].isupper(): c = c[0].upper() + c[1:]
            resultado.append(c)
        else:
            resultado.append(p)
    return " ".join(resultado)


# ======================================
# CSS
# ======================================
def load_custom_css():
    st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Google+Sans:wght@400;500;700&family=Google+Sans+Text:wght@400;500;700&family=Roboto+Mono:wght@400;500&display=swap');
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&display=swap');
:root {
    --bg:#f8f9fa;--s1:#ffffff;--s2:#f1f3f4;--s3:#e8eaed;
    --border:#dadce0;--border2:#bdc1c6;--border-focus:#f97316;
    --text:#202124;--text2:#3c4043;--text3:#5f6368;--text4:#9aa0a6;
    --accent:#f97316;--accent2:#ea580c;--accent3:#c2410c;
    --accent-bg:#fff7ed;--accent-bg2:#ffedd5;--accent-bdr:#fed7aa;
    --green:#059669;--green2:#047857;--green-bg:#ecfdf5;--green-bdr:#a7f3d0;
    --red:#dc2626;--amber:#d97706;--blue:#1a73e8;
    --r:8px;--r2:12px;--r3:16px;--r4:20px;
    --shadow-sm:0 1px 2px rgba(60,64,67,0.1),0 1px 3px rgba(60,64,67,0.08);
    --shadow-md:0 1px 3px rgba(60,64,67,0.12),0 4px 8px rgba(60,64,67,0.08);
    --shadow-lg:0 2px 6px rgba(60,64,67,0.1),0 8px 24px rgba(60,64,67,0.1);
    --transition:all 0.2s cubic-bezier(0.4,0,0.2,1);
}
html,body,[data-testid="stApp"]{
    background:var(--bg)!important;color:var(--text)!important;
    font-family:'Google Sans Text','Inter',-apple-system,BlinkMacSystemFont,'Segoe UI',Roboto,sans-serif;
    font-size:14px;-webkit-font-smoothing:antialiased;letter-spacing:0.01em;
}
#MainMenu,footer,header{visibility:hidden}.stDeployButton{display:none}
.block-container{padding-top:1rem!important;padding-bottom:0!important}
[data-testid="stAppViewBlockContainer"]{padding-top:1rem!important}
.app-header{background:var(--s1);border:1px solid var(--border);border-radius:var(--r3);padding:1rem 1.5rem;margin-bottom:1rem;display:flex;align-items:center;gap:1rem;box-shadow:var(--shadow-sm);position:relative;overflow:hidden;}
.app-header-icon{width:40px;height:40px;background:linear-gradient(135deg,#f97316,#ea580c);border-radius:12px;display:flex;align-items:center;justify-content:center;font-size:1.2rem;color:white;flex-shrink:0;box-shadow:0 2px 8px rgba(249,115,22,0.3);}
.app-header-text{flex:1}
.app-header-title{font-family:'Google Sans',sans-serif;font-size:1.25rem;font-weight:700;color:var(--text);letter-spacing:-0.01em;line-height:1.3}
.app-header-version{font-family:'Roboto Mono',monospace;font-size:0.65rem;color:var(--text3);letter-spacing:0.03em;margin-top:0.15rem}
.app-header-badge{background:var(--accent-bg);border:1px solid var(--accent-bdr);color:var(--accent2);font-family:'Roboto Mono',monospace;font-size:0.6rem;font-weight:500;padding:0.25rem 0.75rem;border-radius:100px;letter-spacing:0.04em;text-transform:uppercase;white-space:nowrap;}
[data-testid="stTabs"] [data-testid="stTabsList"]{background:var(--s1)!important;border:1px solid var(--border)!important;border-radius:var(--r2)!important;padding:4px!important;gap:4px!important;box-shadow:var(--shadow-sm)!important;margin-bottom:0.75rem!important;}
[data-testid="stTabs"] button[data-baseweb="tab"]{font-family:'Google Sans',sans-serif!important;font-size:0.88rem!important;font-weight:500!important;color:var(--text2)!important;border-radius:var(--r)!important;padding:0.45rem 1.2rem!important;border:none!important;background:transparent!important;transition:var(--transition)!important;}
[data-testid="stTabs"] button[data-baseweb="tab"]:hover{background:var(--s2)!important;color:var(--text)!important}
[data-testid="stTabs"] button[data-baseweb="tab"][aria-selected="true"]{background:var(--accent-bg)!important;color:var(--accent2)!important;border:1px solid var(--accent-bdr)!important;font-weight:700!important;}
.metrics-grid{display:grid;grid-template-columns:repeat(5,1fr);gap:0.6rem;margin:0.8rem 0}
.metric-card{background:var(--s1);border:1px solid var(--border);border-radius:var(--r2);padding:0.8rem 0.6rem;text-align:center;transition:var(--transition);box-shadow:var(--shadow-sm);position:relative;overflow:hidden;}
.metric-card::before{content:'';position:absolute;top:0;left:0;right:0;height:3px;border-radius:var(--r2) var(--r2) 0 0}
.metric-card.m-total::before{background:linear-gradient(90deg,#5f6368,#9aa0a6)}
.metric-card.m-unique::before{background:linear-gradient(90deg,#059669,#34d399)}
.metric-card.m-dup::before{background:linear-gradient(90deg,#f59e0b,#fbbf24)}
.metric-card.m-time::before{background:linear-gradient(90deg,#1a73e8,#4285f4)}
.metric-card.m-cost::before{background:linear-gradient(90deg,#f97316,#fb923c)}
.metric-card:hover{transform:translateY(-2px);box-shadow:var(--shadow-lg)}
.metric-val{font-family:'Google Sans',sans-serif;font-size:1.5rem;font-weight:700;line-height:1;margin-bottom:0.3rem;letter-spacing:-0.01em}
.metric-lbl{font-family:'Roboto Mono',monospace;font-size:0.62rem;color:var(--text3);text-transform:uppercase;letter-spacing:0.08em;font-weight:500}
[data-testid="stForm"]{background:var(--s1)!important;border:1px solid var(--border)!important;border-radius:var(--r3)!important;padding:1.2rem 1.5rem!important;box-shadow:var(--shadow-md)!important;}
.sec-label{font-family:'Google Sans',sans-serif;font-size:0.72rem;font-weight:700;color:var(--text2);letter-spacing:0.08em;text-transform:uppercase;padding-bottom:0.3rem;border-bottom:2px solid var(--s3);margin:0.8rem 0 0.5rem;display:flex;align-items:center;gap:0.5rem;}
.sec-label::before{content:'';display:inline-block;width:3px;height:12px;background:linear-gradient(180deg,#f97316,#ea580c);border-radius:2px}
.upload-zone{display:grid;grid-template-columns:repeat(3,1fr);gap:0.6rem;margin:0.3rem 0}
.upload-zone-card{background:var(--s1);border:1.5px dashed var(--border);border-radius:var(--r2);padding:0.6rem 0.8rem;display:flex;align-items:center;gap:0.6rem;transition:var(--transition);}
.upload-zone-card:hover{border-color:var(--accent);border-style:solid;transform:translateY(-1px);box-shadow:var(--shadow-md)}
.upload-zone-icon{width:32px;height:32px;border-radius:8px;display:flex;align-items:center;justify-content:center;font-size:1rem;flex-shrink:0;}
.upload-zone-icon.uz-dossier{background:#fff7ed;color:#f97316}
.upload-zone-icon.uz-region{background:#ecfdf5;color:#059669}
.upload-zone-icon.uz-internet{background:#eff6ff;color:#1a73e8}
.upload-zone-text{flex:1;min-width:0}
.upload-zone-title{font-family:'Google Sans',sans-serif;font-size:0.82rem;font-weight:700;color:var(--text);line-height:1.2}
.upload-zone-desc{font-size:0.7rem;color:var(--text3);line-height:1.3}
[data-testid="stFileUploader"]{background:var(--s1)!important;border:1.5px dashed var(--border)!important;border-radius:var(--r)!important;padding:0.4rem 0.6rem!important;transition:var(--transition)!important;min-height:auto!important;}
[data-testid="stFileUploader"]:hover{border-color:var(--accent)!important;border-style:solid!important;background:var(--accent-bg)!important;}
[data-testid="stFileUploader"] section{padding:0.2rem!important}
[data-testid="stFileUploader"] section>div{font-size:0.78rem!important;color:var(--text2)!important}
[data-testid="stFileUploader"] section small{font-size:0.7rem!important;color:var(--text3)!important}
[data-testid="stFileUploader"] button{background:var(--accent-bg)!important;border:1px solid var(--accent-bdr)!important;color:var(--accent2)!important;font-weight:500!important;font-size:0.75rem!important;border-radius:100px!important;padding:0.25rem 0.8rem!important;font-family:'Google Sans',sans-serif!important;transition:var(--transition)!important;}
[data-testid="stFileUploader"] button:hover{background:var(--accent)!important;color:white!important;border-color:var(--accent)!important}
[data-testid="stTextInput"] input,[data-testid="stTextArea"] textarea{background:var(--s1)!important;border:1.5px solid var(--border)!important;color:var(--text)!important;border-radius:var(--r)!important;font-family:'Google Sans Text',sans-serif!important;font-size:0.9rem!important;padding:0.5rem 0.75rem!important;transition:var(--transition)!important;}
[data-testid="stTextInput"] input:focus,[data-testid="stTextArea"] textarea:focus{border-color:var(--accent)!important;box-shadow:0 0 0 3px rgba(249,115,22,0.12)!important;}
[data-testid="stTextInput"] input::placeholder,[data-testid="stTextArea"] textarea::placeholder{color:var(--text4)!important;font-size:0.85rem!important;}
label[data-testid="stWidgetLabel"] p{font-family:'Google Sans',sans-serif!important;color:var(--text2)!important;font-size:0.82rem!important;font-weight:500!important;margin-bottom:0.15rem!important;}
.stButton>button,[data-testid="stDownloadButton"]>button{background:var(--s1)!important;border:1.5px solid var(--border)!important;color:var(--text)!important;border-radius:100px!important;font-family:'Google Sans',sans-serif!important;font-weight:500!important;font-size:0.88rem!important;transition:var(--transition)!important;padding:0.5rem 1.2rem!important;box-shadow:none!important;}
.stButton>button:hover,[data-testid="stDownloadButton"]>button:hover{border-color:var(--accent)!important;color:var(--accent2)!important;background:var(--accent-bg)!important;box-shadow:var(--shadow-sm)!important;transform:translateY(-1px)!important;}
.stButton>button[kind="primary"],[data-testid="stDownloadButton"]>button[kind="primary"]{background:var(--accent)!important;border:none!important;color:#fff!important;font-weight:500!important;font-size:0.92rem!important;padding:0.6rem 1.5rem!important;box-shadow:0 1px 3px rgba(249,115,22,0.3),0 4px 12px rgba(249,115,22,0.15)!important;letter-spacing:0.01em!important;}
.stButton>button[kind="primary"]:hover,[data-testid="stDownloadButton"]>button[kind="primary"]:hover{background:var(--accent2)!important;box-shadow:0 2px 6px rgba(234,88,12,0.35),0 8px 24px rgba(234,88,12,0.18)!important;transform:translateY(-1px)!important;color:#fff!important;}
[data-testid="stRadio"] label{font-family:'Google Sans Text',sans-serif!important;color:var(--text)!important;font-size:0.88rem!important;font-weight:400!important;}
[data-testid="stRadio"]{margin-bottom:0!important}
[data-testid="stRadio"]>div{gap:0!important}
[data-testid="stStatus"]{background:var(--s1)!important;border:1px solid var(--border)!important;border-radius:var(--r2)!important;font-family:'Roboto Mono',monospace!important;font-size:0.8rem!important;}
[data-testid="stAlert"]{background:var(--s1)!important;border:1px solid var(--border)!important;border-radius:var(--r2)!important;color:var(--text2)!important;font-size:0.85rem!important;padding:0.6rem 0.8rem!important;}
.success-banner{background:linear-gradient(135deg,#ecfdf5,#d1fae5);border:1px solid var(--green-bdr);border-left:4px solid var(--green);border-radius:var(--r2);padding:0.8rem 1.2rem;margin:0.5rem 0 0.8rem;display:flex;align-items:center;gap:0.8rem;}
.success-icon{width:34px;height:34px;background:linear-gradient(135deg,#059669,#047857);border-radius:50%;display:flex;align-items:center;justify-content:center;color:white;font-size:1rem;flex-shrink:0;}
.success-title{font-family:'Google Sans',sans-serif;font-size:1rem;font-weight:700;color:#047857;margin-bottom:0.1rem}
.success-sub{font-size:0.8rem;color:var(--text2)}
.auth-wrap{max-width:380px;margin:8vh auto 0;text-align:center}
.auth-icon{width:60px;height:60px;background:linear-gradient(135deg,#f97316,#ea580c);border-radius:16px;display:inline-flex;align-items:center;justify-content:center;font-size:1.6rem;color:white;margin-bottom:1rem;box-shadow:0 4px 16px rgba(249,115,22,0.3);}
.auth-title{font-family:'Google Sans',sans-serif;font-size:1.5rem;font-weight:700;color:var(--text);margin-bottom:0.3rem}
.auth-sub{font-size:0.85rem;color:var(--text3);margin-bottom:2rem}
.cluster-info{background:var(--accent-bg);border:1px solid var(--accent-bdr);border-radius:var(--r);padding:0.5rem 0.8rem;margin:0.4rem 0;font-family:'Roboto Mono',monospace;font-size:0.68rem;color:var(--text2);line-height:1.6;}
.cluster-info b{color:var(--accent2);font-size:0.72rem}
.config-badge{display:inline-flex;align-items:center;gap:0.4rem;background:var(--s2);border:1px solid var(--border);border-radius:100px;padding:0.2rem 0.7rem;font-family:'Roboto Mono',monospace;font-size:0.62rem;color:var(--text3);margin-bottom:0.6rem;}
[data-testid="stProgressBar"]>div>div{background:linear-gradient(90deg,#f97316,#fb923c,#fdba74)!important;border-radius:100px!important;height:5px!important;}
[data-testid="stDataFrame"]{border:1px solid var(--border)!important;border-radius:var(--r2)!important;box-shadow:var(--shadow-sm)!important;overflow:hidden!important;}
::-webkit-scrollbar{width:6px;height:6px}
::-webkit-scrollbar-track{background:var(--s2);border-radius:3px}
::-webkit-scrollbar-thumb{background:var(--border2);border-radius:3px}
::-webkit-scrollbar-thumb:hover{background:var(--accent)}
.footer{font-family:'Roboto Mono',monospace;font-size:0.6rem;color:var(--text4);text-align:center;padding:0.8rem 0 0.5rem;letter-spacing:0.04em;border-top:1px solid var(--s3);margin-top:1rem;}
.stElementContainer{margin-bottom:0!important}
[data-testid="stVerticalBlock"]>div{gap:0.3rem!important}
[data-testid="stHorizontalBlock"]>div{gap:0.4rem!important}
hr{border-color:var(--s3)!important;margin:0.5rem 0!important}
[data-testid="stSelectbox"]>div>div{font-family:'Google Sans Text',sans-serif!important;font-size:0.88rem!important;color:var(--text)!important;}
@media(max-width:768px){
    .metrics-grid{grid-template-columns:repeat(2,1fr)}
    .upload-zone{grid-template-columns:1fr}
    .app-header{flex-direction:column;text-align:center;gap:0.5rem;padding:1rem}
}
</style>
""", unsafe_allow_html=True)


# ======================================
# Umbrales adaptativos según tamaño del corpus
# ======================================
def _umbrales_adaptativos(n: int) -> dict:
    if n <= 5:
        return dict(
            subtema=0.93,
            tema=0.85,
            dedup_label=0.90,
            fusion_subtemas=0.92,
            fusion_intergrupo=0.95,
            min_pertenencia_subtema=0.80,
            min_pertenencia_tema=0.75,
            coherencia_etiqueta=0.50,
            sim_minima_agrupacion=0.93,
            sim_minima_keywords=0.93,
            max_iter_fusion=1,
            num_temas_max=n,
            usar_paso2b=False,
            usar_fusion_iterativa=False,
        )
    elif n <= 10:
        return dict(
            subtema=0.90,
            tema=0.84,
            dedup_label=0.88,
            fusion_subtemas=0.90,
            fusion_intergrupo=0.93,
            min_pertenencia_subtema=0.72,
            min_pertenencia_tema=0.65,
            coherencia_etiqueta=0.42,
            sim_minima_agrupacion=0.90,
            sim_minima_keywords=0.90,
            max_iter_fusion=2,
            num_temas_max=min(n, 5),
            usar_paso2b=False,
            usar_fusion_iterativa=False,
        )
    elif n <= 20:
        return dict(
            subtema=0.87,
            tema=0.82,
            dedup_label=0.86,
            fusion_subtemas=0.88,
            fusion_intergrupo=0.91,
            min_pertenencia_subtema=0.66,
            min_pertenencia_tema=0.58,
            coherencia_etiqueta=0.38,
            sim_minima_agrupacion=0.87,
            sim_minima_keywords=0.87,
            max_iter_fusion=3,
            num_temas_max=min(n // 2, NUM_TEMAS_MAX),
            usar_paso2b=True,
            usar_fusion_iterativa=True,
        )
    else:
        return dict(
            subtema=UMBRAL_SUBTEMA,
            tema=UMBRAL_TEMA,
            dedup_label=UMBRAL_DEDUP_LABEL,
            fusion_subtemas=UMBRAL_FUSION_SUBTEMAS,
            fusion_intergrupo=UMBRAL_FUSION_INTERGRUPO,
            min_pertenencia_subtema=UMBRAL_MIN_PERTENENCIA_SUBTEMA,
            min_pertenencia_tema=UMBRAL_MIN_PERTENENCIA_TEMA,
            coherencia_etiqueta=UMBRAL_COHERENCIA_ETIQUETA,
            sim_minima_agrupacion=SIM_MINIMA_AGRUPACION_SUBTEMA,
            sim_minima_keywords=SIM_MINIMA_KEYWORDS_RARAS,
            max_iter_fusion=MAX_ITER_FUSION,
            num_temas_max=NUM_TEMAS_MAX,
            usar_paso2b=True,
            usar_fusion_iterativa=True,
        )


# ======================================
# Caché Global de Embeddings
# ======================================
class EmbeddingCache:
    def __init__(self):
        self._cache: Dict[str, List[float]] = {}
        self._hits = 0
        self._misses = 0
        self._dirty = False
        self._disk_path = None
        # Persistencia ligera en disco (local). En Streamlit Cloud el FS es efímero y
        # nunca rompe nada si falla la escritura: todo quede en try/except.
        try:
            d = Path(os.environ.get("GRILL_CACHE_DIR", str(Path.home() / ".grill_cache")))
            d.mkdir(parents=True, exist_ok=True)
            self._disk_path = d / "embeddings.json"
            if self._disk_path.exists() and self._disk_path.stat().st_size < 80 * 1024 * 1024:
                with open(self._disk_path, "r", encoding="utf-8") as f:
                    self._cache.update(json.load(f))
        except Exception:
            self._disk_path = None

    def _key(self, text):
        return hashlib.md5(text[:2000].encode('utf-8', errors='ignore')).hexdigest()

    def get(self, text):
        k = self._key(text)
        if k in self._cache:
            self._hits += 1
            return self._cache[k]
        self._misses += 1
        return None

    def put(self, text, emb):
        self._cache[self._key(text)] = emb
        self._dirty = True

    def get_many(self, textos):
        results = [None] * len(textos)
        missing = []
        for i, t in enumerate(textos):
            c = self.get(t)
            if c is not None:
                results[i] = c
            else:
                missing.append(i)
        return results, missing

    def flush(self):
        """Persiste el caché a disco (una sola escritura; ignora fallos)."""
        if not self._disk_path or not self._dirty:
            return
        try:
            tmp = str(self._disk_path) + ".tmp"
            with open(tmp, "w", encoding="utf-8") as f:
                json.dump(self._cache, f, separators=(",", ":"))
            os.replace(tmp, str(self._disk_path))
            self._dirty = False
        except Exception:
            self._dirty = False

    def stats(self):
        total = self._hits + self._misses
        rate = (self._hits / total * 100) if total > 0 else 0
        return f"Cache: {self._hits} hits, {self._misses} misses ({rate:.0f}%)"

    def clear(self):
        self._cache.clear()
        self._hits = 0
        self._misses = 0

    def reset_stats(self):
        self._hits = 0
        self._misses = 0

if '_emb_cache' not in st.session_state:
    st.session_state['_emb_cache'] = EmbeddingCache()

def get_embedding_cache():
    return st.session_state['_emb_cache']

# ======================================
# Configuración vía Google Sheets (CSV público)
# ======================================
CONFIG_CACHE_TTL = 300  # segundos

@st.cache_data(ttl=CONFIG_CACHE_TTL, show_spinner=False)
def _fetch_map_from_csv(csv_url: str) -> dict:
    df = pd.read_csv(csv_url, header=None, dtype=str)
    df = df.dropna(how="all")
    mapping = pd.Series(
        df.iloc[:, 1].values,
        index=df.iloc[:, 0].astype(str).str.lower().str.strip()
    ).to_dict()
    mapping = {k: v for k, v in mapping.items() if k not in ("nan", "")}
    return mapping

def load_config_from_sheets():
    regiones_url = st.secrets.get("REGIONES_CSV_URL")
    internet_url = st.secrets.get("INTERNET_CSV_URL")

    if not regiones_url or not internet_url:
        st.error(
            "❌ Faltan las URLs de configuración. Agrega REGIONES_CSV_URL e "
            "INTERNET_CSV_URL en los Secrets de la app."
        )
        st.stop()

    try:
        region_map = _fetch_map_from_csv(regiones_url)
        internet_map = _fetch_map_from_csv(internet_url)
    except Exception as e:
        st.error(f"❌ No se pudo leer la configuración desde Google Sheets: {e}")
        st.stop()

    return region_map, internet_map

def refresh_config_cache():
    _fetch_map_from_csv.clear()


# ======================================
# Funciones Auxiliares de Limpieza, Enlaces y Conversión
# ======================================

def check_password():
    if st.session_state.get("password_correct", False):
        return True
    st.markdown("""
    <div class="auth-wrap">
        <div class="auth-icon">◈</div>
        <div class="auth-title">Sistema de Análisis</div>
        <div class="auth-sub">Ingresa tus credenciales para continuar</div>
    </div>""", unsafe_allow_html=True)
    _, col, _ = st.columns([1, 2, 1])
    with col:
        with st.form("pw"):
            pw = st.text_input("Contraseña", type="password", placeholder="Ingresa tu contraseña")
            if st.form_submit_button("Ingresar", use_container_width=True, type="primary"):
                if pw == st.secrets.get("APP_PASSWORD", "INVALID"):
                    st.session_state["password_correct"] = True
                    st.rerun()
                else:
                    st.error("Contraseña incorrecta")
    return False

def call_with_retries(fn, *a, **kw):
    d = 1
    for att in range(3):
        try:
            return fn(*a, **kw)
        except Exception as e:
            if att == 2: raise e
            time.sleep(d)
            d *= 2

async def acall_with_retries(fn, *a, **kw):
    d = 1
    for att in range(3):
        try:
            return await fn(*a, **kw)
        except Exception as e:
            if att == 2: raise e
            await asyncio.sleep(d)
            d *= 2

def norm_key(text):
    if text is None: return ""
    return re.sub(r"[^a-z0-9]+", "", unidecode(str(text).strip().lower()))

def capitalizar_etiqueta(tema):
    if not tema or not tema.strip(): return "Sin tema"
    tema = tema.strip().lower()
    tema = corregir_tildes(tema)
    return tema[0].upper() + tema[1:]

def _frase_esta_completa(texto):
    if not texto or not texto.strip(): return False
    palabras = texto.strip().split()
    if not palabras: return False
    ultima = palabras[-1].lower().rstrip(".,;:!?")
    return unidecode(ultima) not in _TRAILING_INCOMPLETE and len(ultima) > 1

def _recortar_frase_completa(texto, max_palabras=7):
    if not texto: return "Sin tema"
    palabras = texto.strip().split()
    if len(palabras) > max_palabras: palabras = palabras[:max_palabras]
    while palabras and unidecode(palabras[-1].lower().rstrip(".,;:!?")) in _TRAILING_INCOMPLETE:
        palabras.pop()
    if not palabras: return texto.strip().split()[0] if texto.strip() else "Sin tema"
    return " ".join(palabras)

def limpiar_tema(tema):
    if not tema: return "Sin tema"
    tema = tema.strip().strip('"\'')
    for px in ["subtema:", "tema:", "categoría:", "categoria:", "category:"]:
        if tema.lower().startswith(px): tema = tema[len(px):].strip()
    tema = _recortar_frase_completa(tema, max_palabras=MAX_PALABRAS_SUBTEMA)
    return capitalizar_etiqueta(tema) if tema else "Sin tema"

def limpiar_tema_geografico(tema, marca, aliases):
    if not tema: return "Sin tema"
    tl = unidecode(tema.lower())
    for n in [marca] + [a for a in aliases if a]:
        patron = r'\b' + re.escape(unidecode(n.strip().lower())) + r'\b'
        tl = re.sub(patron, '', tl)
    frases_eliminar = [
        "en colombia", "de colombia", "del pais", "en el pais",
        "territorio nacional", "a nivel nacional", "en todo el pais",
    ]
    for frase in frases_eliminar:
        tl = re.sub(r'\b' + re.escape(frase) + r'\b', '', tl)
    tl = re.sub(r'\s+', ' ', tl).strip()
    if not tl: return "Sin tema"
    tokens_orig = tema.split()
    tokens_norm = unidecode(tema.lower()).split()
    norm_disponibles = tl.split()
    resultado_tokens = []
    for orig, norm in zip(tokens_orig, tokens_norm):
        if norm_disponibles and norm == norm_disponibles[0]:
            resultado_tokens.append(orig)
            norm_disponibles.pop(0)
    resultado = " ".join(resultado_tokens).strip()
    resultado = corregir_tildes(resultado) if resultado else ""
    return limpiar_tema(resultado) if resultado.strip() else "Sin tema"

def string_norm_label(s):
    if not s: return ""
    s = unidecode(s.lower())
    s = re.sub(r"[^a-z0-9\s]", " ", s)
    return " ".join(t for t in s.split() if t not in STOPWORDS_ES)
# Rótulos genéricos que NUNCA deben aparecer como subtema (el usuario los rechaza).
_ETIQUETAS_GENERICAS_INVALIDAS = {
    "cobertura de informacion relevante",
    "cobertura informativa general",
    "cobertura de informacion",
    "cobertura de noticias",
    "cobertura noticiosa",
    "informacion relevante",
    "informacion importante",
    "informacion general",
    "noticia relevante",
    "noticias relevantes",
    "noticia importante",
    "noticias importantes",
    "sin tema", "varios", "n a", "general", "noticias", "informacion", "tema", "temas",
    "gestion", "actividades", "acciones", "eventos",
    "actividad corporativa", "gestion corporativa",
    "impacto en la reputacion", "impacto reputacional",
    "reputacion corporativa", "impacto corporativo",
    # ── Rótulos "de relleno" que el modelo inventa cuando no deriva el hecho ──
    # (el usuario los reporta explícitamente: "destacados del sector" y variantes)
    "destacados del sector", "destacados sectoriales", "destacados del gremio",
    "panorama del sector", "panorama sectorial", "panorama general",
    "actualidad del sector", "actualidad sectorial", "actualidad empresarial",
    "actualidad institucional", "actividad institucional", "actividad del sector",
    "noticias del sector", "noticias sectoriales", "informacion del sector",
    "contexto del sector", "contexto general", "contexto general del sector",
    "situacion del sector", "dinamica del sector", "temas del sector",
    "sector", "sectorial", "el sector", "del sector",
    "menciones de la marca", "mencion de la marca", "menciones en medios",
    "presencia en medios", "presencia mediatica", "cobertura mediatica",
    "cobertura en medios", "apariciones en medios", "vision general",
    "resumen de noticias", "resumen informativo", "diversos temas",
    "otros temas", "temas varios", "aspectos generales", "generalidades",
    "declaraciones", "declaraciones de la marca", "comunicado de la marca",
}

# Patrones de etiqueta-relleno: no describen un HECHO, solo el marco/soporte.
# REGLA ESTRUCTURAL, NO LÉXICA POR SECTOR: la etiqueta es relleno si su CABEZA es
# un marco informativo y su complemento es otro término de marco, seguido de
# cualquier calificador de industria ('financiero', 'minero', 'público', 'avícola'...).
# Así funciona para cualquier cliente sin listar sectores uno por uno.
# OJO: el marco solo es relleno si NO va seguido de un asunto real. Por eso el
# complemento debe ser vacío, un adjetivo de marco o un núcleo de marco; así
# 'Cobertura de vacunación infantil' (subtema legítimo) NO se rechaza, pero
# 'Cobertura en medios' / 'Panorama del sector financiero' sí.
_ADJ_MARCO = (r"general(?:es)?|informativ[oa]s?|mediatic[oa]s?|corporativ[oa]s?|"
              r"institucional(?:es)?|sectorial(?:es)?|empresarial(?:es)?|"
              r"relevantes?|importantes?|destacad[oa]s?|varios|varias|diversos|diversas")
_NUC_MARCO = (r"sector(?:es)?|gremio(?:s)?|marca(?:s)?|medios?|mercado(?:s)?|industria(?:s)?|"
              r"empresa(?:s)?|compania(?:s)?|noticia(?:s)?|informacion|informe(?:s)?|"
              r"tema(?:s)?|actualidad|panorama|contexto|prensa|pais|colombia|region|"
              r"entidad(?:es)?|institucion(?:es)?|organizacion(?:es)?|negocio(?:s)?|"
              r"agenda|coyuntura|ambito|entorno|escenario|rubro|ramo|clientes?|publico")
_RE_ETIQUETA_RELLENO = re.compile(
    r"^(?:destacad[oa]s?|panorama|actualidad|generalidades|vision|resumen|"
    r"contexto|situacion|dinamica|coyuntura|menciones?|presencia|cobertura|aparicion(?:es)?|"
    r"informacion|noticias?|novedades|tema|temas|aspectos?|otros|otras|varios|diversos)"
    r"(?:\s+(?:" + _ADJ_MARCO + r"))*"
    r"(?:\s+(?:de|del|de\s+la|en|sobre|para|al|a)\s+"
    r"(?:la\s+|el\s+|los\s+|las\s+)?(?:" + _NUC_MARCO + r")"
    # Calificador de industria/ámbito: CUALQUIER palabra (financiero, minero,
    # público, avícola, energético...). No aporta hecho, solo nombra el rubro.
    r"(?:\s+[a-z]+){0,2}"
    r")?\s*$",
    re.IGNORECASE,
)

# Núcleos que por sí solos NO son un hecho: si la etiqueta es solo marco + lugar
# o marco + 'sector/marca/medios', es relleno.
_NUCLEOS_RELLENO_SUBTEMA = {
    "sector", "sectores", "gremio", "gremios", "marca", "marcas", "medios",
    "mercado", "industria", "empresa", "empresas", "compania", "companias",
    "noticia", "noticias", "informacion", "informe", "cobertura", "presencia",
    "mencion", "menciones", "actualidad", "panorama", "contexto", "generalidades",
}
# Rótulos que jamás deben quedar como subtema final (marcadores vacíos o genéricos).
# Una sola palabra REAL extraída del texto Sí vale, nunca el placeholder 'Sin tema'.
_PLACEHOLDER_SUBTEMA = _ETIQUETAS_GENERICAS_INVALIDAS | {
    'n/a', 'nan', '-', 'noticia', 'informe',
}


# Sustantivos de ACONTECIMIENTO: denotan que algo PASÓ. Se distinguen de los
# nombres de materia/rubro ('salud', 'educación', 'tecnología'), que describen un
# campo pero no un hecho. Sirve para exigir que la etiqueta diga QUÉ PASÓ, en
# cualquier cliente y sector.
_NUCLEOS_ACONTECIMIENTO = {
    "alza", "aumento", "incremento", "subida", "reduccion", "caida", "baja",
    "descenso", "disminucion", "crecimiento", "contraccion",
    "convenio", "alianza", "acuerdo", "pacto", "cooperacion", "colaboracion",
    "memorando", "consorcio", "negociacion",
    "inversion", "expansion", "ampliacion", "adquisicion", "compra", "venta",
    "fusion", "capitalizacion", "financiamiento", "financiacion", "credito",
    "lanzamiento", "estreno", "presentacion", "apertura", "inauguracion",
    "reapertura", "activacion", "cierre", "suspension", "clausura", "cancelacion",
    "investigacion", "indagacion", "denuncia", "demanda", "sancion", "multa",
    "condena", "querella", "imputacion", "fallo", "sentencia", "auditoria",
    "litigio", "captura", "allanamiento", "extincion",
    "premio", "reconocimiento", "galardon", "distincion", "condecoracion",
    "certificacion", "acreditacion", "homenaje", "ranking",
    "nombramiento", "designacion", "posesion", "renuncia", "dimision", "relevo",
    "sucesion", "eleccion", "votacion", "asamblea",
    "reforma", "regulacion", "ley", "decreto", "norma", "normativa", "resolucion",
    "aprobacion", "rechazo", "veto", "sancion presidencial",
    "obra", "construccion", "restauracion", "rehabilitacion", "remodelacion",
    "modernizacion", "renovacion", "mantenimiento", "instalacion", "adecuacion",
    "crisis", "emergencia", "desastre", "sismo", "terremoto", "inundacion",
    "incendio", "derrame", "accidente", "colapso", "falla", "fallas",
    "racionamiento", "apagon", "paro", "bloqueo", "protesta", "marcha", "huelga",
    "brote", "contagio", "epidemia", "pandemia", "vacunacion", "intoxicacion",
    "balance", "resultados", "utilidad", "utilidades", "ingresos", "perdidas",
    "ganancias", "rentabilidad", "presupuesto", "recaudo", "deficit", "superavit",
    "campana", "publicidad", "patrocinio", "comunicado", "declaracion",
    "donacion", "solidaridad", "ayuda", "apoyo", "voluntariado", "asistencia",
    "empleo", "contratacion", "capacitacion", "formacion", "despidos", "vacantes",
    "estudio", "informe", "encuesta", "analisis", "diagnostico", "evaluacion",
    "foro", "congreso", "cumbre", "feria", "seminario", "taller", "encuentro",
    "conversatorio", "jornada", "evento", "festival", "concierto",
    "proyecto", "programa", "plan", "iniciativa", "estrategia", "convocatoria",
    "licitacion", "adjudicacion", "contrato", "concesion",
    "exportaciones", "importaciones", "produccion", "consumo", "demanda",
    "oferta", "ventas", "precio", "precios", "tarifa", "tarifas", "costo", "costos",
    "subasta", "remate", "escasez", "desabastecimiento", "abastecimiento",
    "beca", "becas", "matricula", "graduacion", "publicacion", "hallazgo",
    "descubrimiento", "acuerdo comercial", "alianza estrategica",
}


def _es_calificador_de_ambito(token) -> bool:
    """¿El token es un simple calificador de industria/ámbito, sin hecho propio?

    Detección MORFOLÓGICA, independiente del cliente: los adjetivos relacionales
    de sector en español terminan en -ero/-era, -ario/-aria, -ico/-ica, -al,
    -ivo/-iva, -ano/-ana, -uario, -ista, -tor/-tora ('financiero', 'minero',
    'portuario', 'energético', 'institucional', 'automotor', 'avícola').
    Así 'Destacados del sector X' se rechaza para CUALQUIER sector, sin listarlos.

    No marca sustantivos de hecho ('convenio', 'inversión'): esos se protegen con
    `_es_cabeza_subtema_valida` antes de aplicar el sufijo.
    """
    w = _normaliza_token(str(token or ""))
    if not w or len(w) < 4:
        return False
    if w in {"general", "generales", "relevante", "relevantes", "importante",
             "importantes", "nacional", "nacionales", "internacional", "regional",
             "local", "publico", "publica", "privado", "privada", "global"}:
        return True
    if _es_cabeza_subtema_valida(w):
        return False        # es un sustantivo de evento conocido
    # Adjetivos de contexto clínico/periodístico que no caen en sufijos estándar.
    if w in {"compleja", "complejas", "complejo", "complejos", "alta", "altas",
             "alto", "altos", "corta", "cortas", "corto", "cortos", "larga", "largas",
             "largo", "largos", "nueva", "nuevas", "nuevo", "nuevos", "especial",
             "especiales", "directa", "directas", "directo", "directos", "abierta",
             "abiertas", "abierto", "abiertos", "privada", "privadas", "privado",
             "privados", "extranjera", "extranjeras", "extranjero", "extranjeros",
             "local", "locales", "rural", "rurales", "urbana", "urbanas", "urbano",
             "urbanos", "publica", "publicas", "publico", "publicos"}:
        return True
    if w in _SUSTANTIVOS_SEGUROS_FINALES:
        return False
    return bool(re.search(
        r"(?:ero|era|eros|eras|ario|aria|arios|arias|uario|uaria|"
        r"ico|ica|icos|icas|ivo|iva|ivos|ivas|al|ales|"
        r"ano|ana|anos|anas|ista|istas|tor|tora|tores|"
        r"ense|enses|il|iles|ola|olas|udo|uda)$", w))


def _es_etiqueta_generica(etiqueta) -> bool:
    """True si la etiqueta es un rótulo genérico/vacío (p. ej. 'Cobertura de
    información relevante', 'Destacados del sector') o una sola palabra sin hecho
    concreto. Un subtema debe nombrar un HECHO, no el marco ni el soporte."""
    if not etiqueta or not str(etiqueta).strip():
        return True
    n = re.sub(r"[^a-z0-9\s]", " ", unidecode(str(etiqueta).lower()))
    n = re.sub(r"\s+", " ", n).strip()
    if not n or n in _ETIQUETAS_GENERICAS_INVALIDAS:
        return True
    if len(n.split()) <= 1:
        return True
    # Rótulo-relleno: empieza por un marco informativo ('destacados', 'panorama',
    # 'actualidad', 'menciones'...) y no aporta un hecho.
    if _RE_ETIQUETA_RELLENO.match(n):
        return True
    # 'Sector/Mercado/Industria + <rubro>' sin hecho: 'Sector salud', 'Sector
    # tecnológico', 'Mercado asegurador'. La cabeza es el marco y el resto solo
    # nombra el rubro -> no dice QUÉ PASÓ. Vale para cualquier cliente.
    if re.match(r"^(?:sector(?:es)?|mercado(?:s)?|industria(?:s)?|gremio(?:s)?|"
                r"rubro|ramo|ambito|entorno|negocio(?:s)?)\b", n):
        resto = [t for t in n.split()[1:] if t not in _CONECTORES_ETIQUETA and len(t) >= 3]
        # 'Sector <rubro>' nunca dice QUÉ PASÓ: solo hay hecho si aparece un
        # sustantivo de acontecimiento (alza, convenio, sanción...). Los nombres de
        # rubro ('salud', 'tecnológico', 'asegurador') NO cuentan como hecho.
        if not any(t in _NUCLEOS_ACONTECIMIENTO for t in resto):
            return True
    # 'Marco + núcleo vacío': todos los tokens de contenido son marco/soporte o
    # meros calificadores de industria/ámbito ('Temas del sector financiero').
    # Se detecta por MORFOLOGÍA (sufijos de adjetivo relacional), no por una lista
    # de sectores: funciona igual para minero, portuario, avícola o farmacéutico.
    contenido = [t for t in n.split() if t not in _CONECTORES_ETIQUETA and len(t) >= 3]
    if contenido and all(
        (t in _NUCLEOS_RELLENO_SUBTEMA or _es_calificador_de_ambito(t))
        for t in contenido
    ):
        return True
    return False


_PATRON_LOCATIVO_FINAL = re.compile(
    r'\s+(en|de|del|para)\s+(colombia|bogota|barranquilla|medellin|cali|cartagena|'
    r'la guajira|guajira|santander|antioquia|valle|atlantico|bolivar|meta|cundinamarca|'
    r'norte de santander|sucre|magdalena|tolima|huila|casanare|arauca|putumayo|'
    r'el pais|el exterior|latinoamerica|america latina|la region|la ciudad|el departamento|'
    r'nivel nacional|territorio nacional)\s*$',
    re.IGNORECASE,
)


# Palabras de entidad/cantidad/verbo que no deben encabezar ni aportar 'asunto'
# en el subtema determinista (fallback). Se usan en _construir_frase_accion y
# _derivar_desde_titulos para mantener la etiqueta anclada al hecho, no al marco.
_TOKENS_DEBILES_SUBTEMA_FALLBACK = {
    "gobierno", "gobernacion", "gobernaciones", "alcaldia", "alcaldias", "ministerio",
    "ministerios", "secretaria", "secretarias", "entidad", "entidades", "autoridad",
    "autoridades", "empresa", "empresas", "compania", "companias", "corporacion",
    "corporaciones", "institucion", "instituciones", "fundacion", "fundaciones",
    "nacional", "regional", "local", "pais", "nacion", "se",
    "colombia", "bogota", "barranquilla", "medellin", "cali", "cartagena", "santa marta",
    "monteria", "sincelejo", "valledupar", "riohacha", "soledad", "malambo", "galapa",
    "la guajira", "guajira", "santander", "antioquia", "atlantico", "bolivar", "magdalena",
    "sucre", "cordoba", "cesar", "meta", "tolima", "huila", "narino", "cauca", "valle",
    "cundinamarca", "boyaca", "casanare", "arauca", "putumayo", "quindio", "risaralda",
    "caldas", "norte de santander",
    "millones", "millon", "miles", "billones", "millardos", "pesos", "dolares",
    "monto", "montos", "cifra", "cifras", "suma", "sumas",
    "anuncia", "anuncian", "anuncio", "anunciaron", "anunciado",
    "lanza", "lanzan", "lanzo", "lanzaron", "lanzado",
    "presenta", "presentan", "presento", "presentaron", "presentado",
    "inaugura", "inauguran", "inauguro", "inauguraron", "inaugurado",
    "celebra", "celebran", "celebro", "celebraron", "celebrado",
    "conmemora", "conmemoran", "conmemoro", "conmemoraron",
    "realiza", "realizan", "realizo", "realizaron", "realizado",
    "desarrolla", "desarrollan", "desarrollo", "desarrollaron", "desarrollado",
    "organiza", "organizan", "organizo", "organizaron", "organizado",
    "inicia", "inician", "inicio", "iniciaron", "iniciado",
    "abre", "abrieron", "abrio", "abierto",
    "firma", "firmaron", "firmado", "suscriben", "suscribieron", "suscribio", "suscrito",
    "invierte", "invirtieron", "invirtio", "invertido", "destinan", "destino", "destinaron",
    "entrega", "entregaron", "entregado",
    "recibe", "recibieron", "recibio", "recibido",
    "nombra", "nombraron", "nombrado", "nombrada", "nombramiento",
    "designa", "designaron", "designado", "designada",
    "posesiona", "posesionaron", "posesionado", "posesionada",
    "asume", "asumieron", "asumio", "asumido",
    "fue", "fueron", "sera", "seran", "es", "son", "era", "eran",
    "logra", "lograron", "logrado", "alcanza", "alcanzaron", "alcanzado",
    "supera", "superaron", "superado", "gana", "ganaron", "ganado",
    "otorga", "otorgaron", "otorgado", "premia", "premiaron", "premiado",
    "destaca", "destacaron", "destacado", "reconoce", "reconocieron", "reconocido",
    "elegida", "elegido", "elegidos", "elegidas",
}
def _quitar_locativos_finales(frase):
    if not frase:
        return frase
    s = " ".join(str(frase).split())
    while True:
        m = _PATRON_LOCATIVO_FINAL.search(s)
        if not m:
            return s
        s = s[:m.start()].strip()


# Sustantivos que terminan en -o/-an/-en y son núcleo válido al final de la
# etiqueta: sin esta lista blanca, la limpieza de verbos los cortaría.
_SUSTANTIVOS_SEGUROS_FINALES = {
    "pollo", "huevo", "precio", "costo", "consumo", "mercado", "gremio", "sector",
    "empleo", "servicio", "producto", "proyecto", "convenio", "acuerdo", "premio",
    "reconocimiento", "nombramiento", "lanzamiento", "financiamiento", "crecimiento",
    "aumento", "incremento", "descuento", "documento", "instrumento", "presupuesto",
    "impuesto", "contrato", "decreto", "informe", "balance", "estudio", "cambio",
    "riesgo", "trabajo", "salario", "ingreso", "egreso", "recurso", "proceso",
    "acceso", "negocio", "comercio", "turismo", "periodismo", "programa", "sistema",
    "problema", "esquema", "plan", "pan", "tren", "examen", "origen", "margen",
    "volumen", "resumen", "comun", "dictamen", "certamen", "orden", "aval", "grano",
    "terreno", "gobierno", "cuaderno", "invierno", "puerto", "aeropuerto", "distrito",
    "municipio", "territorio", "laboratorio", "inventario", "salario", "horario",
    "usuario", "beneficiario", "empresario", "ganado", "mercadeo", "bono", "fondo",
    "predio", "credito", "deficit", "vehiculo", "articulo", "capitulo", "titulo",
    "kilo", "kilos", "litro", "litros", "peso", "pesos", "dato", "datos",
    # ── Recursos y materias (cualquier cliente extractivo/agro) ────────────────
    # Sustantivos que terminan en vocal y NO deben tratarse como verbos ni
    # calificadores de rubro: petróleo, carbón, oro, cobre, acero, cemento, etc.
    "petroleo", "petroleos", "carbon", "carbones", "oro", "oras", "cobre", "cobres",
    "acero", "aceros", "cemento", "cementos", "niquel", "niqueles", "platino",
    "esmeralda", "esmeraldas", "sal", "sales", "gas", "gases", "crudo", "crudos",
    "gasolina", "gasolinas", "diesel", "granos", "cafe", "cafes", "cacao", "cacaos",
    "algodon", "arroz", "maiz", "trigo", "soya", "papa", "platanos", "banano",
    "flores", "carne", "carnes", "leche", "lacteos", "azucar", "panela", "tabaco",
    "madera", "maderas", "arena", "grava", "arcilla", "caliza", "fosfato", "ureas",
    "hidrogeno", "litio", "litios", "zinc", "plomo", "estano", "mercurio",
    "polvora", "explosivos", "fertilizantes", "agroquimicos", "semillas",
} | {  # áridos y minerales con acentos
    "petróleo", "carbón", "níquel", "plátano", "azúcar", "energía", "eléctrica",
}

def _es_forma_verbal_es(token) -> bool:
    """¿El token es una forma verbal conjugada? Detección MORFOLÓGICA, para no
    depender de una lista cerrada de verbos (que siempre deja huecos por sector).

    Reconoce futuros (-rá/-rán), pretéritos (-ó/-aron/-ieron), presentes de 3ª
    (-a/-an/-e/-en tras raíz verbal), gerundios (-ando/-iendo) y participios.
    Protege sustantivos con `_SUSTANTIVOS_SEGUROS_FINALES` y los núcleos de hecho.
    """
    w = _normaliza_token(str(token or ""))
    if not w or len(w) < 4:
        return False
    if w in _SUSTANTIVOS_SEGUROS_FINALES or w in _NUCLEOS_ACONTECIMIENTO:
        return False
    if _es_cabeza_subtema_valida(w):
        return False
    if w in _VERBOS_LEAD_SUBTEMA or _RE_VERBO_SUBTEMA.fullmatch(w):
        return True
    # Futuro / condicional: invertira, construira, pagaria
    if re.search(r"(?:ara|era|ira|aran|eran|iran|aria|erian|irian)$", w) and len(w) >= 7:
        return True
    # Pretérito perfecto simple 3ª persona: invirtio, afecto, presento, aumentaron
    if re.search(r"(?:aron|ieron|yeron)$", w) and len(w) >= 6:
        return True
    # Gerundio
    if re.search(r"(?:ando|iendo|yendo)$", w) and len(w) >= 6:
        return True
    # Presente 3ª persona de verbos en -ar/-er/-ir muy comunes en titulares:
    # 'afecta', 'genera', 'supera'. Solo si la raíz no es un sustantivo conocido.
    if re.search(r"(?:ecta|enera|upera|nvierte|eporta|resenta|nuncia|onfirma)$", w):
        return True
    return False


def _quitar_verbos_finales(frase):
    """Elimina verbos conjugados en cola: 'Precio del pollo subió' -> 'Precio del pollo'.
    Un subtema debe ser una FRASE NOMINAL; un verbo final la convierte en titular."""
    if not frase:
        return frase
    palabras = str(frase).split()
    while palabras:
        ultima = unidecode(palabras[-1].lower().rstrip(".,;:!?"))
        if ultima in _VERBOS_LEAD_SUBTEMA or _RE_VERBO_SUBTEMA.fullmatch(ultima):
            palabras.pop()
            continue
        break
    # No dejes preposición/artículo colgando tras quitar el verbo.
    while palabras and unidecode(palabras[-1].lower().rstrip(".,;:!?")) in _TRAILING_INCOMPLETE:
        palabras.pop()
    return " ".join(palabras)


# Verbos en pasado/participio frecuentes en titulares que no deben quedar DENTRO
# de la etiqueta ('Huevo aumentaron 8%'), más allá de los ya listados.
_RE_VERBO_INTERNO_ETIQUETA = re.compile(
    r"\b\w+(?:aron|ieron|aba|abia|aria|arian|ara|iera|ando|iendo|"
    r"ado|ada|ados|adas|ido|ida|idos|idas|o|an|en)\b", re.IGNORECASE)


def _sanear_frase_nominal(frase, max_palabras=None):
    """Saneo único de una etiqueta candidata para que sea FRASE NOMINAL limpia:
      - quita cifras/porcentajes y deícticos temporales ('8%', 'este año')
      - quita verbos conjugados internos y finales
      - quita locativos finales y preposiciones colgantes
      - elimina locuciones adverbiales sueltas ('per cápita') que no son asunto
    'Huevo aumentaron 8% este año' -> 'Exportaciones de huevo' (via el caller)."""
    if not frase:
        return ""
    s = " ".join(str(frase).split())
    # Locución 'per cápita' (y su resto 'capita'): modifica una cifra, no es asunto.
    s = re.sub(r"\bper\s+c[aá]pita\b", " ", s, flags=re.IGNORECASE)
    s = re.sub(r"(?<!\w)c[aá]pita(?!\w)", " ", s, flags=re.IGNORECASE)
    s = " ".join(s.split())
    # Cifras, porcentajes y montos: no describen la categoría.
    s = re.sub(r"\b\d[\d.,]*\s*(?:%|por ciento|mil|millones?|billones?|kilos?|toneladas?)?\b",
               " ", s, flags=re.IGNORECASE)
    # Símbolos/unidades que quedan huérfanos tras borrar la cifra ('%', 'kg', 'm²').
    s = re.sub(r"[%$€°]+", " ", s)
    s = re.sub(r"\b(?:por ciento|puntos porcentuales|kg|kilos?|toneladas?|"
               r"millones?|billones?|mil)\b", " ", s, flags=re.IGNORECASE)
    # Deícticos temporales.
    s = re.sub(r"\b(?:este|esta|estos|estas|el|la)\s+(?:ano|año|mes|semana|dia|día|"
               r"trimestre|semestre|periodo)\b", " ", s, flags=re.IGNORECASE)
    # Periodo suelto tras preposición ('de trimestre', 'en semestre'): no es asunto.
    s = re.sub(r"\b(?:de|del|en|al|para)\s+(?:ano|año|mes|semana|dia|día|trimestre|"
               r"semestre|periodo|bimestre|quincena)\b", " ", s, flags=re.IGNORECASE)
    s = re.sub(r"\b(?:hoy|ayer|manana|mañana|actualmente|recientemente)\b", " ", s,
               flags=re.IGNORECASE)
    s = " ".join(s.split())
    # Verbos conjugados internos (mantiene sustantivos como 'mercado', 'estado').
    palabras, limpias = s.split(), []
    for w in palabras:
        wn = unidecode(w.lower().rstrip(".,;:!?"))
        if wn in _VERBOS_LEAD_SUBTEMA or _RE_VERBO_SUBTEMA.fullmatch(wn) or _es_forma_verbal_es(wn):
            continue
        limpias.append(w)
    s = " ".join(limpias) if limpias else s
    # Preposiciones/artículos consecutivos tras quitar palabras ('a por', 'de en').
    toks, comp = s.split(), []
    for w in toks:
        wn = unidecode(w.lower().rstrip(".,;:!?"))
        if comp and wn in _CONECTORES_ETIQUETA:
            prev = unidecode(comp[-1].lower().rstrip(".,;:!?"))
            if prev in _CONECTORES_ETIQUETA:
                continue          # 'a por' -> 'por'
        comp.append(w)
    s = " ".join(comp)
    s = _recortar_frase_completa(s, max_palabras or MAX_PALABRAS_SUBTEMA)
    s = _quitar_locativos_finales(s)
    s = _quitar_verbos_finales(s)
    # Verbo conjugado en cola no listado ('llegó', 'alcanzó', 'creció'): si la
    # última palabra tiene forma verbal clara y la frase mantiene ≥2 palabras, se cae.
    toks = s.split()
    if len(toks) >= 3:
        ult = unidecode(toks[-1].lower().rstrip(".,;:!?"))
        if (len(ult) >= 4 and re.search(r"(?:o|aron|ieron|ara|era|ira|an|en)$", ult)
                and ult not in _SUSTANTIVOS_SEGUROS_FINALES
                and not _es_cabeza_subtema_valida(ult)):
            toks = toks[:-1]
            s = " ".join(toks)
    # Preposición inicial huérfana tras la limpieza.
    toks = s.split()
    while toks and unidecode(toks[0].lower()) in _CONECTORES_ETIQUETA:
        toks.pop(0)
    # Preposición final huérfana.
    while toks and unidecode(toks[-1].lower().rstrip(".,;:!?")) in _TRAILING_INCOMPLETE:
        toks.pop()
    # Cola 'preposición + término de marco' sin valor ('... a industria', '... del sector'):
    # el asunto ya está en la cabeza; la cola solo repite el ámbito.
    if len(toks) >= 4:
        cola = unidecode(" ".join(toks[-2:]).lower())
        if re.fullmatch(r"(?:a|al|de|del|en|para|con|sobre)\s+"
                        r"(?:la\s+|el\s+|los\s+|las\s+)?(?:" + _NUC_MARCO + r")", cola):
            toks = toks[:-2]
            while toks and unidecode(toks[-1].lower().rstrip(".,;:!?")) in _TRAILING_INCOMPLETE:
                toks.pop()
    return " ".join(toks).strip()


# ── Concordancia mínima para armar frases nominales idiomáticas ───────────────
_SUSTANTIVOS_MASC_SING_COMUNES = {
    "precio", "costo", "consumo", "mercado", "producto", "proyecto", "convenio",
    "acuerdo", "premio", "reconocimiento", "informe", "balance", "estudio",
    "pollo", "huevo", "sector", "gremio", "empleo", "servicio", "programa",
    "plan", "foro", "congreso", "evento", "lanzamiento", "nombramiento",
    "cierre", "aumento", "incremento", "descenso", "sistema", "modelo",
}
_TERMINACIONES_ADJETIVO = (
    "ico", "ica", "icos", "icas", "ivo", "iva", "ivos", "ivas",
    "oso", "osa", "osos", "osas", "able", "ible", "ables", "ibles",
    "al", "ales", "ante", "antes", "iente", "ientes", "ario", "aria",
    "arios", "arias", "ense", "enses", "il", "iles",
    # Gentilicios y relacionales: 'caleños', 'bogotanos', 'andina'
    "eno", "ena", "enos", "enas", "ano", "ana", "anos", "anas",
    "ino", "ina", "inos", "inas",
)
_NO_ADJETIVOS = {
    "digital", "ambiental", "nacional", "regional", "local", "social", "laboral",
    "hospital", "capital", "animal", "canal", "personal", "material", "total",
    "general", "central", "final", "legal", "fiscal", "sanitario", "avicola",
    "plan", "pollo", "huevo", "grano", "mercado", "estado", "gobierno",
    "programa", "sistema", "problema", "empresario", "empresarios",
    "usuario", "usuarios", "beneficiario", "beneficiarios", "diario",
    "ciudadano", "ciudadanos", "colombiano", "colombianos", "campesino",
    "campesinos", "vecino", "vecinos", "alumno", "alumnos", "terreno",
}


def _es_cabeza_subtema_valida(token) -> bool:
    """¿El token es un sustantivo de evento que puede encabezar el subtema?
    Compara por stem Y por prefijo, porque _stem_es no une siempre singular y
    plural ('exportaciones' -> 'export' vs 'exportacion' -> 'exportacion')."""
    w = _normaliza_token(str(token or ""))
    if not w or len(w) < 4:
        return False
    if w in _CABEZAS_SUBTEMA_VALIDAS:
        return True
    sw = _stem_es(w)
    for h in _CABEZAS_SUBTEMA_VALIDAS:
        if len(h) < 4:
            continue
        sh = _stem_es(h)
        if sw == sh or w == h:
            return True
        # prefijo común largo: exportacion/exportaciones, inversion/inversiones
        base = min(len(sw), len(sh))
        if base >= 5 and (sw.startswith(sh[:base]) or sh.startswith(sw[:base])):
            return True
    return False


def _parece_adjetivo_es(palabra) -> bool:
    """Heurística: ¿la palabra es un adjetivo? Evita 'Explotación de sexual'.
    Los adjetivos van pegados al sustantivo, no tras preposición."""
    w = _normaliza_token(str(palabra or ""))
    if not w or len(w) < 4:
        return False
    if w in _NO_ADJETIVOS or w in _SUSTANTIVOS_MASC_SING_COMUNES:
        return False
    if _es_cabeza_subtema_valida(w):
        return False        # es un sustantivo de evento conocido
    # Adjetivos de contexto clínico/periodístico que no caen en sufijos estándar.
    if w in {"compleja", "complejas", "complejo", "complejos", "alta", "altas",
             "alto", "altos", "corta", "cortas", "corto", "cortos", "larga", "largas",
             "largo", "largos", "especial", "especiales", "directa", "directas",
             "directo", "directos", "abierta", "abiertas", "abierto", "abiertos",
             "extranjera", "extranjeras", "extranjero", "extranjeros", "local",
             "locales", "rural", "rurales", "urbana", "urbanas", "urbano", "urbanos"}:
        return True
    return w.endswith(_TERMINACIONES_ADJETIVO)


def _preposicion_de(palabra) -> str:
    """'de' + 'el' -> 'del' cuando el complemento es masculino singular común.
    Da 'Alza del precio' en lugar del más pobre 'Alza de precio'."""
    w = _normaliza_token(str(palabra or ""))
    if w in _SUSTANTIVOS_MASC_SING_COMUNES:
        return "del"
    return "de"


_ACCIONES_OPUESTAS = [
    ({"aprobacion", "aprueba", "apoyo", "acuerdo", "aval", "respaldo"}, {"rechazo", "rechaza", "desacuerdo", "oposicion", "critica"}),
    ({"aumento", "crecimiento", "alza", "subida", "incremento"}, {"caida", "reduccion", "baja", "disminucion", "descenso"}),
    ({"apertura", "inauguracion", "inicio", "lanzamiento", "estreno"}, {"cierre", "suspension", "fin", "clausura", "cancelacion"}),
    ({"exito", "logro", "triunfo", "premio", "reconocimiento"}, {"fracaso", "derrota", "problema", "crisis", "sancion"}),
    ({"demanda", "denuncia", "investigacion", "sancion", "multa"}, {"absolucion", "archivo", "exoneracion", "acuerdo"}),
]

_TOKENS_DEBILES_AGRUPACION = STOPWORDS_ES | {
    "noticia", "noticias", "informe", "informacion", "comunicado", "anuncio",
    "colombia", "pais", "nacional", "regional", "local", "sector", "sectores",
    "empresa", "empresas", "entidad", "entidades", "autoridad", "autoridades",
    "gobierno", "alcaldia", "gobernacion", "ministerio", "nuevo", "nueva",
    "nuevos", "nuevas", "plan", "programa", "proyecto", "iniciativa",
    "actividad", "actividades", "gestion", "tema", "caso", "casos",
}

def _tokens_distintivos(texto: str, min_len: int = 4) -> set:
    norm = string_norm_label(texto)
    return {
        t for t in norm.split()
        if len(t) >= min_len and t not in _TOKENS_DEBILES_AGRUPACION and not t.isdigit()
    }

def _overlap_distintivo(a: str, b: str) -> float:
    ta, tb = _tokens_distintivos(a), _tokens_distintivos(b)
    if not ta or not tb: return 0.0
    return len(ta & tb) / max(1, min(len(ta), len(tb)))

def _hay_conflicto_accion(a: str, b: str) -> bool:
    ta, tb = _tokens_distintivos(a, min_len=3), _tokens_distintivos(b, min_len=3)
    for grupo_a, grupo_b in _ACCIONES_OPUESTAS:
        if (ta & grupo_a and tb & grupo_b) or (ta & grupo_b and tb & grupo_a):
            return True
    return False

def _etiquetas_compatibles(a: str, b: str, min_overlap: float = 0.45) -> bool:
    na, nb = string_norm_label(a), string_norm_label(b)
    if not na or not nb: return False
    if _hay_conflicto_accion(na, nb): return False
    if SequenceMatcher(None, na, nb).ratio() >= 0.90: return True
    return _overlap_distintivo(na, nb) >= min_overlap

def _grupos_contenido_compatibles(
    textos_a: list,
    textos_b: list,
    etiqueta_a: str = "",
    etiqueta_b: str = "",
    min_sim: float = 0.88,
    min_overlap: float = 0.20,
) -> bool:
    muestra_a = [str(t) for t in textos_a[:20] if str(t).strip()]
    muestra_b = [str(t) for t in textos_b[:20] if str(t).strip()]
    if not muestra_a or not muestra_b: return False
    texto_a = " ".join(muestra_a)[:2500]
    texto_b = " ".join(muestra_b)[:2500]
    if _hay_conflicto_accion(f"{etiqueta_a} {texto_a}", f"{etiqueta_b} {texto_b}"):
        return False
    overlap = _overlap_distintivo(f"{etiqueta_a} {texto_a}", f"{etiqueta_b} {texto_b}")
    labels_muy_cercanas = _etiquetas_compatibles(etiqueta_a, etiqueta_b, min_overlap=0.55)
    if overlap < min_overlap and not labels_muy_cercanas:
        return False
    embs = get_embeddings_batch([texto_a, texto_b])
    if len(embs) < 2 or embs[0] is None or embs[1] is None:
        return labels_muy_cercanas and overlap >= min_overlap
    sim = cosine_similarity(
        np.array(embs[0]).reshape(1, -1),
        np.array(embs[1]).reshape(1, -1)
    )[0][0]
    return sim >= min_sim

def _validar_estructura_subtema(etiqueta: str) -> bool:
    if not etiqueta or len(etiqueta.split()) < 2: return False
    if len(etiqueta.split()) > MAX_PALABRAS_SUBTEMA: return False
    if _PATRON_TITULAR.match(etiqueta): return False
    if _PATRON_ESTADO.search(etiqueta): return False
    # Un subtema es una FRASE NOMINAL, no un titular copiado. Los marcadores de
    # encabezado periodístico (dos puntos, guiones largos, barra, punto y coma,
    # envoltura entre comillas) delatan un titular literal -> se rechazan.
    if re.search(r'[:：—–|;»\u201d\uff1a]', etiqueta):
        return False
    palabras = etiqueta.split()
    if len(palabras) == 2:
        # 'Sustantivo + adjetivo' ('Cirugía compleja', 'Cobertura regional') es un
        # subtema válido SIN preposición; 'keyword + keyword' ('precio pollo') no.
        if _parece_adjetivo_es(palabras[1]):
            return True
    if len(palabras) <= 4:
        nexos = {
            "de","del","para","sobre","en","con","por","ante","hacia",
            "entre","sin","al","las","los","una","uno","que","como",
            "y","o","a","e","u",
        }
        tiene_nexo = any(unidecode(p.lower().rstrip(".,;:!?")) in nexos for p in palabras[1:])
        if not tiene_nexo: return False
    return True

def _es_nombre_o_fragmento_marca(etiqueta: str, marca: str, aliases=None) -> bool:
    """Detecta etiquetas que solo repiten total o parcialmente el nombre de la marca."""
    vacias = {"de", "del", "la", "el", "los", "las", "y", "e", "grupo"}
    tokens_etiqueta = {t for t in _normalizar_mencion(etiqueta).split() if t not in vacias}
    if not tokens_etiqueta:
        return True
    for nombre in _variantes_marca(marca, aliases):
        tokens_marca = {t for t in _normalizar_mencion(nombre).split() if t not in vacias}
        if not tokens_marca:
            continue
        comunes = tokens_etiqueta & tokens_marca
        # Rechaza cualquier etiqueta compuesta casi exclusivamente por tokens de la marca.
        if len(comunes) >= 2 and len(comunes) / len(tokens_etiqueta) >= 0.70:
            return True
    return False

def _es_verboso_con_marca(etiqueta, marca, aliases=None):
    """Detecta etiquetas vacías tipo 'Investigación sobre universidad simón bolívar':
    un verbo-muletilla (investigación/estudio/análisis/informe) + 'sobre/de' + SOLO la marca.
    Eso no describe un hecho; se debe regenerar con un asunto real."""
    s = (etiqueta or "").strip()
    m = re.match(
        r"^(investigaci[oó]n|estudio|an[aá]lisis|informe|trabajo)\s+"
        r"(sobre|acerca de|de|del|sobre la|sobre el|de la|de el)\s+(.+)$",
        s, re.IGNORECASE,
    )
    if not m:
        return False
    resto = m.group(3)
    return bool(resto.strip()) and _es_nombre_o_fragmento_marca(resto, marca, aliases)

def extract_link(cell):
    if hasattr(cell, "hyperlink") and cell.hyperlink:
        return {"value": "Link", "url": cell.hyperlink.target}
    if isinstance(cell.value, str) and "=HYPERLINK" in cell.value:
        m = re.search(r'=HYPERLINK\("([^"]+)"', cell.value)
        if m: return {"value": "Link", "url": m.group(1)}
    return {"value": cell.value, "url": None}

def extract_link_from_cell(cell):
    if cell.hyperlink and cell.hyperlink.target:
        return cell.hyperlink.target
    return None

def convert_html_entities(text):
    if not isinstance(text, str):
        return text
    text = html.unescape(text)
    html_entities = {
        '&#xF3;': 'ó', '&#xE1;': 'á', '&#xE9;': 'é', '&#xED;': 'í',
        '&#xFA;': 'ú', '&#xF1;': 'ñ', '&#xDC;': 'Ü', '&#xFC;': 'ü',
        '&#xC1;': 'Á', '&#xC9;': 'É', '&#xCD;': 'Í', '&#xD3;': 'Ó',
        '&#xDA;': 'Ú', '&#xD1;': 'Ñ', '&#xC7;': 'Ç', '&#xE7;': 'ç',
    }
    for entity, char in html_entities.items():
        text = text.replace(entity, char)

    def replace_hex_entity(match):
        try:
            return chr(int(match.group(1), 16))
        except Exception:
            return match.group(0)

    def replace_decimal_entity(match):
        try:
            return chr(int(match.group(1)))
        except Exception:
            return match.group(0)

    text = re.sub(r'&#x([0-9A-Fa-f]+);', replace_hex_entity, text)
    text = re.sub(r'&#(\d+);', replace_decimal_entity, text)

    for bad, good in {'\u201c': '"', '\u201d': '"', '\u2018': "'", '\u2019': "'",
                      'Â': '', 'â': '', '€': '', '™': ''}.items():
        text = text.replace(bad, good)
    return text

def clean_text(text):
    if not isinstance(text, str):
        return text
    return convert_html_entities(text).strip()

def clean_cuerpo(text):
    if not isinstance(text, str) or text.strip() == '':
        return text
    text = convert_html_entities(text)
    text = re.sub(r'<br\s*/?>', '\n', text, flags=re.IGNORECASE)
    text = re.sub(r'<[^>]+>', '', text)
    return text.strip()


# ======================================
# FUNCIÓN DE NORMALIZACIÓN DE TÍTULOS (MEJORADA)
# ======================================
def normalize_title_for_comparison(title):
    if not isinstance(title, str): 
        return ""
    
    cleaned = re.sub(r"\s+[\|–—-]\s+[^\|–—-]+$", "", title).strip()
    
    if ":" in cleaned:
        parts = cleaned.split(":", 1)
        suffix = parts[1].strip()
        if len(suffix) >= 10:
            cleaned = suffix
    cleaned = unidecode(cleaned)
    return re.sub(r"\W+", " ", cleaned).lower().strip()


def clean_title_for_output(title):
    return re.sub(r"\s*\|\s*[\w\s]+$", "", str(title)).strip()

def corregir_texto(text):
    if not isinstance(text, str): return text
    text = re.sub(r"(<br>|\[\.\.\.\]|\s+)", " ", text).strip()
    m = re.search(r"[A-ZÁÉÍÓÚÑ]", text)
    if m: text = text[m.start():]
    if text and not text.endswith("..."): text = text.rstrip(".") + "..."
    return text

def normalizar_tipo_medio(tipo_raw):
    if not isinstance(tipo_raw, str): return str(tipo_raw)
    t = unidecode(tipo_raw.strip().lower())
    return {
        'online': 'Internet', 'internet': 'Internet',
        'diario': 'Prensa',
        'am': 'Radio', 'fm': 'Radio', 'radio': 'Radio',
        'aire': 'Televisión', 'cable': 'Televisión', 'tv': 'Televisión',
        'television': 'Televisión', 'televisión': 'Televisión',
        'revista': 'Revistas', 'revistas': 'Revistas',
    }.get(t, str(tipo_raw).strip().title() or "Otro")

def parse_numeric(val):
    if val is None:
        return None
    if isinstance(val, (int, float)):
        if isinstance(val, float) and val.is_integer():
            return int(val)
        return val
    s = str(val).strip()
    if not s:
        return None
    if 'e' in s.lower():
        s = s.replace(',', '.')
    else:
        if ',' in s and '.' in s:
            if s.rfind('.') < s.rfind(','):
                s = s.replace('.', '').replace(',', '.')
            else:
                s = s.replace(',', '')
        elif ',' in s:
            parts = s.split(',')
            if len(parts) > 2 or (len(parts) == 2 and len(parts[1]) == 3 and not s.lower().startswith('0,')):
                s = s.replace(',', '')
            else:
                s = s.replace(',', '.')
        elif '.' in s:
            parts = s.split('.')
            if len(parts) > 2 or (len(parts) == 2 and len(parts[1]) == 3 and not s.lower().startswith('0.')):
                s = s.replace('.', '')
    try:
        f_val = float(s)
        if f_val.is_integer():
            return int(f_val)
        return f_val
    except ValueError:
        return None

def texto_para_embedding(titulo, resumen, max_len=1800):
    t = str(titulo or "").strip()
    r = str(resumen or "").strip()
    return f"{t}. {t}. {t}. {r}"[:max_len]

def _normalizar_mencion(texto: str) -> str:
    """Minúsculas, sin tildes, sin puntuación. utb == UTB, tecnologica == tecnológica."""
    return re.sub(r"\s+", " ", re.sub(r"[^a-z0-9]+", " ", unidecode(str(texto).lower()))).strip()

def _acronimo_de_nombre(nombre: str) -> str:
    vacias = {"de", "del", "la", "el", "los", "las", "y", "e", "da", "do", "di", "grupo"}
    toks = [t for t in _normalizar_mencion(nombre).split() if t not in vacias]
    if len(toks) < 2:
        return ""
    ac = "".join(t[0] for t in toks)
    return ac if 2 <= len(ac) <= 6 else ""

def _lista_alias(marca, aliases=None):
    nombres = []
    if marca:
        nombres.extend(str(marca).split(";"))
    if isinstance(aliases, str):
        nombres.extend(aliases.split(";"))
    else:
        nombres.extend(str(a) for a in (aliases or []))
    vistos, out = set(), []
    for n in nombres:
        k = _normalizar_mencion(n)
        if k and k not in vistos:
            vistos.add(k)
            out.append(n.strip())
    return out

def _variantes_marca(marca, aliases=None):
    """Formas digitadas + acrónimos (Universidad Tecnológica de Bolívar → utb)."""
    base = _lista_alias(marca, aliases)
    extra = []
    for n in base:
        ac = _acronimo_de_nombre(n)
        if ac:
            extra.append(ac)
    return _lista_alias(";".join(base + extra), None)

def _coincide_nombre_completo(texto: str, nombre: str) -> bool:
    nombre = _normalizar_mencion(nombre)
    if len(nombre) < 2:
        return False
    return bool(re.search(rf"(?<![a-z0-9]){re.escape(nombre)}(?![a-z0-9])", texto))

def _menciona_marca_o_alias(texto: str, marca: str, aliases=None) -> bool:
    normalizado = _normalizar_mencion(texto)
    nombres = _variantes_marca(marca, aliases)
    if any(_coincide_nombre_completo(normalizado, nombre) for nombre in nombres if str(nombre).strip()):
        return True
    tokens_texto = set(normalizado.split())
    vacias = {"de", "del", "la", "el", "los", "las", "y", "grupo"}
    for nombre in nombres:
        tokens_nombre = [t for t in _normalizar_mencion(nombre).split() if len(t) >= 3 and t not in vacias]
        if not tokens_nombre:
            continue
        coincidencias = len(set(tokens_nombre) & tokens_texto)
        if coincidencias >= min(2, len(set(tokens_nombre))) and coincidencias / len(set(tokens_nombre)) >= 0.60:
            return True
        if len(tokens_nombre) == 1:
            token = tokens_nombre[0]
            if len(token) >= 6 and any(
                len(candidate) >= 6 and SequenceMatcher(None, token, candidate).ratio() >= 0.88
                for candidate in tokens_texto
            ):
                return True
        else:
            # Compare each brand/alias token against nearby text tokens. This accepts
            # small spelling differences while still requiring most of the name.
            fuzzy_hits = 0
            for token in set(tokens_nombre):
                if any(
                    candidate == token or (
                        len(token) >= 5 and len(candidate) >= 5
                        and SequenceMatcher(None, token, candidate).ratio() >= 0.86
                    )
                    for candidate in tokens_texto
                ):
                    fuzzy_hits += 1
            required = max(1, int(np.ceil(len(set(tokens_nombre)) * 0.60)))
            if fuzzy_hits >= required and (fuzzy_hits >= 2 or len(set(tokens_nombre)) == 1):
                return True
    return False

def _default_text_column_index(columns, preferred_names, fallback=0):
    """Find common title/summary column spellings without accents or case sensitivity."""
    normalized = [_normalizar_mencion(str(c)).replace("-", " ") for c in columns]
    preferred = [_normalizar_mencion(x).replace("-", " ") for x in preferred_names]
    for wanted in preferred:
        for i, current in enumerate(normalized):
            if current == wanted:
                return i
    for wanted in preferred:
        for i, current in enumerate(normalized):
            if wanted in current or current in wanted:
                return i
    return min(fallback, max(0, len(columns) - 1))

def _safe_filename_part(value):
    cleaned = re.sub(r'[^A-Za-z0-9_-]+', '_', unidecode(str(value or '')).strip())
    return cleaned.strip('_') or 'marca'

_TERMINAL_PUNCT = re.compile(r'[.!?…]')
# Fin de oración REAL: excluye decimales (1.6), miles (3.200), siglas (S.A.S.),
# abreviaturas comunes y URLs. Antes '1.6 litros' cortaba el contexto en '1.'.
_RE_FIN_ORACION_REAL = re.compile(
    r'(?<![A-Z])'                     # no tras inicial de sigla: 'J. Pérez'
    r'(?<!\b[Ss]r)(?<!\b[Ss]ra)(?<!\b[Dd]r)(?<!\b[Dd]ra)'
    r'(?<!\bEE\.\sUU)(?<!\bpág)(?<!\bNo)(?<!\bnúm)(?<!\betc)'
    r'[.!?…]+'
    r'(?!\d)'                         # no seguido de dígito: 1.6 / 3.200
    r'(?=\s+[«"“(¿¡A-ZÁÉÍÓÚÑ0-9]|\s*$)'   # arranca oración nueva o fin de texto
)
MIN_PALABRAS_CONTEXTO = 10
# Contexto amplio: más caracteres alrededor de la mención de la marca/alias para un análisis
# más preciso de tono y subtema (el usuario pidió ampliarlo). ~600 carácter. mínimos garantizados,
# tope de ~2600.
CONTEXTO_MIN_CHARS = 600
MAX_CONTEXTO_CHARS = 2600

def _texto_hasta_terminal(texto, n=1):
    """Recorta 'texto' para que termine justo tras el enésimo signo de cierre (punto)."""
    if not texto:
        return ""
    for m in _RE_FIN_ORACION_REAL.finditer(texto):
        n -= 1
        if n == 0:
            return texto[:m.end()].strip(" \n\t")
    return texto.strip(" \n\t")

def _contexto_para_excel(contexto, max_chars=1400, umbral_corto=127):
    """Paso final de limpieza del 'Contexto analizado' para el Excel: toma hasta 3 oraciones
    (más contexto de la marca) con tope ~max_chars, ampliado tras el reciente aumento de caracteres."""
    if not contexto:
        return ""
    texto = str(contexto).strip()
    tramo = _texto_hasta_terminal(texto, n=2)
    if 0 < len(tramo) < umbral_corto:
        ampliada = _texto_hasta_terminal(texto, n=3)
        if len(ampliada) > len(tramo):
            tramo = ampliada
    if len(tramo) > max_chars:
        tramo = tramo[:max_chars].rstrip()
    return tramo.strip()

def _construir_texto_basico(row, tc, sc, bn, al):
    """Texto base de análisis: prioridad Título → Contexto analizado → Resumen-Aclaración,
    con el CONTEXTO AMPLIO (párrafo completo de la marca, hasta ~2200 car)."""
    titulo = str(row.get(tc, "") or "").strip()
    resumen = str(row.get(sc, "") or "").strip()[:300]
    ctx = extraer_contexto_marca(row.get(tc, ""), row.get(sc, ""), bn, al, row.get("Cuerpo Completo"))
    if ctx:
        ctx = ctx[:2200]
    return f"{titulo}. {titulo}. {ctx}. {resumen}".strip(" .")

def _extraer_parrafo_marca(fuente, marca, aliases):
    """Contexto amplio de la marca: parte del párrafo que la menciona y añade los párrafos
    siguientes hasta cubrir CONTEXTO_MIN_CHARS (≈600 car.), con tope MAX_CONTEXTO_CHARS.
    Así el análisis del tono/subtema de la marca cuenta con MÁS caracteres (pedido del usuario)."""
    parrafos = [p.strip() for p in re.split(r'\n+', fuente) if p and p.strip()]
    if not parrafos:
        parrafos = [fuente.strip()]
    con_hits = [i for i, p in enumerate(parrafos) if _menciona_marca_o_alias(p, marca, aliases)]
    if not con_hits:
        # Respaldo por ORACIÓN: el párrafo puede ser demasiado largo o venir sin
        # saltos, pero la mención existe en alguna oración. Se devuelve la oración
        # con la marca más su vecina, para no perder el fragmento a analizar.
        return _extraer_oracion_marca(fuente, marca, aliases)
    i = con_hits[0]  # primer párrafo que menciona la marca
    partes = [parrafos[i]]
    acum = len(parrafos[i])
    # Ampliar con el texto siguiente (hasta su 2º punto) mientras no se cubra el mínimo.
    for j in range(i + 1, len(parrafos)):
        if acum >= CONTEXTO_MIN_CHARS:
            break
        siguientes = parrafos[j]
        tramo1 = _texto_hasta_terminal(siguientes, n=1)
        if 0 < len(tramo1) < 60:
            siguientes = _texto_hasta_terminal(siguientes, n=2)
        partes.append(siguientes)
        acum += len(siguientes)
    tramo = " ".join(p for p in partes if p).strip()
    return tramo[:MAX_CONTEXTO_CHARS]


_RE_FIN_ORACION = _RE_FIN_ORACION_REAL


def _extraer_oracion_marca(fuente, marca, aliases):
    """Fragmento centrado en la ORACIÓN que menciona la marca (+ la siguiente).
    Segundo nivel de extracción cuando la división por párrafos no encuentra la
    mención: garantiza que el tono se evalúe sobre el texto donde aparece el
    cliente, en lugar de devolver '' (que antes forzaba 'Neutro' sin análisis)."""
    txt = " ".join(str(fuente or "").split())
    if not txt:
        return ""
    oraciones = [o.strip() for o in _RE_FIN_ORACION_REAL.split(txt) if o and o.strip()]
    if not oraciones:
        oraciones = [txt]
    idx = next((k for k, o in enumerate(oraciones)
                if _menciona_marca_o_alias(o, marca, aliases)), None)
    if idx is None:
        return ""
    partes = [oraciones[idx]]
    acum = len(oraciones[idx])
    # Contexto inmediato: oración siguiente y, si sigue corto, la anterior.
    if idx + 1 < len(oraciones) and acum < CONTEXTO_MIN_CHARS:
        partes.append(oraciones[idx + 1])
        acum += len(oraciones[idx + 1])
    if idx > 0 and acum < CONTEXTO_MIN_CHARS:
        partes.insert(0, oraciones[idx - 1])
    return " ".join(partes).strip()[:MAX_CONTEXTO_CHARS]


def _brand_audit(titulo, resumen, marca, aliases, cuerpo=None):
    d = extraer_contexto_marca_detallado(titulo, resumen, marca, aliases, cuerpo)
    return d['contexto'], d['coincidencia'], d['origen']

def extraer_contexto_marca(titulo, resumen, marca, aliases=None, cuerpo=None, ventana=320):
    """Contexto analizado de la marca: párrafo completo desde su inicio hasta su
    punto final. Prioridad de fuentes: Cuerpo Completo → Resumen → Título. Si el
    tramo queda muy corto, extiende hasta el segundo punto del texto siguiente.

    El fragmento SIEMPRE se ancla en la mención del cliente. Si la marca aparece
    solo en el título, se devuelve el título más la primera oración del resumen
    (contexto mínimo para juzgar el impacto); si aparece solo en el resumen/cuerpo,
    se devuelve ese fragmento. Nunca devuelve '' cuando la marca está en el texto.
    """
    titulo   = clean_text(str(titulo or "")).strip()
    resumen  = clean_text(str(resumen or "")).strip()
    cuerpo   = clean_text(str(cuerpo or "")).strip()
    if not marca:
        return ""
    fuentes = []
    if cuerpo:   fuentes.append(cuerpo)
    if resumen:  fuentes.append(resumen)
    if titulo:   fuentes.append(titulo)
    fuente = ""
    for f in fuentes:
        if _menciona_marca_o_alias(f, marca, aliases):
            fuente = f
            break
    if not fuente:
        return ""
    ctx = _extraer_parrafo_marca(fuente, marca, aliases)
    # La marca solo está en el TÍTULO: el título por sí solo suele ser muy corto
    # para evaluar impacto; se añade la primera oración del resumen como apoyo.
    if fuente == titulo and resumen:
        apoyo = _texto_hasta_terminal(resumen, n=1)
        if apoyo and apoyo not in ctx:
            ctx = f"{ctx} {apoyo}".strip()
    if not ctx:
        # Última red: la marca se menciona partida entre título y resumen.
        combinado = f"{titulo}. {resumen}".strip(" .")
        ctx = _extraer_oracion_marca(combinado, marca, aliases)
    return ctx[:MAX_CONTEXTO_CHARS]

def extraer_contexto_marca_detallado(titulo, resumen, marca, aliases=None, cuerpo=None):
    """Return auditable brand match metadata for sentiment analysis."""
    titulo, resumen = str(titulo or "").strip(), str(resumen or "").strip()
    cuerpo = str(cuerpo or "").strip()
    nombres = _variantes_marca(marca, aliases)
    title_hit = _menciona_marca_o_alias(titulo, marca, aliases)
    summary_hit = _menciona_marca_o_alias(resumen, marca, aliases)
    body_hit = bool(cuerpo) and _menciona_marca_o_alias(cuerpo, marca, aliases)
    if not title_hit and not summary_hit and not body_hit:
        return {"contexto": "", "marca_encontrada": "No", "origen": "", "coincidencia": ""}
    origen = ", ".join(x for x, ok in (("Título", title_hit), ("Resumen", summary_hit), ("Cuerpo", body_hit)) if ok)
    source = f"{titulo}. {resumen}. {cuerpo}".strip(" .")
    source_norm = _normalizar_mencion(source)
    matched = next((n for n in nombres if _coincide_nombre_completo(source_norm, n)), marca)
    return {
        "contexto": extraer_contexto_marca(titulo, resumen, marca, aliases, cuerpo),
        "marca_encontrada": "Sí", "origen": origen, "coincidencia": matched,
    }

def _validar_etiqueta_completa(etiqueta, titulos_grp=None, resumenes_grp=None, marca="", aliases=None, fallback_fn=None):
    if (not etiqueta or etiqueta.strip().lower() in ("sin tema", "varios", "n/a")
            or _es_etiqueta_generica(etiqueta)):
        if fallback_fn: return fallback_fn(titulos_grp or [])
        return "Sin tema"
    # Un subtema real no es un titular copiado ni inventa su núcleo de hecho.
    if not _validar_estructura_subtema(etiqueta):
        if fallback_fn: return fallback_fn(titulos_grp or [])
        return "Sin tema"
    if _frase_esta_completa(etiqueta):
        # CABEZA ANCLADA: el núcleo del hecho debe estar respaldado en el texto.
        if fallback_fn and titulos_grp:
            fuentes = [str(t) for t in titulos_grp if str(t).strip()]
            if resumenes_grp:
                fuentes += [str(r) for r in resumenes_grp if str(r).strip()]
            if not _head_anclada(etiqueta, fuentes):
                return fallback_fn(titulos_grp)
        return etiqueta
    recortada = _recortar_frase_completa(etiqueta, max_palabras=MAX_PALABRAS_SUBTEMA)
    if _frase_esta_completa(recortada) and len(recortada.split()) >= 2 and not _es_etiqueta_generica(recortada):
        return capitalizar_etiqueta(recortada)
    if titulos_grp and len(titulos_grp) > 0:
        try:
            prompt = (
                f"La frase '{etiqueta}' está incompleta o es genérica. "
                f"Genera una frase temática COMPLETA en español de 3-5 palabras "
                f"con preposición (de/del/para/sobre/en):\n\n"
                + "\n".join(f"  · {t[:120]}" for t in titulos_grp[:4])
                + "\n\nREGLAS: frase nominal con preposición, terminar en sustantivo/adjetivo, "
                f"tildes y ñ correctas. La etiqueta debe explicar el hecho relacionado con '{marca}', "
                "no limitarse al nombre de la institución. Usa SOLO palabras que aparezcan en el "
                "texto y NO devuelvas rótulos genéricos ('Cobertura de información relevante', "
                "'Cobertura informativa general').\n"
                "CORRECTO: 'Proyecto de terminal de transportes', 'Operación del Canal del Dique'\n"
                "INCORRECTO: 'Terminal transportes', 'Operación canal'\n"
                'JSON: {"subtema":"..."}'
            )
            resp = call_with_retries(
                openai.ChatCompletion.create,
                model=OPENAI_MODEL_CLASIFICACION,
                messages=[{"role": "user", "content": prompt}],
                max_tokens=200,
                temperature=0.1,
                response_format={"type": "json_object"}
            )
            u = resp.get('usage', {}) if isinstance(resp, dict) else getattr(resp, 'usage', {})
            if u:
                st.session_state['tokens_input'] += (u.get('prompt_tokens') if isinstance(u, dict) else getattr(u, 'prompt_tokens', 0)) or 0
                st.session_state['tokens_output'] += (u.get('completion_tokens') if isinstance(u, dict) else getattr(u, 'completion_tokens', 0)) or 0
            raw = json.loads(resp.choices[0].message.content).get("subtema", "")
            if raw:
                cleaned = limpiar_tema(raw)
                if (_frase_esta_completa(cleaned) and len(cleaned.split()) >= 2
                        and not _es_etiqueta_generica(cleaned)):
                    return capitalizar_etiqueta(cleaned)
        except:
            pass
    if fallback_fn: return fallback_fn(titulos_grp or [])
    return capitalizar_etiqueta(recortada) if recortada and len(recortada.split()) >= 2 and not _es_etiqueta_generica(recortada) else "Sin tema"

def dedup_labels(etiquetas, umbral=UMBRAL_DEDUP_LABEL):
    unique = list(dict.fromkeys(etiquetas))
    if len(unique) <= 1:
        return etiquetas
    normed = [string_norm_label(u) for u in unique]
    n = len(unique)
    parent = list(range(n))

    def find(x):
        while parent[x] != x:
            parent[x] = parent[parent[x]]
            x = parent[x]
        return x

    def union(a, b):
        ra, rb = find(a), find(b)
        if ra != rb:
            parent[rb] = ra

    def _es_fusion_segura(s1, s2):
        return _etiquetas_compatibles(s1, s2, min_overlap=0.45)

    for i in range(n):
        if not normed[i]: continue
        for j in range(i + 1, n):
            if not normed[j] or find(i) == find(j): continue
            if SequenceMatcher(None, normed[i], normed[j]).ratio() >= max(umbral, 0.88):
                if _es_fusion_segura(normed[i], normed[j]):
                    union(i, j)
                    
    for i in range(n):
        if not normed[i]: continue
        tokens_i = set(normed[i].split())
        if len(tokens_i) < 2: continue
        for j in range(i + 1, n):
            if not normed[j] or find(i) == find(j): continue
            tokens_j = set(normed[j].split())
            if len(tokens_j) < 2: continue
            interseccion = tokens_i & tokens_j
            menor = min(len(tokens_i), len(tokens_j))
            if menor > 0 and len(interseccion) / menor >= 0.78:
                if _es_fusion_segura(normed[i], normed[j]):
                    union(i, j)
                    
    le = get_embeddings_batch(unique)
    vp = [(i, le[i]) for i in range(n) if le[i] is not None]
    if len(vp) >= 2:
        vi, vv = zip(*vp)
        sm = cosine_similarity(np.array(vv))
        for pi in range(len(vi)):
            for pj in range(pi + 1, len(vi)):
                if sm[pi][pj] >= max(umbral, 0.90):
                    if find(vi[pi]) != find(vi[pj]):
                        if _es_fusion_segura(normed[vi[pi]], normed[vi[pj]]):
                            union(vi[pi], vi[pj])

    freq = Counter(etiquetas)
    grupos = defaultdict(list)
    for i in range(n):
        grupos[find(i)].append(i)
    canon = {}
    for root, members in grupos.items():
        cands = [unique[m] for m in members]
        vc = [c for c in cands if c.lower() not in ("sin tema", "varios") and _frase_esta_completa(c)]
        va = [c for c in cands if c.lower() not in ("sin tema", "varios")]
        if vc:
            canon[root] = max(vc, key=lambda c: (freq[c], len(c)))
        elif va:
            best = max(va, key=lambda c: (freq[c], len(c)))
            r = _recortar_frase_completa(best)
            canon[root] = r if _frase_esta_completa(r) else best
        else:
            canon[root] = cands[0]
    lm = {unique[i]: canon[find(i)] for i in range(n)}
    return [capitalizar_etiqueta(lm.get(e, e)) for e in etiquetas]

def _fusionar_subtemas_semanticos(subtemas, textos_por_subtema, marca, aliases, umbral=UMBRAL_FUSION_SUBTEMAS):
    unique_subs = list(dict.fromkeys(subtemas))
    if len(unique_subs) <= 1: return subtemas
    repr_texts = []
    for sub in unique_subs:
        txts = textos_por_subtema.get(sub, [])
        palabras = []
        for t in txts[:20]:
            for w in string_norm_label(str(t)).split():
                if len(w) > 3: palabras.append(w)
        top_kw = " ".join(w for w, _ in Counter(palabras).most_common(10))
        repr_texts.append(f"{sub}. {sub}. {sub}. {top_kw}"[:600])
    emb_repr = get_embeddings_batch(repr_texts)
    valid = [(i, emb_repr[i]) for i in range(len(unique_subs)) if emb_repr[i] is not None]
    if len(valid) < 2: return subtemas
    v_idx, v_emb = zip(*valid)
    sim = cosine_similarity(np.array(v_emb))
    n = len(v_idx)
    parent = list(range(n))

    def find(x):
        while parent[x] != x:
            parent[x] = parent[parent[x]]
            x = parent[x]
        return x

    def union(a, b):
        ra, rb = find(a), find(b)
        if ra != rb: parent[rb] = ra

    for i in range(n):
        for j in range(i + 1, n):
            if find(i) == find(j): continue
            sub_i, sub_j = unique_subs[v_idx[i]], unique_subs[v_idx[j]]
            if sim[i][j] >= max(umbral, 0.88) and _grupos_contenido_compatibles(
                textos_por_subtema.get(sub_i, []),
                textos_por_subtema.get(sub_j, []),
                sub_i,
                sub_j,
                min_sim=max(umbral, 0.88),
                min_overlap=0.22,
            ):
                union(i, j)
            
    grupos = defaultdict(list)
    for i in range(n): grupos[find(i)].append(v_idx[i])
    freq = Counter(subtemas)
    lm = {}
    for root, members in grupos.items():
        cands = [unique_subs[m] for m in members]
        if len(cands) == 1:
            lm[cands[0]] = cands[0]
            continue
        vc = [c for c in cands if c.lower() not in ("sin tema", "varios") and _frase_esta_completa(c)]
        best = max(vc, key=lambda c: (freq.get(c, 0), len(c))) if vc else max(cands, key=lambda c: (freq.get(c, 0), len(c)))
        if len(cands) <= 3:
            unified = _unificar_subtemas_llm(cands, textos_por_subtema, marca, aliases)
            if unified and _frase_esta_completa(unified): best = unified
        for c in cands: lm[c] = capitalizar_etiqueta(best)
    return [lm.get(s, s) for s in subtemas]

def _unificar_subtemas_llm(subtemas_a_unificar, textos_por_subtema, marca, aliases):
    subs_str = "\n".join(f"  · {s}" for s in subtemas_a_unificar)
    all_kw = []
    for sub in subtemas_a_unificar:
        for t in textos_por_subtema.get(sub, [])[:5]:
            for w in string_norm_label(str(t)).split():
                if len(w) > 3: all_kw.append(w)
    kw_str = " · ".join(w for w, _ in Counter(all_kw).most_common(8))
    prompt = (
        f"Estos subtemas son variaciones del MISMO tema. "
        f"Genera UN subtema unificado (4-6 palabras) como frase nominal completa:\n\n"
        f"{subs_str}\n\nKeywords: {kw_str}\n\n"
        f"REGLAS: frase coherente vinculada con '{marca}', con preposición (de/del/para/sobre/en), "
        "tildes y ñ correctas, terminar en sustantivo/adjetivo y explicar el hecho, no solo la marca.\n"
        "CORRECTO: 'Regulación de tarifas eléctricas', 'Apertura de nuevas sucursales'\n"
        "INCORRECTO: 'Tarifas energía', 'Apertura sucursales', 'Actividad corporativa'\n"
        'JSON: {"subtema":"..."}'
    )
    try:
        resp = call_with_retries(
            openai.ChatCompletion.create,
            model=OPENAI_MODEL_CLASIFICACION,
            messages=[{"role": "user", "content": prompt}],
            max_tokens=160,
            temperature=0.05,
            response_format={"type": "json_object"}
        )
        u = resp.get('usage', {}) if isinstance(resp, dict) else getattr(resp, 'usage', {})
        if u:
            st.session_state['tokens_input'] += (u.get('prompt_tokens') if isinstance(u, dict) else getattr(u, 'prompt_tokens', 0)) or 0
            st.session_state['tokens_output'] += (u.get('completion_tokens') if isinstance(u, dict) else getattr(u, 'completion_tokens', 0)) or 0
        raw = json.loads(resp.choices[0].message.content).get("subtema", "")
        if raw: return limpiar_tema(raw)
    except:
        pass
    return None

def get_embeddings_batch(textos, batch_size=100):
    if not textos: return []
    cache = get_embedding_cache()
    resultados, missing = cache.get_many(textos)
    if not missing: return resultados
    mt = [textos[i][:2000] if textos[i] else "" for i in missing]
    for i in range(0, len(mt), batch_size):
        batch = mt[i:i + batch_size]
        bidx = missing[i:i + batch_size]
        try:
            resp = call_with_retries(openai.Embedding.create, input=batch, model=OPENAI_MODEL_EMBEDDING)
            u = resp.get('usage', {}) if isinstance(resp, dict) else getattr(resp, 'usage', {})
            if u:
                st.session_state['tokens_embedding'] += (u.get('total_tokens') if isinstance(u, dict) else getattr(u, 'total_tokens', 0)) or 0
            for j, d in enumerate(resp["data"]):
                oi = bidx[j]
                emb = d["embedding"]
                resultados[oi] = emb
                cache.put(textos[oi], emb)
        except:
            for j, t in enumerate(batch):
                oi = bidx[j]
                try:
                    r = openai.Embedding.create(input=[t], model=OPENAI_MODEL_EMBEDDING)
                    emb = r["data"][0]["embedding"]
                    resultados[oi] = emb
                    cache.put(textos[oi], emb)
                except:
                    pass
    cache.flush()
    return resultados

class DSU:
    def __init__(self, n):
        self.p = list(range(n))
        self.rank = [0] * n

    def find(self, i):
        path = []
        while self.p[i] != i:
            path.append(i)
            i = self.p[i]
        for node in path: self.p[node] = i
        return i

    def union(self, i, j):
        ri, rj = self.find(i), self.find(j)
        if ri == rj: return
        if self.rank[ri] < self.rank[rj]: ri, rj = rj, ri
        self.p[rj] = ri
        if self.rank[ri] == self.rank[rj]: self.rank[ri] += 1

    def grupos(self, n):
        c = defaultdict(list)
        for i in range(n): c[self.find(i)].append(i)
        return dict(c)

def agrupar_textos_similares(textos, umbral):
    if not textos: return {}
    embs = get_embeddings_batch(textos)
    valid = [(i, e) for i, e in enumerate(embs) if e is not None]
    if len(valid) < 2: return {}
    idxs, M = zip(*valid)
    labels = AgglomerativeClustering(
        n_clusters=None, distance_threshold=1 - umbral, metric="cosine", linkage="average"
    ).fit(np.array(M)).labels_
    g = defaultdict(list)
    for k, lbl in enumerate(labels): g[lbl].append(idxs[k])
    return dict(enumerate(g.values()))

def agrupar_por_titulo_similar(titulos):
    gid, grupos, used = 0, {}, set()
    norm = [normalize_title_for_comparison(t) for t in titulos]
    for i in range(len(norm)):
        if i in used or not norm[i]: continue
        grp = [i]
        used.add(i)
        for j in range(i + 1, len(norm)):
            if j in used or not norm[j]: continue
            if SequenceMatcher(None, norm[i], norm[j]).ratio() >= SIMILARITY_THRESHOLD_TITULOS:
                grp.append(j)
                used.add(j)
        if len(grp) >= 2:
            grupos[gid] = list(set(grp))
            gid += 1
    return grupos

def seleccionar_representante(indices, textos):
    embs = get_embeddings_batch([textos[i] for i in indices])
    validos = [(indices[k], e) for k, e in enumerate(embs) if e is not None]
    if not validos: return indices[0], textos[indices[0]]
    idxs, M = zip(*validos)
    centro = np.mean(M, axis=0, keepdims=True)
    best = int(np.argmax(cosine_similarity(np.array(M), centro)))
    return idxs[best], textos[idxs[best]]

def construir_grupos_consistentes(titulos, resumenes):
    """Agrupa republicaciones y notas equivalentes con criterios conservadores."""
    titulos = [str(x or "") for x in titulos]
    resumenes = [str(x or "") for x in resumenes]
    textos = [texto_para_embedding(t, r) for t, r in zip(titulos, resumenes)]
    n = len(textos)
    dsu = DSU(n)
    embs = get_embeddings_batch(textos)
    norm = [normalize_title_for_comparison(t) for t in titulos]

    # Bloqueo por palabras distintivas para evitar una comparación O(n²) completa.
    indice = defaultdict(set)
    for i, titulo in enumerate(norm):
        for token in _tokens_distintivos(titulo):
            indice[token].add(i)
    pares = set()
    for idxs in indice.values():
        if len(idxs) > 100:
            continue
        orden = sorted(idxs)
        pares.update((orden[a], orden[b]) for a in range(len(orden)) for b in range(a + 1, len(orden)))

    for i, j in pares:
        if _hay_conflicto_accion(textos[i], textos[j]):
            continue
        title_sim = SequenceMatcher(None, norm[i], norm[j]).ratio() if norm[i] and norm[j] else 0.0
        overlap = _overlap_distintivo(textos[i], textos[j])
        semantic = 0.0
        if embs[i] is not None and embs[j] is not None:
            semantic = cosine_similarity(
                np.array(embs[i]).reshape(1, -1), np.array(embs[j]).reshape(1, -1)
            )[0][0]
        if title_sim >= SIMILARITY_THRESHOLD_TITULOS or (semantic >= SIMILARITY_THRESHOLD_TONO and overlap >= 0.45):
            dsu.union(i, j)
    return dsu.grupos(n)

def construir_grafo_equivalencia(titulos, resumenes, contextos=None):
    # Grafo UNICO de 'noticias equivalentes' (misma historia), con criterios ESTRICTOS.
    # Reutilizado por tono, tema y subtema para que no se contradigan entre si.
    n = len(titulos)
    dsu = DSU(n)
    tn = [norm_key(str(t or "")) for t in titulos]
    rn = [norm_key(str(r or "")) for r in resumenes] if resumenes is not None else None
    cn = [norm_key(str(c or "")) for c in contextos] if contextos is not None else None
    # Tokens distintivos por título: exige asunto compartido antes de unir por
    # similitud de cadena, para no pegar noticias distintas de redacción parecida.
    tok = [_tokens_distintivos(str(t or "")) for t in titulos]
    for i in range(n):
        if not tn[i]:
            continue
        ti = tn[i]
        for j in range(i + 1, n):
            tj = tn[j]
            if not tj or dsu.find(i) == dsu.find(j):
                continue
            if _hay_conflicto_accion(str(titulos[i] or ""), str(titulos[j] or "")):
                continue          # aprobación vs rechazo: nunca son la misma noticia
            igual = (ti == tj)
            if not igual and len(ti) >= 10 and len(tj) >= 10:
                igual = (ti in tj or tj in ti)                     # mismo titular con/sin subtitulo
            ratio_t = SequenceMatcher(None, ti, tj).ratio()
            if not igual and ratio_t >= 0.88:
                igual = True
            # CUERPO IDÉNTICO y sustancial = republicación con titular reescrito.
            # Es la señal más fuerte de "misma noticia" y cubre el caso del usuario:
            # el mismo despacho publicado por varios medios con otro titular.
            # Se exige además asunto compartido (≥2 tokens distintivos o cobertura
            # ≥0.5) para no unir notas distintas que comparten un cuerpo boilerplate.
            if not igual and rn and rn[i] and rn[i] == rn[j] and len(rn[i]) >= 60:
                comp = len(tok[i] & tok[j])
                cob_t = comp / max(1, min(len(tok[i]), len(tok[j]))) if tok[i] and tok[j] else 0.0
                if comp >= 2 or cob_t >= 0.5:
                    igual = True
            if not igual and rn and rn[i] and rn[j] and ratio_t >= 0.80 and SequenceMatcher(None, rn[i], rn[j]).ratio() >= 0.70:
                igual = True                                       # titulo parecido Y cuerpo parecido
            # Reescritura de titular con el MISMO asunto: los títulos comparten casi
            # todos sus tokens distintivos y el cuerpo es muy parecido. Cubre el caso
            # del usuario: "noticias iguales o muy parecidas deben compartir etiquetas".
            if not igual and rn and rn[i] and rn[j] and tok[i] and tok[j]:
                inter = len(tok[i] & tok[j])
                jacc = inter / max(1, len(tok[i] | tok[j]))
                cobertura = inter / max(1, min(len(tok[i]), len(tok[j])))
                sim_cuerpo = SequenceMatcher(None, rn[i], rn[j]).ratio()
                if (jacc >= 0.60 or cobertura >= 0.80) and sim_cuerpo >= 0.75:
                    igual = True
            if not igual and cn and cn[i] and cn[i] == cn[j]:
                igual = True                                       # mismo Contexto analizado
            if igual:
                dsu.union(i, j)
    return dsu

def aplicar_consistencia_grupos(df, titulo_col, resumen_col,
                                tono_col="Tono IA", tema_col="Tema", subtema_col="Subtema"):
    # Asigna 'Grupo noticia' y unifica Tono IA / Tema / Subtema de las noticias equivalentes
    # usando UN UNICO grafo de equivalencia (mismo criterio para los tres).
    if df.empty:
        return df
    grupos = construir_grupos_consistentes(df[titulo_col].fillna(''), df[resumen_col].fillna(''))
    df = df.copy()
    df["Grupo noticia"] = ""
    for numero, idxs in enumerate(grupos.values(), start=1):
        gid = f"G{numero:05d}"
        for i in idxs:
            df.at[df.index[i], "Grupo noticia"] = gid

    if subtema_col in df.columns:
        df[subtema_col] = df[subtema_col].apply(
            lambda x: capitalizar_etiqueta(_recortar_frase_completa(str(x), MAX_PALABRAS_SUBTEMA))
            if str(x).strip().lower() not in {"", "nan", "n/a", "-"} else x
        )

    n = len(df)
    contextos = ([str(x) for x in df['Contexto analizado'].fillna('')]
                 if 'Contexto analizado' in df.columns else None)

    # Grafo de equivalencia estricto + union semantica por 'Grupo noticia'.
    dsu = construir_grafo_equivalencia(
        [str(x) for x in df[titulo_col].fillna('')],
        [str(x) for x in df[resumen_col].fillna('')],
        contextos,
    )
    por_grupo = defaultdict(list)
    for i in range(n):
        g = str(df.iloc[i].get('Grupo noticia') or "").strip()
        if g and g.lower() not in ("", "nan", "none"):
            por_grupo[g].append(i)
    for idxs in por_grupo.values():
        if len(idxs) < 2:
            continue
        base = idxs[0]
        for k in idxs[1:]:
            dsu.union(base, k)

    grupos_eq = defaultdict(list)
    for i in range(n):
        grupos_eq[dsu.find(i)].append(i)

    def _canon_mas_frecuente(idxs, col):
        vals = [str(df.iloc[i][col]).strip() for i in idxs]
        vals = [v for v in vals if v and v.lower() not in ("nan", "none", "-", "n/a", "")]
        if not vals:
            return None
        order = []
        for v in vals:
            if v not in order:
                order.append(v)
        freq = {v: vals.count(v) for v in order}
        return max(order, key=lambda v: (freq[v], -order.index(v)))

    def _canon_par_mas_frecuente(idxs):
        """Canoniza (Tema, Subtema) como PAR ATÓMICO.

        Antes se elegía el Tema más frecuente y el Subtema más frecuente por
        separado: si el grupo traía (A,x) y (B,y), el resultado podía ser (B,x),
        un par que NUNCA existió en los datos -> 'subtema que no corresponde al
        tema'. Ahora se vota el par completo, así el Tema siempre es el que
        realmente acompañaba a ese Subtema.
        """
        vacios = ("nan", "none", "-", "n/a", "")
        pares = []
        for i in idxs:
            sub = str(df.iloc[i][subtema_col]).strip() if subtema_col in df.columns else ""
            tem = str(df.iloc[i][tema_col]).strip() if tema_col in df.columns else ""
            if sub.lower() in vacios and tem.lower() in vacios:
                continue
            if _es_etiqueta_generica(sub) and sub.lower() not in vacios:
                continue          # no dejes que un genérico gane la votación
            pares.append((tem, sub))
        if not pares:
            return None
        orden = []
        for p in pares:
            if p not in orden:
                orden.append(p)
        freq = {p: pares.count(p) for p in orden}
        # Empate -> gana el par cuyo Tema y Subtema son ambos no vacíos y el más antiguo.
        def _score(p):
            completo = 1 if (p[0].strip().lower() not in vacios and p[1].strip().lower() not in vacios) else 0
            return (freq[p], completo, -orden.index(p))
        return max(orden, key=_score)

    for idxs in grupos_eq.values():
        if len(idxs) < 2:
            continue
        par = _canon_par_mas_frecuente(idxs)
        if par:
            tem_c, sub_c = par
            for i in idxs:
                if subtema_col in df.columns and sub_c.strip():
                    df.at[df.index[i], subtema_col] = capitalizar_etiqueta(sub_c)
                if tema_col in df.columns and tem_c.strip():
                    df.at[df.index[i], tema_col] = capitalizar_etiqueta(tem_c)
        # Tono: Positivo/Negativo 'gana' sobre Neutro; conflicto Pos+Neg no se toca.
        if tono_col in df.columns:
            tvals = [str(df.iloc[i][tono_col]).strip().title() for i in idxs]
            if "Positivo" in tvals and "Negativo" in tvals:
                continue
            canon_tono = "Positivo" if "Positivo" in tvals else ("Negativo" if "Negativo" in tvals else None)
            if canon_tono:
                for i in idxs:
                    cur = str(df.iloc[i][tono_col]).strip().title()
                    if cur in ("Neutro", "N/A", "", "Nan"):
                        df.at[df.index[i], tono_col] = canon_tono
    return df


# TONO (Sistema Reputacional por IA)
# ======================================

# ── Pre-despacho determinista de tono (solo casos inequívocos, alta precisión) ──
# Reduce llamadas al LLM y elimina varianza en hechos estructurales. Si no es concluyente,
# devuelve None y decide el LLM. Desactivar con env GRILL_TONO_DETERMINISTA=0.
_TONO_POS_SUJETO = re.compile(
    r"\b(?:es|fue|result[oó]|qued[oó]|queda|resulta|se alz[oó]|se consagr[oó])\b"
    r".{0,30}?\b(?:ganador|ganadora|campe[oó]n|premiad[oa]|galardonad[oa]|reconocid[oa]|"
    r"condecorad[oa]|finalista|n[uú]mero uno|primer (?:lugar|puesto))\b"
)
_TONO_POS_VERBO = re.compile(
    r"\b(?:gan[oó]|obtuv[oó]|alz[oó]|recibi[oó]|logr[oó]|consigui[oó])\b"
    r".{0,30}?\b(?:premio|galard[oó]n|reconocimiento|honoris causa|triunfo|victoria|"
    r"primer (?:lugar|puesto)|n[uú]mero uno|distinci[oó]n|condecoraci[oó]n)\b"
)
_TONO_NEG_PASIVO = re.compile(
    r"\b(?:fue|es|siendo|qued[oó]|result[oó])\s+(?:demandad[oa]|denunciad[oa]|multad[oa]|"
    r"sancionad[oa]|investigad[oa]|imputad[oa]|condenad[oa]|judicializad[oa])\b"
)
_TONO_NEG_CONTRA = re.compile(
    r"\b(?:demanda|denuncia|multa|sanci[oó]n|investigaci[oó]n|proceso|querella)\b"
    r".{0,40}?\b(?:contra|a)\s+"
)


def _tono_determinista(eval_txt: str, marca: str, aliases=None):
    tex = unidecode((eval_txt or "").lower())
    if not tex:
        return None
    nombres = [unidecode(n.lower()) for n in [marca] + [a for a in (aliases or []) if a] if n]
    spans = []
    for n in nombres:
        spans.extend(m.start() for m in re.finditer(re.escape(n), tex))
    if not spans:
        return None
    pos = min(spans)
    # Positivo / negativo-pasivo: la marca es sujeto, el verbo/hecho va justo después.
    ventana = tex[max(0, pos - 10): pos + 80]
    if _TONO_POS_SUJETO.search(ventana) or _TONO_POS_VERBO.search(ventana):
        return {"tono": "Positivo", "confianza": "Alta",
                "justificacion": "Regla determinista: la marca es premiada/reconocida/ganadora."}
    if _TONO_NEG_PASIVO.search(ventana):
        return {"tono": "Negativo", "confianza": "Alta",
                "justificacion": "Regla determinista: la marca es demandada/multada/sancionada/investigada."}
    # Negativo "contra la marca": el desencadenante (demanda/denuncia/...) puede ir algo antes.
    ventana_contra = tex[max(0, pos - 50): pos + 80]
    if _TONO_NEG_CONTRA.search(ventana_contra):
        return {"tono": "Negativo", "confianza": "Alta",
                "justificacion": "Regla determinista: demanda/denuncia/sanción contra la marca."}
    return None


class ClasificadorTono:
    def __init__(self, marca, aliases):
        nombres = _variantes_marca(marca, aliases)
        self.marca = nombres[0] if nombres else str(marca or "").strip()
        self.aliases = [n for n in nombres[1:] if n]
        self._all_names = [self.marca] + self.aliases

    def _menciona_marca(self, texto):
        return _menciona_marca_o_alias(texto, self.marca, self.aliases)

    async def _clasificar_llm(self, texto, sem, contexto_marca=""):
        async with sem:
            eval_txt = (contexto_marca or texto or "").strip()
            if not eval_txt or not self._menciona_marca(eval_txt):
                return {"tono": "Neutro"}

            # Pre-despacho determinista (ahorra LLM en hechos inequívocos y elimina varianza)
            if os.environ.get("GRILL_TONO_DETERMINISTA", "1") != "0":
                det = _tono_determinista(eval_txt, self.marca, self.aliases)
                if det:
                    det["evidencia"] = eval_txt[:2400]
                    return det

            aliases_str = f" (también conocida como: {', '.join(self.aliases)})" if self.aliases else ""
            prompt = (
                f"Eres un experto analista en Relaciones Públicas y Gestión de Reputación. "
                f"Evalúa el impacto reputacional DIRECTO sobre la marca '{self.marca}'{aliases_str}.\n\n"
                f"El tono GENERAL de la noticia NO importa. Si el artículo es neutro o habla de otro actor, "
                f"pero la mención a '{self.marca}' es favorable, el tono es Positivo. "
                f"Si '{self.marca}' gana, es premiada, finalista, reconocida o se alza como ganadora "
                f"(aunque aparezca junto a otras instituciones), el tono es Positivo. "
                f"Estar en una LISTA DE GANADORES no es Neutro. "
                f"Si el artículo es positivo o trágico a nivel país/sector, pero '{self.marca}' queda "
                f"criticada, demandada o cuestionada, el tono es Negativo.\n\n"
                f"TEXTO CENTRADO EN LA MARCA:\n{eval_txt[:2400]}\n\n"
                f"REGLAS DE CLASIFICACIÓN ESTRICTAS:\n"
                f"🔴 NEGATIVO: un hecho perjudica, cuestiona o expone directamente a '{self.marca}' "
                f"(demandas, multas, fraudes, fallas propias, quejas, investigaciones, pérdidas o retiro de productos).\n"
                f"🟢 POSITIVO: el hecho acredita directamente un logro, mejora o aporte verificable de '{self.marca}' "
                f"(premio, crecimiento, lanzamiento exitoso, inversión realizada, innovación, expansión o reconocimiento). "
                f"También es POSITIVO si '{self.marca}' expresa SOLIDARIDAD o respaldo, emite un comunicado de apoyo humanitario, "
                f"dona, ayuda a comunidades/afectados tras una crisis o se pronuncia apoyando a su sector. "
                f"Un comunicado o muestra de solidaridad de la marca NUNCA es Neutro: refuerza su imagen pública.\n"
                f"⚪ NEUTRO: La marca se menciona SIN impacto a su imagen. Ejemplos:\n"
                f"  - La noticia habla de una crisis del sector/país, pero la marca solo es mencionada informando o adaptándose.\n"
                f"  - Se menciona a la marca de paso, sin rol (no aplica si es ganadora, premiada o protagonista).\n"
                f"  - Una persona, autoridad, proveedor o tercero es quien recibe el efecto positivo o negativo.\n"
                f"  - Emite un comunicado puramente rutinario/informativo SIN solidaridad, ayuda ni logro (p. ej. cambio de horario, cierre de oficina).\n"
                f"  - Critica, denuncia o advierte sobre un problema de terceros o del sector; la crítica de la marca NO es una crítica contra la marca.\n\n"
                f"⚠️ ATENCIÓN: Ignora el tono del sector o de terceros. Evalúa ÚNICAMENTE cómo el hecho afecta "
                f"la reputación corporativa de '{self.marca}': mejora (Positivo), empeora (Negativo) o no cambia (Neutro). "
                f"Recordatorio: solidaridad, comunicados de respaldo y ayuda humanitaria de la marca = POSITIVO, nunca Neutro.\n\n"
                f'Responde ÚNICAMENTE con JSON: {{"tono":"Positivo|Negativo|Neutro", '
                f'"confianza":"Alta|Media|Baja", "justificacion":"explicación concreta de máximo 35 palabras"}}'
            )

            try:
                resp = await acall_with_retries(
                    openai.ChatCompletion.acreate,
                    model=OPENAI_MODEL_CLASIFICACION,
                    messages=[{"role": "user", "content": prompt}],
                    max_tokens=300,
                    temperature=0.0,
                    response_format={"type": "json_object"}
                )
                
                u = resp.get('usage', {}) if isinstance(resp, dict) else getattr(resp, 'usage', {})
                if u:
                    st.session_state['tokens_input'] += (u.get('prompt_tokens') if isinstance(u, dict) else getattr(u, 'prompt_tokens', 0)) or 0
                    st.session_state['tokens_output'] += (u.get('completion_tokens') if isinstance(u, dict) else getattr(u, 'completion_tokens', 0)) or 0
                
                resultado = json.loads(resp.choices[0].message.content)
                tono = str(resultado.get("tono", "Neutro")).strip().title()
                
                tono = tono if tono in ("Positivo", "Negativo", "Neutro") else "Neutro"
                confianza = str(resultado.get("confianza", "Media")).strip().title()
                if confianza not in ("Alta", "Media", "Baja"):
                    confianza = "Media"
                return {"tono": tono, "confianza": confianza,
                        "justificacion": str(resultado.get("justificacion", "")).strip()[:400],
                        "evidencia": eval_txt[:2400]}
            except Exception as e:
                return {"tono": "Neutro", "confianza": "Baja", "justificacion": "Error de clasificación", "evidencia": eval_txt[:2400]}

    async def procesar_lote_async(self, textos, pbar, resumenes, titulos, cuerpos=None):
        n = len(textos)
        txts = textos.tolist()
        pbar.progress(0.05, "Agrupando noticias para análisis de tono...")
        
        txts_emb = [texto_para_embedding(str(titulos.iloc[i]), str(resumenes.iloc[i])) for i in range(n)]
        dsu = DSU(n)
        
        embs = get_embeddings_batch(txts_emb)
        candidatos = agrupar_textos_similares(txts_emb, SIMILARITY_THRESHOLD_TONO)
        candidatos.update({len(candidatos) + k: v for k, v in agrupar_por_titulo_similar(titulos.tolist()).items()})
        for idxs in candidatos.values():
            for pos, i in enumerate(idxs):
                for j in idxs[pos + 1:]:
                    ti, tj = normalize_title_for_comparison(titulos.iloc[i]), normalize_title_for_comparison(titulos.iloc[j])
                    titulo_casi_igual = SequenceMatcher(None, ti, tj).ratio() >= 0.96
                    contenido_casi_igual = (
                        embs[i] is not None and embs[j] is not None
                        and cosine_similarity(np.array(embs[i]).reshape(1, -1), np.array(embs[j]).reshape(1, -1))[0][0] >= SIMILARITY_THRESHOLD_TONO
                        and _overlap_distintivo(txts_emb[i], txts_emb[j]) >= 0.45
                    )
                    if (titulo_casi_igual or contenido_casi_igual) and not _hay_conflicto_accion(txts_emb[i], txts_emb[j]):
                        dsu.union(i, j)
                
        grupos = dsu.grupos(n)
        contextos = [
            extraer_contexto_marca(
                str(titulos.iloc[i]), str(resumenes.iloc[i]), self.marca, self.aliases,
                cuerpos.iloc[i] if cuerpos is not None else None
            )
            for i in range(n)
        ]
        reps = {}
        for cid, idxs in grupos.items():
            con_marca = [i for i in idxs if contextos[i]]
            if con_marca:
                ri, _ = seleccionar_representante(con_marca, contextos)
                reps[cid] = (ri, contextos[ri])
            else:
                # Sin contexto extraído: NO se asume Neutro a ciegas. Si la marca
                # aparece en título/resumen, se analiza ese texto (título + 1ª
                # oración del resumen). Solo queda Neutro si no hay mención real.
                ri = idxs[0]
                respaldo = ""
                for i in idxs:
                    t_i = str(titulos.iloc[i] or "")
                    r_i = str(resumenes.iloc[i] or "")
                    if self._menciona_marca(t_i) or self._menciona_marca(r_i):
                        combinado = f"{t_i}. {r_i}".strip(" .")
                        respaldo = (_extraer_oracion_marca(combinado, self.marca, self.aliases)
                                    or combinado[:MAX_CONTEXTO_CHARS])
                        ri = i
                        break
                reps[cid] = (ri, respaldo)
        
        sem = asyncio.Semaphore(CONCURRENT_REQUESTS)
        cids = list(reps.keys())
        
        async def _clasificar_con_cid(cid):
            _idx, ctx = reps[cid]
            if not ctx:
                return cid, {"tono": "Neutro"}
            return cid, await self._clasificar_llm(txts[_idx], sem, contexto_marca=ctx)

        tasks = [_clasificar_con_cid(c) for c in cids]
        rpg = {}
        
        for i, f in enumerate(asyncio.as_completed(tasks)):
            cid, r = await f
            rpg[cid] = r
            pbar.progress(0.1 + 0.85 * (i + 1) / len(tasks), f"Evaluando Reputación {i + 1}/{len(tasks)}")
            
        final = [None] * n
        
        for cid, idxs in grupos.items():
            r = rpg.get(cid, {"tono": "Neutro"})
            for i in idxs: final[i] = r

        tonos = [f["tono"] if f else "Neutro" for f in final]
        tonos = _propagar_tono_equivalentes(tonos, titulos.tolist(), resumenes.tolist())
        final = [{"tono": t} for t in tonos]
            
        pbar.progress(1.0, "Análisis de Tono completado")
        return final

def _propagar_tono_equivalentes(tonos, titulos, resumenes, contextos=None):
    # Noticias equivalentes (mismo grafo de equivalencia que tema/subtema):
    # si una es Positivo/Negativo y otra Neutro, se alinean. No propaga conflictos Pos+Neg.
    n = len(tonos)
    if n < 2:
        return list(tonos)
    dsu = construir_grafo_equivalencia(titulos, resumenes, contextos)
    out = list(tonos)
    for idxs in dsu.grupos(n).values():
        if len(idxs) < 2:
            continue
        vals = [out[i] for i in idxs]
        if "Positivo" in vals and "Negativo" in vals:
            continue
        if "Positivo" in vals:
            canon = "Positivo"
        elif "Negativo" in vals:
            canon = "Negativo"
        else:
            continue
        for i in idxs:
            if out[i] in ("Neutro", "N/A", "", "Nan"):
                out[i] = canon
    return out


def _predict_pkl_in_batches(pipeline, textos, progress=None, batch_size=64):
    """Run sklearn PKL inference in bounded batches so Streamlit can show progress."""
    values = list(textos)
    if not values:
        return []
    predictions = []
    total = len(values)
    for start in range(0, total, batch_size):
        end = min(start + batch_size, total)
        predictions.extend(pipeline.predict(values[start:end]))
        if progress is not None:
            progress.progress(end / total, f"Prediciendo PKL {end}/{total}")
    return predictions


def _snippet_tono_pkl(titulo, contexto):
    # Shape canonico del texto que recibe el PKL de tono (centrado en la marca).
    # IMPORTANTE: entrena el PKL con este MISMO formato para maxima consistencia.
    tit = str(titulo or "").strip()
    ctx = str(contexto or "").strip()
    if ctx and tit and ctx.lower().startswith(tit[:20].lower()):
        return ctx
    if ctx and tit:
        return f"{tit}. {ctx}"
    return ctx or tit


def analizar_tono_con_pkl(textos, pkl_file, titulos=None, resumenes=None, marca="", aliases=None, progress=None, cuerpos=None):
    try:
        if progress is not None:
            progress.progress(0.02, "Cargando modelo PKL...")
        if hasattr(pkl_file, "seek"):
            pkl_file.seek(0)
        pipeline = joblib.load(pkl_file)
        TM = {
            1: "Positivo", "1": "Positivo", "positivo": "Positivo", "Positivo": "Positivo",
            0: "Neutro", "0": "Neutro", "neutro": "Neutro", "Neutro": "Neutro",
            -1: "Negativo", "-1": "Negativo", "negativo": "Negativo", "Negativo": "Negativo",
        }
        def _norm_pred(p):
            if p in TM: return TM[p]
            s = str(p).strip()
            return TM.get(s, TM.get(s.title(), s.title() if s.title() in ("Positivo", "Negativo", "Neutro") else "Neutro"))

        if marca and titulos is not None and resumenes is not None:
            titulos = list(titulos)
            resumenes = list(resumenes)
            n = len(titulos)
            snippets, flags = [], []
            for i in range(n):
                ctx = extraer_contexto_marca(titulos[i], resumenes[i], marca, aliases, cuerpos[i] if cuerpos is not None else None)
                if ctx:
                    snippets.append(_snippet_tono_pkl(titulos[i], ctx)[:2400])
                    flags.append(True)
                else:
                    snippets.append("")
                    flags.append(False)
            result = [{"tono": "Neutro"}] * n
            idx_pred = [i for i, f in enumerate(flags) if f]
            if idx_pred:
                preds = _predict_pkl_in_batches(pipeline, [snippets[i] for i in idx_pred], progress)
                for i, p in zip(idx_pred, preds):
                    result[i] = {"tono": _norm_pred(p)}
            if titulos is not None and resumenes is not None:
                tonos = _propagar_tono_equivalentes([r["tono"] for r in result], list(titulos), list(resumenes))
                return [{"tono": t} for t in tonos]
            return result
        preds = [{"tono": _norm_pred(p)} for p in _predict_pkl_in_batches(pipeline, textos, progress)]
        if titulos is not None and resumenes is not None:
            tonos = _propagar_tono_equivalentes([r["tono"] for r in preds], list(titulos), list(resumenes))
            return [{"tono": t} for t in tonos]
        return preds
    except Exception as e:
        st.error(f"Error pkl tono: {e}")
        return None

def analizar_temas_con_pkl(textos, pkl_file):
    try:
        pipeline = joblib.load(pkl_file)
        predicciones = pipeline.predict(textos)
        return [str(p).strip() for p in predicciones]
    except Exception as e:
        st.error(f"Error pkl temas: {e}")
        return None

# ======================================
# SUBTEMAS
# ======================================
class ClasificadorSubtema:
    def __init__(self, marca, aliases):
        nombres = _variantes_marca(marca, aliases)
        self.marca = nombres[0] if nombres else str(marca or "")
        self.aliases = nombres[1:]
        self._cache = {}
        self._umbrales: dict = {}

    def _paso1(self, titulos, resumenes, dsu):
        def nt(t, n):
            return ' '.join(re.sub(r'[^a-z0-9\s]', '', unidecode(str(t).lower())).split()[:n])
        bt, br = defaultdict(list), defaultdict(list)
        for i, (ti, re_) in enumerate(zip(titulos, resumenes)):
            a, b = nt(ti, 40), nt(re_, 15)
            if a: bt[hashlib.md5(a.encode()).hexdigest()].append(i)
            b = nt(re_, 120)
            if len(b.split()) >= 25: br[hashlib.md5(b.encode()).hexdigest()].append(i)
        for bk in (bt, br):
            for idxs in bk.values():
                for j in idxs[1:]: dsu.union(idxs[0], j)

    def _paso2(self, titulos, dsu):
        norm = [normalize_title_for_comparison(t) for t in titulos]
        n = len(norm)
        for i in range(n):
            if not norm[i]: continue
            for j in range(i + 1, n):
                if not norm[j] or dsu.find(i) == dsu.find(j): continue
                ratio = SequenceMatcher(None, norm[i], norm[j]).ratio()
                comparte_asunto = _overlap_distintivo(norm[i], norm[j]) >= 0.40
                if ratio >= SIMILARITY_THRESHOLD_TITULOS and comparte_asunto and not _hay_conflicto_accion(norm[i], norm[j]):
                    dsu.union(i, j)

    def _paso2b_keywords(self, titulos, dsu, ae):
        sim_min = self._umbrales.get('sim_minima_keywords', SIM_MINIMA_KEYWORDS_RARAS)
        stop = {
            'el','la','los','las','un','una','unos','unas','de','del','al',
            'en','con','por','para','que','se','su','sus','es','son','fue',
            'como','mas','pero','sin','sobre','entre','tras','esta','este',
            'esto','hay','ser','han','ha','ya','muy','otro','otra','otros',
            'otras','todo','toda','todos','todas','puede','desde','hasta',
            'donde','cuando','quien','cual','cada','nos','les','ante','bajo',
            'nueva','nuevo','nuevos','nuevas','forma','hace','asi','sera',
            'segun','tiene','fueron','sido','hacer','dice','dijo','tambien',
        }
        titulo_words = []
        for t in titulos:
            ws = set()
            for w in re.findall(r'[a-z]+', unidecode(str(t).lower())):
                if len(w) >= 5 and w not in stop: ws.add(w)
            titulo_words.append(ws)
        word_freq = Counter()
        for ws in titulo_words:
            for w in ws: word_freq[w] += 1
        n = len(titulos)
        # 'Raras' = palabras distintivas que conectan noticias: aparecen en >=2 títulos
        # pero en una minoría (<= 8% del corpus). El 3% anterior se colapsaba a 2 en corpus
        # pequeños (solo palabras con frecuencia exacta 2), dejando el paso casi inerte.
        max_freq = max(3, int(n * 0.08))
        rare_index = defaultdict(list)
        for i, ws in enumerate(titulo_words):
            for w in ws:
                if 2 <= word_freq[w] <= max_freq: rare_index[w].append(i)
        for idxs in rare_index.values():
            for a in range(len(idxs)):
                for b in range(a + 1, len(idxs)):
                    ia, ib = idxs[a], idxs[b]
                    if dsu.find(ia) == dsu.find(ib): continue
                    ea, eb = ae[ia], ae[ib]
                    if ea is None or eb is None: continue
                    sim = cosine_similarity(
                        np.array(ea).reshape(1, -1),
                        np.array(eb).reshape(1, -1)
                    )[0][0]
                    if sim >= sim_min and not _hay_conflicto_accion(str(titulos[ia]), str(titulos[ib])):
                        dsu.union(ia, ib)

    def _paso3(self, et, ae, dsu, pbar, ps):
        umbral_cluster = max(self._umbrales.get('subtema', UMBRAL_SUBTEMA), 0.82)
        sim_min = max(self._umbrales.get('sim_minima_agrupacion', SIM_MINIMA_AGRUPACION_SUBTEMA), 0.90)
        n = len(et)
        if n < 2: return

        def _puede_unir(i, j):
            if _hay_conflicto_accion(et[i], et[j]):
                return False
            if _overlap_distintivo(et[i], et[j]) >= 0.30:
                return True
            return SequenceMatcher(
                None,
                normalize_title_for_comparison(et[i]),
                normalize_title_for_comparison(et[j])
            ).ratio() >= 0.96

        B = 500
        if n <= B:
            pbar.progress(ps, "Clustering semántico...")
            ok = [(k, e) for k, e in enumerate(ae) if e is not None]
            if len(ok) < 2: return
            io_, M = zip(*ok)
            sim_matrix = cosine_similarity(np.array(M))
            linkage = 'complete' if n <= 10 else 'average'
            labels = AgglomerativeClustering(
                n_clusters=None, distance_threshold=1 - umbral_cluster,
                metric='precomputed', linkage=linkage
            ).fit(1 - sim_matrix).labels_
            g = defaultdict(list)
            for k, lbl in enumerate(labels): g[lbl].append(io_[k])
            for cl in g.values():
                if len(cl) < 2: continue
                vecs = np.array([ae[i] for i in cl if ae[i] is not None])
                if len(vecs) < 2: continue
                centroid = np.mean(vecs, axis=0)
                sims_al_centroid = cosine_similarity(vecs, centroid.reshape(1, -1)).flatten()
                todos_ok = all(s >= sim_min for s in sims_al_centroid)
                if todos_ok:
                    for j in cl[1:]:
                        if _puede_unir(cl[0], j):
                            dsu.union(cl[0], j)
                else:
                    mejor_idx = int(np.argmax(sims_al_centroid))
                    repr_vec = np.array(ae[cl[mejor_idx]]).reshape(1, -1)
                    for k_local, i_global in enumerate(cl):
                        if ae[i_global] is None: continue
                        sim_vs_repr = cosine_similarity(
                            np.array(ae[i_global]).reshape(1, -1), repr_vec
                        )[0][0]
                        if sim_vs_repr >= sim_min and _puede_unir(cl[mejor_idx], i_global):
                            dsu.union(cl[mejor_idx], i_global)
            pbar.progress(ps + 0.18, "Clustering completado")
            return

        tb = max(1, (n + B - 1) // B)
        for bn_, bs in enumerate(range(0, n, B)):
            bi = list(range(bs, min(bs + B, n)))
            ok = [(idx, ae[idx]) for idx in bi if ae[idx] is not None]
            if len(ok) < 2: continue
            io_, M = zip(*ok)
            sim_matrix = cosine_similarity(np.array(M))
            labels = AgglomerativeClustering(
                n_clusters=None, distance_threshold=1 - umbral_cluster,
                metric='precomputed', linkage='average'
            ).fit(1 - sim_matrix).labels_
            g = defaultdict(list)
            for k, lbl in enumerate(labels): g[lbl].append(io_[k])
            for cl in g.values():
                if len(cl) < 2: continue
                vecs = np.array([ae[i] for i in cl if ae[i] is not None])
                if len(vecs) < 2: continue
                centroid = np.mean(vecs, axis=0)
                sims = cosine_similarity(vecs, centroid.reshape(1, -1)).flatten()
                mejor_idx = int(np.argmax(sims))
                repr_vec = np.array(ae[cl[mejor_idx]]).reshape(1, -1)
                for k_local, i_global in enumerate(cl):
                    if ae[i_global] is None: continue
                    s = cosine_similarity(np.array(ae[i_global]).reshape(1, -1), repr_vec)[0][0]
                    if s >= sim_min and _puede_unir(cl[mejor_idx], i_global):
                        dsu.union(cl[mejor_idx], i_global)
            pbar.progress(ps + 0.15 * (bn_ + 1) / tb, f"Clustering {bn_ + 1}/{tb}...")

        pbar.progress(ps + 0.16, "Unificando...")
        usar_fusion = self._umbrales.get('usar_fusion_iterativa', True)
        if usar_fusion: self._fusion(et, ae, dsu, pbar, ps + 0.16)

    def _fusion(self, textos, ae, dsu, pbar, ps):
        n = len(textos)
        umbral_inter = self._umbrales.get('fusion_intergrupo', UMBRAL_FUSION_INTERGRUPO)
        max_iter = self._umbrales.get('max_iter_fusion', MAX_ITER_FUSION)
        sim_min = self._umbrales.get('sim_minima_agrupacion', SIM_MINIMA_AGRUPACION_SUBTEMA)
        for it in range(max_iter):
            grupos = dsu.grupos(n)
            if len(grupos) < 2: break
            centroids, vg = [], []
            for gid, idxs in grupos.items():
                vecs = [ae[i] for i in idxs[:50] if ae[i] is not None]
                if vecs:
                    centroids.append(np.mean(vecs, axis=0))
                    vg.append(gid)
            if len(vg) < 2: break
            sim = cosine_similarity(np.array(centroids))
            umbral_efectivo = max(umbral_inter, sim_min)
            pairs = sorted(
                [(sim[i][j], i, j) for i in range(len(vg)) for j in range(i + 1, len(vg))
                 if sim[i][j] >= umbral_efectivo], reverse=True
            )
            fus = 0
            for _, i, j in pairs:
                ri, rj = grupos[vg[i]][0], grupos[vg[j]][0]
                if dsu.find(ri) != dsu.find(rj):
                    textos_i = [textos[k] for k in grupos[vg[i]][:20]]
                    textos_j = [textos[k] for k in grupos[vg[j]][:20]]
                    if _grupos_contenido_compatibles(
                        textos_i,
                        textos_j,
                        "",
                        "",
                        min_sim=umbral_efectivo,
                        min_overlap=0.16,
                    ):
                        dsu.union(ri, rj)
                        fus += 1
            pbar.progress(min(ps + 0.04 * (it + 1), 0.52), f"Fusión {it + 1}: {fus}")
            if fus == 0: break

    def _generar_etiqueta(self, textos_grp, titulos_grp, resumenes_grp, subtemas_existentes=None, evitar_etiqueta=None):
        tn = sorted(set(normalize_title_for_comparison(t) for t in titulos_grp if t))
        existentes_key = "|".join(sorted(string_norm_label(s) for s in (subtemas_existentes or []))[:20])
        evitar_key = string_norm_label(evitar_etiqueta) if evitar_etiqueta else ""
        ck = hashlib.md5(("|".join(tn[:12]) + f"#{len(titulos_grp)}#{existentes_key}#{evitar_key}").encode()).hexdigest()
        if ck in self._cache: return self._cache[ck]

        tm = list(dict.fromkeys(str(t)[:160] for t in titulos_grp if pd.notna(t) and str(t).strip() and str(t).strip().lower() != 'nan'))[:6]
        rm = [str(r)[:260] for r in resumenes_grp[:3] if r and len(str(r)) > 20]

        # Fuentes para el "grounding": título + resumen + texto rico (ya incluye contexto de la marca).
        fuentes_grounding = [str(t) for t in titulos_grp if t and str(t).strip()]
        fuentes_grounding += [str(r) for r in resumenes_grp if r and str(r).strip()]
        fuentes_grounding += [str(t) for t in textos_grp[:3] if t and str(t).strip()]

        lista_existentes = ""
        if subtemas_existentes and len(subtemas_existentes) > 0:
            lista_existentes = (
                "\n\nSUBTEMAS YA CREADOS (REUTILÍZALOS SOLO SI ES EXACTAMENTE EL MISMO HECHO):\n"
                + ", ".join(f"'{s}'" for s in subtemas_existentes[:15])
                + "\nSi este grupo de noticias trata EXACTAMENTE el mismo hecho que uno de los subtemas ya creados, "
                "responde con ese subtema palabra por palabra. Si es otro hecho, crea uno nuevo."
            )
        if evitar_etiqueta:
            lista_existentes += f"\nNO uses '{evitar_etiqueta}': es un hecho distinto, genera un subtema nuevo y específico."

        bloq_resumenes = ("\nRESÚMENES:\n" + "\n".join(f"  · {r}" for r in rm)) if rm else ""

        # Contexto amplio de la MARCA (extraído del texto rico `_txt`): llega al prompt del
        # subtema con más caracteres para que el asunto se derive de la mención de la marca.
        ctx_txt = [str(t) for t in textos_grp[:2] if t and str(t).strip()]
        bloq_contexto = ("\n\nCONTEXTO AMPLIADO (puede incluir la mención de la marca/alias):\n"
                         + "\n".join(f"  · {c[:700]}" for c in ctx_txt)) if ctx_txt else ""

        prompt = (
            f"Eres analista de reputación que monitorea noticias sobre '{self.marca}'.\n"
            "Lee las noticias y resume EL HECHO periodístico central RELACIONADO CON "
            f"'{self.marca}' en UNA sola frase nominal descriptiva de 2 a 5 palabras, "
            "gramaticalmente correcta y con sentido lógico completo.\n\n"
            "REGLAS DE HECHO:\n"
            "  - El hecho debe describir QUÉ HIZO O QUÉ LE PASÓ a la marca/cliente según el texto. "
            "  - Si la marca solo ATIENDE a un paciente o víctima (hospital, clínica), el hecho es de "
            "ATENCIÓN, TRATAMIENTO, CIRUGÍA o REHABILITACIÓN, NO un reconocimiento ni un premio. "
            "  - NO inventes el tipo de hecho: si no hay premio, no digas 'Reconocimiento'. "
            "  - Cada palabra de contenido del subtema debe aparecer (o derivar de) el texto. "
            "  - NO copies el titular casi literal ni uses dos puntos o guiones (eso es un titular, no un subtema).\n"
            "CÓMO CONSTRUIRLA:\n"
            "  1. Identifica primero el TIPO de hecho: lanzamiento, convenio, alianza, inversión, "
            "proyecto, campaña, foro, premiación, reconocimiento, nombramiento, designación, posesión, "
            "renuncia, investigación, sanción, publicación de un libro, apertura, intercambio, "
            "atención médica, cirugía, tratamiento, etc.\n"
            "  2. Escribe: [tipo de hecho] + [preposición: de/del/para/sobre/en] + [objeto o asunto concreto]. "
            "La frase debe leerse como una categoría, no copiar un titular.\n"
            "  3. Usa SOLO palabras que aparezcan en el texto analizado (o sus derivadas directas, "
            "ej. 'renunció' → 'renuncia'). NO inventes nombres, lugares, cargos ni términos. "
            "Deriva la frase del TÍTULO o del RESUMEN-ACLARACIÓN: cada palabra de contenido debe aparecer en esos textos.\n"
            "  4. Sintetiza el hecho; NO copies el titular completo ni frases sueltas del texto.\n"
            "  5. Si el hecho NO está vinculado con la marca, describe el tema real de la noticia sin forzar la relación.\n\n"
            "PROHIBIDO (se rechaza automáticamente):\n"
            "  - Empezar por nombre de persona o cargo ('Jesús Martínez', 'Ever Pallares', 'Alcalde', 'Gobernador', 'Superintendente').\n"
            "  - Empezar por un lugar o país ('La Guajira', 'Colombia', 'Barranquilla').\n"
            "  - Verbo conjugado ('presenta', 'lanza', 'plantea', 'renunció', 'asume', 'fue').\n"
            "  - Dos sustantivos pegados sin preposición ('Guajira escenario', 'Colombia escudo').\n"
            "  - Adjetivo después de 'de' cuando debe ir pegado ('Explotación de sexual' es incorrecto; correcto: 'Explotación sexual').\n"
            "  - Etiquetas genéricas ('Cobertura de información relevante', 'Cobertura informativa general', "
            "'Información relevante', 'Gestión corporativa', 'Actividad institucional').\n"
            "  - RÓTULOS DE RELLENO que describen el marco y no el hecho: 'Destacados del sector', "
            "'Panorama del sector', 'Actualidad del sector', 'Noticias del sector', 'Menciones de la marca', "
            "'Presencia en medios', 'Cobertura mediática', 'Contexto general', 'Otros temas', 'Aspectos generales'. "
            "Si dudas, pregúntate: ¿mi etiqueta dice QUÉ PASÓ? Si solo dice DÓNDE se publicó o DE QUÉ SECTOR es, está mal.\n"
            "  - Solo una palabra clave suelta ('Pollo', 'Precio'): debe ser una categoría, no un keyword.\n\n"
            f"TÍTULOS:\n" + "\n".join(f"  · {t}" for t in tm)
            + bloq_resumenes
            + bloq_contexto
            + lista_existentes
            + "\n\nEJEMPLOS CORRECTOS: 'Convenio de cooperación científica', 'Reconocimiento al liderazgo regional', "
            "'Intercambio intercultural en La Guajira', 'Explotación sexual de menores', 'Posesión del superintendente de Notariado'\n"
            "EJEMPLOS INCORRECTOS: 'Jesus de martinez', 'Ever de pallares', 'Guajira de escenario', 'Colombia de escudo', "
            "'Foro de plantea', 'Memoria de caribe', 'Divorcio de sirena'\n\n"
            'JSON: {"subtema":"..."}'
        )

        _VERBOS_FRASES = re.compile(
            r'\b(presenta|presentan|anuncia|anuncian|lanza|lanzan|inaugura|inauguran|'
            r'realiza|realizan|desarrolla|desarrollan|ejecuta|ejecutan|gestiona|gestionan|'
            r'impulsa|impulsan|promueve|promueven|lidera|lideran|encabeza|encabezan|'
            r'aprueba|aprueban|firma|firman|suscribe|suscriben|invierte|invierten|'
            r'construye|construyen|instala|instalan|entrega|entregan|recibe|reciben|'
            r'solicita|solicitan|visita|visitan|atiende|atienden|destaca|destacan|'
            r'señala|señalan|indica|indican|expresa|expresan|afirma|afirman|'
            r'propone|proponen|pide|piden|exige|exigen|apoya|apoyan|'
            r'informa|informan|reporta|reportan|advierte|advierten|'
            r'levanta|levantan|levantaron|levanto|impacta|impactan|encarece|encarecen|'
            r'encarecio|sube|suben|subio|baja|bajan|bajaron|gano|ganan|ganaron|'
            r'pierde|pierden|perdio|logra|logran|busca|buscan|crece|crecen|'
            r'aumenta|aumentan|conquista|conquistan|derrumba|derrumban|recupera|recuperan|'
            r'plantea|plantean|planteo|renuncia|renuncian|renuncio|asume|asumen|asumio|'
            r'posesiona|posesionan|posesiono|nombra|nombran|nombro|designa|designan|designo|'
            r'representa|representan|dimite|dimitio)\b',
            re.IGNORECASE
        )

        def _tiene_verbo_conjugado(s): return bool(_VERBOS_FRASES.search(s))

        def _primera_palabra_verbo(s):
            prim = unidecode((s or "").strip().lower() or "").split()
            if not prim:
                return False
            p0 = prim[0].rstrip(".,!?;:")
            return p0 in _VERBOS_LEAD_SUBTEMA or bool(_RE_VERBO_SUBTEMA.search(p0))

        try:
            resp = call_with_retries(
                openai.ChatCompletion.create,
                model=OPENAI_MODEL_CLASIFICACION,
                messages=[{"role": "user", "content": prompt}],
                max_tokens=180,
                temperature=0.0,
                response_format={"type": "json_object"}
            )
            u = resp.get('usage', {}) if isinstance(resp, dict) else getattr(resp, 'usage', {})
            if u:
                st.session_state['tokens_input'] += (u.get('prompt_tokens') if isinstance(u, dict) else getattr(u, 'prompt_tokens', 0)) or 0
                st.session_state['tokens_output'] += (u.get('completion_tokens') if isinstance(u, dict) else getattr(u, 'completion_tokens', 0)) or 0

            raw = json.loads(resp.choices[0].message.content).get("subtema", "Varios")
            et = limpiar_tema(raw)

            if not et or et.strip().lower() == "sin tema":
                et = self._refinar(tm, None, rm, forzar_preposicion=True, prohibir_verbos=True)
            if _tiene_verbo_conjugado(et):
                et = self._refinar(tm, None, rm, forzar_preposicion=True, prohibir_verbos=True)

            def _es_robotico(s):
                palabras = s.split()
                if len(palabras) <= 3:
                    nexos = {"de", "del", "para", "sobre", "en", "con", "por",
                             "ante", "hacia", "entre", "sin", "al", "las", "los",
                             "una", "uno", "que", "como", "y", "o", "a", "e", "u"}
                    tiene_nexo = any(unidecode(p.lower()) in nexos for p in palabras[1:])
                    if not tiene_nexo: return True
                return False

            genericas = {"gestión", "gestion", "actividades", "acciones", "noticias",
                         "información", "informacion", "eventos", "varios", "sin tema",
                         "actividad corporativa", "gestion corporativa",
                         "impacto en la reputacion", "impacto reputacional",
                         "reputacion corporativa", "impacto corporativo"}
            es_gen = string_norm_label(et) in {string_norm_label(g) for g in genericas} or _es_etiqueta_generica(et)
            es_solo_marca = _es_nombre_o_fragmento_marca(et, self.marca, self.aliases)
            es_verboso_marca = _es_verboso_con_marca(et, self.marca, self.aliases)
            es_rob = _es_robotico(et)

            if es_gen or es_solo_marca or es_verboso_marca or es_rob or len(et.split()) < 3:
                et = self._refinar(tm, None, rm, forzar_preposicion=True, prohibir_verbos=True)

            if not _validar_estructura_subtema(et):
                et = self._refinar(tm, None, rm, forzar_preposicion=True, prohibir_verbos=True)
                if not _validar_estructura_subtema(et):
                    et = self._fallback(titulos_grp, resumenes_grp)

            # Refuerzo 1: nombres propios / cargos / números / siglas
            if _empieza_por_nombre_propio(et, titulos_grp) or _contiene_numero_o_acronimo(et):
                et = self._refinar(tm, None, rm, forzar_preposicion=True, prohibir_verbos=True, prohibir_nombres=True)
            if _empieza_por_nombre_propio(et, titulos_grp) or _contiene_numero_o_acronimo(et):
                et = self._fallback(titulos_grp, resumenes_grp)

            # Refuerzo 2: grounding (no inventos)
            if not _subtema_grounded(et, fuentes_grounding):
                et = self._refinar(tm, None, rm, forzar_preposicion=True, prohibir_verbos=True)
            if not _subtema_grounded(et, fuentes_grounding):
                et = self._fallback(titulos_grp, resumenes_grp)

            # Refuerzo 2b: CABEZA ANCLADA. El núcleo de hecho del subtema debe estar
            # en el texto. Mata 'Reconocimiento a fundación santa fe' cuando el texto
            # solo dice que atendieron a un paciente: 'reconocimiento' no aparece.
            if not _head_anclada(et, fuentes_grounding):
                et = self._refinar(tm, None, rm, forzar_preposicion=True, prohibir_verbos=True)
            if not _head_anclada(et, fuentes_grounding):
                et = self._fallback(titulos_grp, resumenes_grp)
            # Refuerzo 2c: sin marcadores de titular (dos puntos, guion) -> recorte
            # crudo, no un subtema real.
            if not _validar_estructura_subtema(et):
                et = self._fallback(titulos_grp, resumenes_grp)

            # Refuerzo final anti-frases-sin-sentido
            if _tiene_verbo_conjugado(et) or _primera_palabra_verbo(et) or _es_robotico(et) or _empieza_por_nombre_propio(et, titulos_grp):
                et = self._fallback(titulos_grp, resumenes_grp)

            et = _validar_etiqueta_completa(
                et, titulos_grp=titulos_grp, resumenes_grp=resumenes_grp,
                marca=self.marca, aliases=self.aliases, fallback_fn=lambda tl: self._fallback(tl, resumenes_grp)
            )
            if _es_nombre_o_fragmento_marca(et, self.marca, self.aliases):
                et = self._refinar(tm, None, rm, forzar_preposicion=True, prohibir_verbos=True, prohibir_nombres=True)
            if _es_nombre_o_fragmento_marca(et, self.marca, self.aliases):
                et = self._fallback(titulos_grp, resumenes_grp)
        except:
            et = self._fallback(titulos_grp, resumenes_grp)

        et = capitalizar_etiqueta(et)
        self._cache[ck] = et
        return et

    def _refinar(self, titulos, kw=None, resumenes=None, forzar_preposicion=False, prohibir_verbos=False, prohibir_nombres=False):
        ctx = ("\nContexto: " + " | ".join(str(r)[:140] for r in (resumenes or [])[:3])) if resumenes else ""
        restricciones = []
        if forzar_preposicion:
            restricciones.append("Incluye una preposición (de/del/para/sobre/en) entre los conceptos.")
        if prohibir_verbos:
            restricciones.append("Prohibido: verbos conjugados y empezar por cargo o nombre de persona.")
        if prohibir_nombres:
            restricciones.append("Prohibido: empezar por nombre de persona o de lugar.")
        bloque_rest = ("\n".join("  - " + r for r in restricciones)) if restricciones else "  - (ninguna adicional)"

        prompt = (
            f"Eres analista de reputación de '{self.marca}'. Reescribe el hecho de las noticias "
            "en UNA frase nominal descriptiva de 3-5 palabras, correcta y completa, con orden lógico.\n\n"
            f"Títulos:\n" + "\n".join(f"  · {t[:150]}" for t in titulos[:5]) + ctx + "\n\n"
            f"Restricciones:\n{bloque_rest}\n\n"
            "Formato correcto: [tipo de hecho] + [preposición] + [objeto/asunto]. "
            "Ej.: 'Convenio de cooperación científica', 'Investigación por fallas operativas', "
            "'Explotación sexual de menores'.\n"
            "Usa SOLO palabras del texto. Tildes y ñ correctas. No copies el titular literal.\n"
            "PROHIBIDO: rótulos genéricos ('Cobertura de información relevante', 'Cobertura informativa general').\n"
            'JSON: {"subtema":"..."}'
        )
        try:
            resp = call_with_retries(
                openai.ChatCompletion.create,
                model=OPENAI_MODEL_CLASIFICACION,
                messages=[{"role": "user", "content": prompt}],
                max_tokens=180,
                temperature=0.2,
                response_format={"type": "json_object"}
            )
            raw = json.loads(resp.choices[0].message.content).get("subtema", "Varios")
            et = limpiar_tema(raw)
            if not _frase_esta_completa(et):
                et = _recortar_frase_completa(et)
                if not _frase_esta_completa(et): return self._fallback(titulos, resumenes)
            return et
        except:
            return self._fallback(titulos or [], resumenes)

    def _extraer_desde_texto(self, titulos, resumenes):
        """Un intento LLM ESTRICTO: la frase se arma ÚNICAMENTE con palabras que
        aparecen literalmente en Título / Resumen-Aclaración. '' si no valida."""
        if not titulos and not resumenes:
            return ""
        tm = list(dict.fromkeys(str(t)[:160] for t in titulos if str(t).strip()))[:4]
        rm = [str(r)[:200] for r in resumenes[:2] if r and len(str(r)) > 20]
        if not tm and not rm:
            return ""
        bloq_t = ("\n".join(f"  · {t}" for t in tm)) if tm else "(no hay títulos)"
        bloq_r = ("\n".join(f"  · {r}" for r in rm)) if rm else "(no hay resúmenes)"
        prompt = (
            "Eres analista de reputación. Extrae del texto periodístico la frase nominal "
            "de 3 a 5 palabras que mejor describa EL HECHO central del grupo.\n\n"
            f"TÍTULOS:\n{bloq_t}\n\nRESÚMENES-ACLARACIÓN:\n{bloq_r}\n\n"
            "REGLAS ESTRICTAS:\n"
            "  - Usa ÚNICAMENTE palabras que aparezcan literalmente en los títulos o "
            "resúmenes (puedes eliminar palabras o cambiar su orden, pero NO añadir ni inventar ninguna).\n"
            "  - Formato nominal: [tipo de hecho] + [preposición] + [asunto]. "
            "Ej.: 'Convenio de cooperación científica', 'Ampliación de plataforma digital'.\n"
            "  - PROHIBIDO: rótulos genéricos ('Cobertura de información relevante', "
            "'Cobertura informativa general', 'Información relevante'), empezar por nombre "
            "de persona, cargo o lugar, y verbos conjugados.\n"
            'JSON: {"subtema":"..."}'
        )
        try:
            resp = call_with_retries(
                openai.ChatCompletion.create,
                model=OPENAI_MODEL_CLASIFICACION,
                messages=[{"role": "user", "content": prompt}],
                max_tokens=120,
                temperature=0.0,
                response_format={"type": "json_object"}
            )
            u = resp.get('usage', {}) if isinstance(resp, dict) else getattr(resp, 'usage', {})
            if u:
                st.session_state['tokens_input'] += (u.get('prompt_tokens') if isinstance(u, dict) else getattr(u, 'prompt_tokens', 0)) or 0
                st.session_state['tokens_output'] += (u.get('completion_tokens') if isinstance(u, dict) else getattr(u, 'completion_tokens', 0)) or 0
            raw = json.loads(resp.choices[0].message.content).get("subtema", "")
            et = limpiar_tema(raw)
            fuentes = [str(t) for t in titulos if str(t).strip()] + [str(r) for r in resumenes if str(r).strip()]
            if (et and not _es_etiqueta_generica(et)
                    and _frase_esta_completa(et)
                    and _validar_estructura_subtema(et)
                    and _subtema_grounded(et, fuentes)
                    and not _empieza_por_nombre_propio(et, titulos)):
                return et
        except Exception:
            pass
        return ""

    def _derivar_desde_texto_nominal(self, textos, fuentes, es_resumen=False):
        """Deriva la frase nominal DESDE la oración donde vive el hecho (prioridad:
        el resumen, que es donde suele aparecer la mención del cliente).

        Para la fila real "Fue trasladado de urgencia a la Fundación Santa Fe y sometido
        a una cirugía compleja", produce 'Cirugía de alta complejidad' (hecho anclado)
        en lugar de 'Reconocimiento a fundación santa fe' (inventado) o de recortar el
        titular. Detecta el hecho y toma el OBJETO de la propia oración del hecho.
        """
        for t in (textos or [])[:4]:
            s = _sanear_frase_nominal(str(t), max_palabras=60)
            if not s or len(s) < 8:
                continue
            sl = unidecode(s.lower())
            # Todos los hechos detectables, con su posición; prioriza los de salud.
            hits = []
            for pat, nom in (_NUCLEOS_HECHO_EVENTO + self._NUCLEOS_HECHO):
                m = re.search(pat, sl)
                if m:
                    hits.append((m.start(), len(nom) if nom != "Cirugía" else 2, nom))
            if not hits:
                continue
            # El hecho más relevante: el de salud primero (cirugía/atención/traslado).
            _PRIO_SALUD = {"Cirugía": 6, "Rehabilitación": 5, "Hospitalización": 5,
                           "Tratamiento": 5, "Atención a paciente": 5, "Atención a víctima": 5,
                           "Diagnóstico": 4, "Atención": 4, "Atención de urgencia": 3,
                           "Traslado": 2, "Uso": 1}
            def _cla(h):
                pos, _, nom = h
                return (_PRIO_SALUD.get(nom, 0), -pos)
            pos, _, hecho = max(hits, key=_cla)
            # Si el hecho YA incluye complemento ('Atención a víctima', 'Atención de
            # urgencia'), es completo: no agregar objeto adicional.
            if re.match(r"^[A-Za-zÁÉÍÓÚÑÜáéíóúñü]+\s+(?:a|de|del|a\s+la|de\s+la|al)\b", hecho):
                return capitalizar_etiqueta(_sanear_frase_nominal(hecho))
            # OBJETO: la palabra de contenido que SIGUE al hecho dentro de la misma
            # oración (evita 'nació parálisis' del inicio del texto).
            palabras_hecho = {_normaliza_token(w) for w in hecho.lower().split()}
            seg = sl[pos: len(sl)]
            objeto = ""
            for w in re.findall(r"[a-z]+", seg):
                wn = _normaliza_token(w)
                if (len(wn) >= 4 and wn not in _CONECTORES_ETIQUETA
                        and wn not in _VERBOS_LEAD_SUBTEMA
                        and not _RE_VERBO_SUBTEMA.search(wn)
                        and not _es_forma_verbal_es(wn)
                        and wn not in _CARGOS_SUBTEMA
                        and wn not in _TOKENS_DEBILES_SUBTEMA_FALLBACK
                        and wn not in _NUCLEOS_ACONTECIMIENTO
                        and wn not in palabras_hecho          # no repetir el hecho
                        and not _es_nombre_o_fragmento_marca(wn, self.marca, self.aliases)
                        and wn not in {_normaliza_token(x) for x in self.aliases + [self.marca]}):
                    objeto = wn
                    break
            if not objeto:
                continue
            if _parece_adjetivo_es(objeto) or objeto in {"alta", "alta", "compleja"}:
                # 'Cirugía compleja': el adjetivo va pegado, no tras 'de'. Si el
                # objeto es un participio verbal ('trasladada'), no sirve de objeto.
                if objeto.startswith(("activ", "traslad", "atend", "realiz", "llev",
                                      "somet", "oper", "trat", "intern", "hosp", "ingres")):
                    continue
                frase = f"{hecho} {_ADJ_POSIBLE.get(objeto, objeto)}"
            else:
                frase = f"{hecho} de {objeto}"
                nxt = ""
                seg2 = seg[seg.find(objeto) + len(objeto):]
                for w in re.findall(r"[a-z]+", seg2):
                    if w != objeto and _parece_adjetivo_es(w):
                        nxt = w
                        break
                if nxt and len(f"{frase} {nxt}".split()) <= MAX_PALABRAS_SUBTEMA:
                    frase = f"{frase} {nxt}"
            frase = capitalizar_etiqueta(_sanear_frase_nominal(frase))
            if (_frase_esta_completa(frase) and len(frase.split()) >= 2
                    and not _es_etiqueta_generica(frase)
                    and not _es_nombre_o_fragmento_marca(frase, self.marca, self.aliases)):
                return frase
        return ""

    def _construir_frase_accion(self, titulos, resumenes):
        """Frase 'Acción de asunto' con palabras reales del texto. Determinista y grounded."""
        if not titulos and not resumenes:
            return ""
        nombres_iniciales = set()
        for t in titulos:
            nombres_iniciales |= _nombres_propios_iniciales_titulo(t)

        acciones = [
            (r"\b(lanzamiento|lanza|lanzo|estrena|estreno|presenta|presento|presentacion)\b", "Lanzamiento"),
            (r"\b(anuncia|anuncio|anuncian)\b", "Anuncio"),
            (r"\b(inaugura|inauguro|inauguracion|apertura|abre|abrio)\b", "Apertura"),
            (r"\b(firma|firmo|suscribe|suscribio|convenio|alianza|pacto|acuerdo)\b", "Convenio"),
            (r"\b(recibe|recibio|premio|reconocimiento|galardon|distincion|honoris|causa|premiacion)\b", "Reconocimiento"),
            (r"\b(investiga|investigacion|sancion|sancio|denuncia|demanda|multa)\b", "Investigación"),
            (r"\b(renuncia|renuncio|dimite|dimitio)\b", "Renuncia"),
            (r"\b(designa|designo|nombra|nombro|nombramiento|asume|asumio|posesiona|posesiono|representante|nombrada|nombrado|designada|designado|elegida|elegido)\b", "Designación"),
            (r"\b(invierte|inversion|inversiones|invirtio|invertira|destina|destino|destinara)\b", "Inversión"),
            (r"\b(proyecto|proyectos)\b", "Proyecto"),
            (r"\b(reforma|reformas)\b", "Reforma"),
            (r"\b(convocatoria|convoca|convocan)\b", "Convocatoria"),
            (r"\b(licitacion|licita|adjudica|adjudicacion)\b", "Licitación"),
            (r"\b(campana|campañas)\b", "Campaña"),
            (r"\b(foro|foros)\b", "Foro"),
            (r"\b(congreso)\b", "Congreso"),
            (r"\b(cumbre)\b", "Cumbre"),
            (r"\b(encuentro|encuentros)\b", "Encuentro"),
            (r"\b(seminario|seminarios)\b", "Seminario"),
            (r"\b(taller|talleres)\b", "Taller"),
            (r"\b(conversatorio|simposio)\b", "Conversatorio"),
            (r"\b(ampliacion|amplia|ampliara)\b", "Ampliación"),
            (r"\b(construccion|construira|construye|construyen)\b", "Construcción"),
            (r"\b(reactivacion|reactiva|reapertura)\b", "Reactivación"),
            (r"\b(cierre|suspension|clausura)\b", "Cierre"),
            (r"\b(balance|balances)\b", "Balance"),
            (r"\b(resultados|resultado)\b", "Resultados"),
            (r"\b(estudio|estudios)\b", "Estudio"),
            (r"\b(publicacion|publica obra|publica el libro|lanzamiento del libro)\b", "Publicación"),
            (r"\b(capacitacion|entrenamiento)\b", "Capacitación"),
        ]
        texto_total = " ".join(str(t) for t in titulos[:5])
        accion_par = next(
            ((patron, nombre) for patron, nombre in acciones
             if re.search(patron, unidecode(texto_total.lower()), re.IGNORECASE)),
            None,
        )
        accion = accion_par[1] if accion_par else None

        tokens_marca = set(_normalizar_mencion(" ".join([self.marca] + self.aliases)).split())
        excluir = (tokens_marca | STOPWORDS_ES | _VERBOS_LEAD_SUBTEMA | _CARGOS_SUBTEMA
                   | nombres_iniciales | _TOKENS_DEBILES_SUBTEMA_FALLBACK | {
            "universidad", "empresa", "compania", "corporacion", "fundacion", "institucion",
            "anuncio", "anuncia", "lanzamiento", "lanza", "presenta", "presencia",
            "invitado", "especial", "principal", "marca", "cliente", "noticia",
            "ultimo", "ultima", "ultimos", "ultimas", "nuevo", "nueva", "nuevos", "nuevas",
            "poder", "suerte", "gran", "grande", "grandes", "coccion", "lenta", "medio",
            "parte", "manera", "forma", "tipo", "asi", "pues", "mismo", "misma",
            "escenario", "escenarios", "contexto", "contextos", "varios", "toneladas",
        })
        if accion_par:
            # Evita que la propia cabeza ('Foro', 'Inversión', 'Campaña'...) o sus
            # formas verbales se cuelen como 'palabra clave' del asunto.
            excluir = excluir | {
                _normaliza_token(w) for w in re.findall(r"[A-Za-zÁÉÍÓÚÑÜáéíóúñü]+", accion_par[0])
            }

        palabras_ordenadas = []
        for t in titulos[:5]:
            for w in re.findall(r"[A-Za-zÁÉÍÓÚÑÜáéíóúñü]+", str(t)):
                norm = _normaliza_token(w)
                if (len(norm) >= 4 and norm not in excluir
                        and norm not in _TRAILING_INCOMPLETE
                        and not _RE_VERBO_SUBTEMA.search(norm)):
                    palabras_ordenadas.append((norm, w))
        if not palabras_ordenadas and resumenes:
            # Sin palabras útiles en títulos, se ancla en el resumen (también es texto real).
            for r in resumenes[:3]:
                for w in re.findall(r"[A-Za-zÁÉÍÓÚÑÜáéíóúñü]+", str(r)):
                    norm = _normaliza_token(w)
                    if (len(norm) >= 4 and norm not in excluir
                            and norm not in _TRAILING_INCOMPLETE
                            and not _RE_VERBO_SUBTEMA.search(norm)):
                        palabras_ordenadas.append((norm, w))
        if palabras_ordenadas:
            cnt = Counter(p[0] for p in palabras_ordenadas)
            top_norm = [n for n, _ in cnt.most_common(3)]
            origen = {}
            for norm, orig in palabras_ordenadas:
                origen.setdefault(norm, orig)
            top = [origen[n].lower() for n in top_norm]
            if accion and top:
                frase = _recortar_frase_completa(f"{accion} de {' '.join(top[:2])}", MAX_PALABRAS_SUBTEMA)
            elif accion:
                frase = _recortar_frase_completa(accion, MAX_PALABRAS_SUBTEMA)
            else:
                frase = ""
            if _frase_esta_completa(frase) and not _es_nombre_o_fragmento_marca(frase, self.marca, self.aliases):
                return capitalizar_etiqueta(frase)
        return ""

    def _derivar_desde_titulos(self, titulos):
        """Frase nominal tomada del título más representativo: 100% palabras del texto.
        La cabeza debe ser un SUSTANTIVO: si el título arranca por verbo, se
        nominaliza con el núcleo de hecho ('investiga' -> 'Investigación por ...')."""
        for t in titulos:
            s = str(t).strip()
            if len(s) < 12 or len(s.split()) < 3:
                continue
            for n in [self.marca] + [a for a in (self.aliases or []) if a]:
                n = str(n).strip()
                if n:
                    s = re.sub(r'\b' + re.escape(n) + r'\b', ' ', s, flags=re.IGNORECASE)
            s = re.sub(r'\s+', ' ', s).strip()
            toks = s.split()
            nombres = _nombres_propios_iniciales_titulo(t)
            i = 0
            verbo_saltado = None
            while i < len(toks):
                tok = toks[i].lower().rstrip('.,;:!?')
                tn = unidecode(tok)
                es_verbo = (tn in _VERBOS_LEAD_SUBTEMA or bool(_RE_VERBO_SUBTEMA.search(tn)))
                if (re.match(r'^\d', tn)
                        or tn in _ARTICULOS_SUBTEMA
                        or tn in _CONECTORES_ETIQUETA
                        or es_verbo
                        or tn in _CARGOS_SUBTEMA
                        or tn in _TOKENS_DEBILES_SUBTEMA_FALLBACK
                        or tn in {"nuevo", "nueva", "nuevos", "nuevas", "gran", "grande", "grandes",
                                  "este", "esta", "estos", "estas"}
                        or _PATRON_TITULAR.match(toks[i])
                        or _RE_VERBO_SUBTEMA.search(tn)
                        or (tn not in _CABEZAS_SUBTEMA_VALIDAS and tn in nombres)):
                    if es_verbo and verbo_saltado is None:
                        verbo_saltado = tn      # recuerda el hecho para nominalizar
                    i += 1
                    continue
                break
            rest = toks[i:]
            if len(rest) < 2:
                continue
            frase = _sanear_frase_nominal(" ".join(rest))
            # Tras el saneo puede quedar un resto sin cabeza de hecho ('Gas a
            # industria', 'Energía solar'): se re-ancla con el núcleo del texto
            # ('Racionamiento de gas', 'Inversión en energía solar').
            if frase:
                toks_f = [_normaliza_token(w) for w in frase.split()]
                tiene_hecho = any(t in _NUCLEOS_ACONTECIMIENTO or _es_cabeza_subtema_valida(t)
                                  for t in toks_f if t)
                if not tiene_hecho:
                    nucleo = self._nucleo_hecho_de_texto(f"{t} {verbo_saltado or ''}")
                    if nucleo and _normaliza_token(nucleo) not in set(toks_f):
                        frase = _sanear_frase_nominal(
                            f"{nucleo} {_preposicion_de(frase.split()[0])} {frase}")
            # La cabeza no puede ser un adjetivo/adverbio suelto ('Per capita de pollo'):
            # se antepone el núcleo de hecho detectado en el título.
            if frase:
                cabeza = unidecode(frase.split()[0].lower())
                necesita_nucleo = (
                    cabeza in _CONECTORES_ETIQUETA
                    or cabeza in {"per", "mas", "menos", "muy", "solo", "casi", "tras"}
                    or _parece_adjetivo_es(cabeza)
                    or not _es_cabeza_subtema_valida(cabeza)
                )
                if necesita_nucleo:
                    nucleo = self._nucleo_hecho_de_texto(f"{t} {verbo_saltado or ''}")
                    if nucleo and _normaliza_token(nucleo) not in {
                            _normaliza_token(w) for w in frase.split()}:
                        frase = _sanear_frase_nominal(f"{nucleo} de {frase}")
            # Locuciones adverbiales que no aportan asunto por sí solas.
            if frase and re.fullmatch(r"(?i)\w+\s+de\s+per\s+capita", unidecode(frase)):
                frase = _sanear_frase_nominal(frase.split()[0])
            if (_frase_esta_completa(frase) and len(frase.split()) >= 2
                    and not _es_nombre_o_fragmento_marca(frase, self.marca, self.aliases)
                    and not _es_etiqueta_generica(frase)):
                return capitalizar_etiqueta(frase)
        return ""

    def _nucleo_hecho_de_texto(self, texto):
        """Sustantivo de hecho ('Investigación', 'Consumo', 'Alza') presente en el texto."""
        tnorm = unidecode(str(texto or "").lower())
        return next((nom for pat, nom in self._NUCLEOS_HECHO if re.search(pat, tnorm)), None)

    def _palabra_clave_mas_frecuente(self, titulos, resumenes):
        """Sustantivos de contenido más frecuentes, unidos como FRASE NOMINAL real.

        Antes devolvía el pegote 'precio pollo' / 'consumo pollo': dos keywords
        pegadas, que es justo lo que el usuario rechaza ("subtemas que son solo
        palabras clave"). Ahora se antepone un núcleo de hecho y se inserta la
        preposición, produciendo 'Precio del pollo' o 'Alza del precio del pollo'.
        """
        excluir = STOPWORDS_ES | _VERBOS_LEAD_SUBTEMA | _CARGOS_SUBTEMA | _TRAILING_INCOMPLETE | {
            "noticia", "noticias", "informe", "informacion", "comunicado", "nota",
            "colombia", "pais", "nacional", "regional", "local", "sector", "empresa",
            "entidad", "autoridad", "gobierno", "alcaldia", "gobernacion", "ministerio",
            "nuevo", "nueva", "nuevos", "nuevas",
        }
        for n in [self.marca] + [a for a in (self.aliases or []) if a]:
            n = str(n).strip()
            if n:
                excluir = excluir | set(_normalizar_mencion(n).split())
        cnt = Counter()
        origen = {}
        for t in titulos[:5]:
            for w in re.findall(r"[A-Za-zÁÉÍÓÚÑÜáéíóúñü]+", str(t)):
                norm = _normaliza_token(w)
                if len(norm) >= 4 and norm not in excluir and not _RE_VERBO_SUBTEMA.search(norm):
                    cnt[norm] += 1
                    origen.setdefault(norm, w)
        if not cnt and resumenes:
            for r in resumenes[:3]:
                for w in re.findall(r"[A-Za-zÁÉÍÓÚÑÜáéíóúñü]+", str(r)):
                    norm = _normaliza_token(w)
                    if len(norm) >= 4 and norm not in excluir and not _RE_VERBO_SUBTEMA.search(norm):
                        cnt[norm] += 1
                        origen.setdefault(norm, w)
        if not cnt:
            return ''
        tops = [origen[n] for n, _ in cnt.most_common(3)]
        return self._frase_nominal_desde_tokens(tops, titulos, resumenes)

    # ── Construcción de FRASE NOMINAL (evita pegotes de keywords) ───────────────
    # Núcleo de hecho detectado en el texto → se usa como cabeza de la etiqueta.
    _NUCLEOS_HECHO = [
        (r"\b(alza|aumento|incremento|encarec|subida|sube|subio)\w*", "Alza"),
        (r"\b(caida|baja|reduccion|descenso|disminucion)\w*", "Reducción"),
        (r"\b(precio|tarifa|costo)\w*", "Precio"),
        (r"\b(consumo|demanda)\w*", "Consumo"),
        (r"\b(venta|ventas|comercializacion)\w*", "Ventas"),
        (r"\b(exportacion|exporta)\w*", "Exportaciones"),
        (r"\b(importacion|importa)\w*", "Importaciones"),
        (r"\b(produccion|produce|produjo)\w*", "Producción"),
        (r"\b(inversion|invierte|invirtio)\w*", "Inversión"),
        (r"\b(lanzamiento|lanza|estreno|presentacion)\w*", "Lanzamiento"),
        (r"\b(convenio|alianza|acuerdo|pacto|cooperacion)\w*", "Convenio"),
        (r"\b(premio|galardon|reconocimiento|distincion)\w*", "Reconocimiento"),
        (r"\b(investigacion|investiga|indaga|denuncia|demanda|sancion|multa)\w*", "Investigación"),
        (r"\b(nombramiento|designacion|posesion|nombra|designa)\w*", "Nombramiento"),
        (r"\b(renuncia|dimision|dimite)\w*", "Renuncia"),
        (r"\b(apertura|inauguracion|inaugura|abre)\w*", "Apertura"),
        (r"\b(cierre|suspension|clausura)\w*", "Cierre"),
        (r"\b(ampliacion|expansion|amplia)\w*", "Ampliación"),
        (r"\b(construccion|obra|construye)\w*", "Construcción"),
        (r"\b(reforma|regulacion|decreto|ley)\w*", "Reforma"),
        (r"\b(racionamiento|raciona|apagon|desabastecimiento|escasez)\w*", "Racionamiento"),
        (r"\b(crisis|emergencia|desastre|sismo|inundacion)\w*", "Crisis"),
        (r"\b(campana|publicidad|patrocinio)\w*", "Campaña"),
        (r"\b(foro|congreso|cumbre|seminario|encuentro|feria)\w*", "Foro"),
        (r"\b(proyecto|iniciativa|programa|plan)\w*", "Proyecto"),
        (r"\b(balance|resultado|utilidad|ingreso)\w*", "Balance"),
        (r"\b(estudio|informe|encuesta|analisis)\w*", "Estudio"),
        (r"\b(donacion|solidaridad|ayuda|apoyo)\w*", "Apoyo"),
        (r"\b(capacitacion|formacion|curso|taller)\w*", "Capacitación"),
        (r"\b(vacunacion|brote|contagio|virus|gripe|influenza)\w*", "Situación sanitaria"),
        (r"\b(empleo|contratacion|trabajador|nomina)\w*", "Empleo"),
    ]

    def _frase_nominal_desde_tokens(self, tokens, titulos=None, resumenes=None):
        """Convierte tokens de contenido en una frase nominal con preposición.
        'precio', 'pollo' -> 'Alza del precio del pollo'. NUNCA devuelve dos
        sustantivos pegados sin nexo, ni un adjetivo tras 'de'."""
        toks = [str(t).strip() for t in (tokens or []) if str(t).strip()]
        if not toks:
            return ''
        texto = " ".join(str(x) for x in (list(titulos or [])[:5] + list(resumenes or [])[:2]))
        tnorm = unidecode(texto.lower())
        nucleo = next((nom for pat, nom in self._NUCLEOS_HECHO if re.search(pat, tnorm)), None)

        vistos, limpios = set(), []
        for t in toks:
            k = _normaliza_token(t)
            if not k or k in vistos:
                continue
            vistos.add(k)
            limpios.append(t.lower())
        if not limpios:
            return ''

        # 1) Si uno de los tokens centrales (top-2) ya es un sustantivo de evento
        #    válido ('exportaciones', 'convenio'), ese es el mejor núcleo: viene del
        #    texto. 2) Si no, el núcleo léxico detectado. 3) Si no, el primer token.
        cabeza, resto = None, list(limpios)
        for cand in limpios[:2]:
            if _es_cabeza_subtema_valida(cand):
                cabeza = cand.capitalize()
                resto = [x for x in limpios if _normaliza_token(x) != _normaliza_token(cand)]
                break
        if cabeza is None and nucleo:
            if _normaliza_token(nucleo) in {_normaliza_token(x) for x in limpios}:
                cabeza = nucleo
                resto = [x for x in limpios if _normaliza_token(x) != _normaliza_token(nucleo)]
            else:
                cabeza, resto = nucleo, limpios
        if cabeza is None:
            cabeza, resto = limpios[0].capitalize(), limpios[1:]

        frase = cabeza
        usados = 0
        for comp in resto:
            if usados >= 2:
                break
            if _parece_adjetivo_es(comp) and usados >= 1:
                # Adjetivo: se pega al sustantivo anterior, nunca tras 'de'.
                cand = f"{frase} {comp}"
            else:
                cand = f"{frase} {_preposicion_de(comp)} {comp}"
            if len(cand.split()) > MAX_PALABRAS_SUBTEMA:
                break
            frase = cand
            usados += 1
        frase = _sanear_frase_nominal(frase)
        if (len(frase.split()) >= 2 and _frase_esta_completa(frase)
                and not _es_etiqueta_generica(frase)
                and not _es_nombre_o_fragmento_marca(frase, self.marca, self.aliases)):
            return capitalizar_etiqueta(frase)
        return ''

    def _derivar_ultimo_recurso(self, titulos, resumenes):
        '''Última red absoluta: SIEMPRE una etiqueta con palabras reales del texto
        (Título / Resumen-Aclaración). Nunca 'Sin tema' mientras exista texto.'''
        fuentes = [str(t).strip() for t in (titulos or [])
                   if str(t).strip() and str(t).strip().lower() != 'nan']
        if not fuentes and resumenes:
            fuentes = [str(r).strip() for r in (resumenes or []) if str(r).strip()]
        for s in fuentes:
            if len(s) < 4:
                continue
            s2 = s
            for n in [self.marca] + [a for a in (self.aliases or []) if a]:
                n = str(n).strip()
                if n:
                    s2 = re.sub(r'\b' + re.escape(n) + r'\b', ' ', s2, flags=re.IGNORECASE)
            toks = [w for w in s2.split() if w]
            i = 0
            while i < len(toks):
                tn = unidecode(toks[i].lower().rstrip('.,;:!?'))
                if (tn[:1].isdigit()
                        or tn in _ARTICULOS_SUBTEMA
                        or tn in _CONECTORES_ETIQUETA
                        or tn in _VERBOS_LEAD_SUBTEMA
                        or tn in _CARGOS_SUBTEMA
                        or tn in _TOKENS_DEBILES_SUBTEMA_FALLBACK
                        or tn in {'nuevo', 'nueva', 'nuevos', 'nuevas', 'gran', 'grande',
                                  'grandes', 'este', 'esta', 'estos', 'estas'}
                        or _PATRON_TITULAR.match(toks[i])
                        or _RE_VERBO_SUBTEMA.search(tn)):
                    i += 1
                else:
                    break
            resto = toks[i:]
            if resto:
                frase = _sanear_frase_nominal(' '.join(resto))
                if (_frase_esta_completa(frase) and len(frase.split()) >= 2
                        and not _es_etiqueta_generica(frase)
                        and not _es_nombre_o_fragmento_marca(frase, self.marca, self.aliases)):
                    return capitalizar_etiqueta(frase)
            # Título reducido a nada útil (solo marca) -> se usa el título mismo.
            frase2 = _sanear_frase_nominal(s, max_palabras=4)
            if _frase_esta_completa(frase2) and len(frase2.split()) >= 2:
                return capitalizar_etiqueta(frase2)
        return ''

    def _fallback(self, titulos, resumenes=None):
        """Última red: SIEMPRE deriva una etiqueta del texto real (Título / Resumen-Aclaración).
        Nunca devuelve rótulos genéricos tipo 'Cobertura de información relevante'.
        Cada candidato se valida ANTES de devolverse (estructura de subtema real,
        sin titulares copiados, sin cabezas inventadas). Para los casos del cliente,
        la oración del RESUMEN (donde suele vivivir la mención) tiene prioridad."""
        titulos = [t for t in (titulos or []) if t is not None and str(t).strip() and str(t).strip().lower() != 'nan']
        resumenes = [r for r in (resumenes or []) if r is not None and str(r).strip()]
        fuentes = list(titulos) + list(resumenes)

        def _sirve(et):
            return (et and _validar_estructura_subtema(et)
                    and not _es_etiqueta_generica(et)
                    and not _es_nombre_o_fragmento_marca(et, self.marca, self.aliases)
                    and _head_anclada(et, fuentes))

        # 1) Extracción LLM estricta: frase armada SOLO con palabras del texto.
        et = self._extraer_desde_texto(titulos, resumenes)
        if _sirve(et):
            return et
        # 1b) La oración del RESUMEN es la fuente del hecho del cliente: derivar
        #     de ella directamente da subtemas como 'Atención de urgencia' en lugar
        #     de recortar el titular.
        et = self._derivar_desde_texto_nominal(resumenes or titulos, fuentes, es_resumen=True)
        if _sirve(et):
            return et

        # 2) Frase nominal derivada del título más representativo (100% grounded).
        et = self._derivar_desde_titulos(titulos)
        if _sirve(et):
            return et

        # 3) Acción detectada + palabras clave del texto (100% grounded).
        et = self._construir_frase_accion(titulos, resumenes)
        if _sirve(et):
            return et

        # 4) Palabras de contenido más frecuentes (aún grounded, exige 2+ palabras reales).
        et = self._palabra_clave_mas_frecuente(titulos, resumenes)
        if _sirve(et):
            return et

        # 5) Última red: oración que menciona al cliente en el resumen/título, si
        #    existe, recortada a frase nominal válida; sino fragmento del texto.
        return self._derivar_ultimo_recurso(titulos, resumenes)

    def _consolidar_sinonimos_llm(self, subtemas_unicos):
        if len(subtemas_unicos) <= 1:
            return {s: s for s in subtemas_unicos}
            
        prompt = (
            "Eres un analista de datos. Tienes la siguiente lista de subtemas periodísticos:\n"
            f"{', '.join(subtemas_unicos)}\n\n"
            "Tu tarea es encontrar SUBTEMAS SINÓNIMOS que signifiquen exactamente lo mismo "
            "(aunque usen palabras ligeramente distintas) y unificarlos bajo el nombre más claro y representativo.\n"
            "REGLAS:\n"
            "1. NO fusiones temas que sean distintos (ej. 'Inversión en vías' y 'Mantenimiento de vías' son distintos).\n"
            "2. SÍ fusiona sinónimos (ej. 'Lanzamiento de plataforma web' y 'Estreno de portal digital').\n"
            "3. Devuelve un objeto JSON donde las claves sean los subtemas originales y el valor sea el subtema unificado.\n\n"
            'Ejemplo de salida:\n'
            '{"Tendencias de consumo de pollo": "Tendencias de consumo de pollo", "Hábitos de compra de aves": "Tendencias de consumo de pollo"}'
        )
        try:
            resp = call_with_retries(
                openai.ChatCompletion.create,
                model=OPENAI_MODEL_CLASIFICACION,
                messages=[{"role": "user", "content": prompt}],
                max_tokens=1000,
                temperature=0.0,
                response_format={"type": "json_object"}
            )
            return json.loads(resp.choices[0].message.content)
        except:
            return {s: s for s in subtemas_unicos}

    def procesar_lote(self, col, pbar, res_puros, tit_puros):
        textos   = col.tolist()
        titulos  = tit_puros.tolist()
        resumenes = res_puros.tolist()
        n = len(textos)

        self._umbrales = _umbrales_adaptativos(n)
        u = self._umbrales
        st.caption(
            f"📐 Corpus: **{n}** noticias · Umbral subtema: **{u['subtema']}** · "
            f"Sim mínima: **{u['sim_minima_agrupacion']}**"
        )

        et = textos  # base de agrupación: Título+Contexto+Resumen (ya en `col` = _txt)

        pbar.progress(0.05, "Fase 1 · Idénticas...")
        dsu = DSU(n)
        self._paso1(titulos, resumenes, dsu)
        
        pbar.progress(0.12, "Fase 2 · Títulos...")
        self._paso2(titulos, dsu)

        pbar.progress(0.18, "Embeddings...")
        ae = get_embeddings_batch(et)

        if u['usar_paso2b']:
            pbar.progress(0.15, "Fase 2b · Keywords raras (con validación semántica)...")
            self._paso2b_keywords(titulos, dsu, ae)

        pbar.progress(0.20, "Fase 3 · Clustering...")
        self._paso3(et, ae, dsu, pbar, 0.20)

        gf = dsu.grupos(n)
        ng = len(gf)
        pbar.progress(0.55, f"Fase 4 · Etiquetando {ng} grupos...")
        mapa = {}
        sg = sorted(gf.items(), key=lambda x: -len(x[1]))  # grupos grandes primero (revert: lógica estable)
        subtemas_aprobados = []
        textos_por_subtema_aprobado = defaultdict(list)

        def _generar_etiqueta_segura(idxs):
            # Cada miembro del grupo DSU comparte la etiqueta; reutiliza los subtemas ya
            # aprobados para que noticias equivalentes compartan EXACTAMENTE el mismo subtema.
            sample = idxs[:MAX_GRUPO_ETIQUETA]
            textos_grp = [textos[i] for i in sample]
            titulos_grp = [titulos[i] for i in sample]
            resumenes_grp = [resumenes[i] for i in sample]
            etiqueta = self._generar_etiqueta(
                textos_grp, titulos_grp, resumenes_grp,
                subtemas_existentes=subtemas_aprobados
            )
            if etiqueta in textos_por_subtema_aprobado:
                previos = textos_por_subtema_aprobado.get(etiqueta, [])
                if not _grupos_contenido_compatibles(
                    textos_grp, previos, etiqueta, etiqueta,
                    min_sim=max(u['sim_minima_agrupacion'], 0.88), min_overlap=0.24,
                ):
                    rechazada = etiqueta
                    etiqueta = self._generar_etiqueta(
                        textos_grp, titulos_grp, resumenes_grp,
                        subtemas_existentes=subtemas_aprobados,
                        evitar_etiqueta=rechazada
                    )
                    if etiqueta in textos_por_subtema_aprobado:
                        previos2 = textos_por_subtema_aprobado.get(etiqueta, [])
                        if not _grupos_contenido_compatibles(
                            textos_grp, previos2, etiqueta, etiqueta,
                            min_sim=max(u['sim_minima_agrupacion'], 0.88), min_overlap=0.24,
                        ):
                            etiqueta = capitalizar_etiqueta(self._fallback(titulos_grp, resumenes_grp))
            if etiqueta not in subtemas_aprobados:
                subtemas_aprobados.append(etiqueta)
            textos_por_subtema_aprobado[etiqueta].extend(textos_grp)
            return etiqueta

        for k, (lid, idxs) in enumerate(sg):
            if k % 10 == 0: pbar.progress(0.55 + 0.25 * (k / max(ng, 1)), f"Etiquetando {k + 1}/{ng}...")
            e = _generar_etiqueta_segura(idxs)
            for i in idxs: mapa[i] = e

        subtemas = [mapa.get(i, "Varios") for i in range(n)]



        pbar.progress(0.80, "Fase 4b · Coherencia (sin reasignar)...")
        # 0.35 cosine-to-label is not event membership. Jumping rows onto
        # another Subtema (or minting a new phrase) over-grouped and paraphrased.

        pbar.progress(0.86, "Fase 5 · Sin fusión cruzada de etiquetas...")
        # Skip corpus-wide dedup_labels / _fusionar_subtemas_semanticos.
        # Those glued distinct events that shared a 5-word paraphrase.

        pbar.progress(0.90, "Fase 6 · Consistencia...")
        subtemas = self._consistencia(subtemas, ae, pbar, u)

        indices_reclass = [i for i, s in enumerate(subtemas) if s == "_RECLASSIFICAR"]
        if indices_reclass:
            pbar.progress(0.93, f"Fase 6b · Reclasificando...")
            for i in indices_reclass:
                et_ind = self._generar_etiqueta([textos[i]], [titulos[i]], [resumenes[i]], subtemas_existentes=subtemas_aprobados)
                subtemas[i] = capitalizar_etiqueta(et_ind)
                if et_ind not in subtemas_aprobados: subtemas_aprobados.append(et_ind)

        pbar.progress(0.93, "Fase 7 · Completitud...")
        subtemas = self._validar_completitud_final(subtemas, textos, titulos, resumenes)
        pbar.progress(0.95, "Fase 7b · Depurando rótulos genéricos...")
        por_etiqueta = defaultdict(list)
        for i, s in enumerate(subtemas):
            por_etiqueta[s].append(i)
        for s in [s for s in por_etiqueta if _es_etiqueta_generica(s)]:
            idxs = por_etiqueta[s]
            muestra_t = [titulos[i] for i in idxs[:MAX_GRUPO_ETIQUETA]]
            muestra_r = [resumenes[i] for i in idxs[:3]]
            nueva = self._fallback(muestra_t, muestra_r)
            nueva = capitalizar_etiqueta(nueva) if nueva else "Varios"
            for i in idxs:
                subtemas[i] = nueva

        pbar.progress(0.96, 'Fase 7c · Garantizando subtema para cada noticia...')
        por_etq = defaultdict(list)
        for i, s in enumerate(subtemas):
            por_etq[s].append(i)
        for s in [s for s in por_etq]:
            n = ' '.join(unidecode(str(s).lower()).split()) if isinstance(s, str) else ''
            if n not in _PLACEHOLDER_SUBTEMA:
                continue
            idxs = por_etq[s]
            muestra_t = [titulos[i] for i in idxs[:MAX_GRUPO_ETIQUETA]]
            muestra_r = [resumenes[i] for i in idxs[:3]]
            et_n = self._fallback(muestra_t, muestra_r)
            if et_n and et_n.strip().lower() not in ('sin tema', 'varios'):
                nueva = capitalizar_etiqueta(et_n)
            else:
                txt_i = ''
                for i in idxs:
                    t_i = titulos[i]
                    r_i = resumenes[i]
                    tx_i = textos[i]
                    if t_i is not None and str(t_i).strip() and str(t_i).strip().lower() != 'nan':
                        txt_i = str(t_i).strip()
                        break
                    if r_i is not None and str(r_i).strip() and str(r_i).strip().lower() != 'nan':
                        txt_i = str(r_i).strip()
                        break
                    if tx_i is not None and str(tx_i).strip() and str(tx_i).strip().lower() != 'nan':
                        txt_i = str(tx_i).strip()
                        break
                nueva = capitalizar_etiqueta(_recortar_frase_completa(txt_i, max_palabras=4)) if txt_i else 'Sin tema'
            for i in idxs:
                subtemas[i] = nueva

        pbar.progress(0.97, "Fase 8 · Sin dedup ni sinónimos cruzados...")
        subtemas = [capitalizar_etiqueta(s) for s in subtemas]
        nf = len(set(subtemas))
        pbar.progress(1.0, f"{nf} subtemas")
        st.info(f"Subtemas: **{nf}** · Grupos originales: **{ng}**")
        return subtemas

    def _validar_completitud_final(self, subtemas, textos, titulos, resumenes):
        por_subtema = defaultdict(list)
        for i, s in enumerate(subtemas): por_subtema[s].append(i)
        resultado = list(subtemas)
        for sub, idxs in por_subtema.items():
            if _frase_esta_completa(sub): continue
            recortada = _recortar_frase_completa(sub)
            if _frase_esta_completa(recortada) and len(recortada.split()) >= 2:
                for i in idxs: resultado[i] = capitalizar_etiqueta(recortada)
                continue
            tit_grp = [titulos[i] for i in idxs[:6]]
            res_grp = [resumenes[i] for i in idxs[:3]]
            nueva = _validar_etiqueta_completa(
                sub, titulos_grp=tit_grp, resumenes_grp=res_grp,
                marca=self.marca, aliases=self.aliases, fallback_fn=lambda tl: self._fallback(tl, res_grp)
            )
            for i in idxs: resultado[i] = capitalizar_etiqueta(nueva)
        return resultado

    def _consistencia(self, subtemas, ae, pbar, umbrales=None):
        min_sub = umbrales.get('min_pertenencia_subtema', UMBRAL_MIN_PERTENENCIA_SUBTEMA)
        ps = defaultdict(list)
        for i, s in enumerate(subtemas): ps[s].append(i)
        r = list(subtemas)
        centroids = {}
        for sub, idxs in ps.items():
            vecs = [ae[i] for i in idxs if ae[i] is not None]
            if vecs: centroids[sub] = np.mean(vecs, axis=0)
        for sub in [s for s in centroids if len(ps[s]) >= 3]:
            idxs = ps[sub]
            if sub.lower() in ("sin tema", "varios") or len(idxs) < 3: continue
            vi = [(i, ae[i]) for i in idxs if ae[i] is not None]
            if len(vi) < 3: continue
            v_i, v_v = zip(*vi)
            M = np.array(v_v)
            sims = cosine_similarity(M, centroids[sub].reshape(1, -1)).flatten()
            thr = max(0.60, np.mean(sims) - 2 * np.std(sims))
            for k, (oi, sv) in enumerate(zip(v_i, sims)):
                if sv >= thr: continue
                bs, bsim = sub, sv
                emb = ae[oi]
                for os_, oc in centroids.items():
                    if os_ == sub: continue
                    s2 = cosine_similarity(np.array(emb).reshape(1, -1), oc.reshape(1, -1))[0][0]
                    if s2 > bsim and s2 > 0.75: bsim = s2; bs = os_
                if bs != sub: r[oi] = bs
                elif sv < min_sub: r[oi] = "_RECLASSIFICAR"
        return r

# ======================================
# TEMAS  
# ======================================
def _construir_representacion_grupo(subtema, textos_grupo, max_textos=30):
    palabras = []
    for t in textos_grupo[:max_textos]:
        for w in string_norm_label(str(t)).split():
            if len(w) > 3: palabras.append(w)
    kw_str = " ".join(w for w, _ in Counter(palabras).most_common(12))
    return f"{subtema}. {subtema}. {kw_str}"[:500]

def _validar_estructura_tema(tema: str) -> bool:
    if not tema or len(tema.split()) < 2: return False
    if len(tema.split()) > 5: return False
    if re.match(r'^[0-9]', tema): return False
    num_palabras = re.compile(
        r'^(uno|dos|tres|cuatro|cinco|seis|siete|ocho|nueve|diez|'
        r'once|doce|veinte|cien|varios|cada)', re.IGNORECASE
    )
    if num_palabras.match(tema): return False
    if _PATRON_TITULAR.match(tema): return False
    if _PATRON_ESTADO.search(tema): return False
    # Un TEMA tampoco puede ser un rótulo de relleno ('Destacados del sector').
    if _es_etiqueta_generica(tema): return False
    genericos = {
        "economia", "politica", "tecnologia", "seguridad", "justicia",
        "actualidad", "nacional", "internacional", "empresas", "sociedad",
        "negocios", "informacion", "noticias", "varios", "general",
    }
    if string_norm_label(tema) in genericos: return False
    return True


# ── Taxonomía editorial de respaldo: garantiza que NINGUNA fila quede sin Tema ──
# Cada entrada es (Tema, patrón de disparo). Se evalúa contra el subtema + el
# título/contexto, así el Tema queda derivado del texto y no de un default fijo.
_TAXONOMIA_TEMAS = [
    ("Resultados y desempeño financiero",
     r"\b(utilidad(es)?|ingreso(s)?|ganancia(s)?|perdida(s)?|balance|ebitda|"
     r"facturacion|ventas|rentabilidad|dividendo(s)?|trimestre|estado(s) financiero)\b"),
    ("Inversión y expansión",
     r"\b(inversion(es)?|invertira|expansion|nueva(s)? sede(s)?|apertura|ampliacion|"
     r"planta|adquisicion|compra de|fusion|capitalizacion|infraestructura)\b"),
    ("Lanzamiento de productos y servicios",
     r"\b(lanzamiento|nuevo producto|nueva linea|portafolio|servicio nuevo|"
     r"estreno|presentacion de|innovacion|prototipo|modelo)\b"),
    ("Precios y comportamiento del mercado",
     r"\b(precio(s)?|tarifa(s)?|costo(s)?|inflacion|alza|encarecimiento|"
     r"demanda|oferta|consumo|mercado|competencia|exportacion(es)?|importacion(es)?)\b"),
    ("Regulación y política pública",
     r"\b(ley|decreto|reforma|regulacion|norma(tiva)?|proyecto de ley|congreso|"
     r"ministerio|superintendencia|resolucion|arancel|impuesto|politica publica)\b"),
    ("Litigios, investigaciones y sanciones",
     r"\b(demanda|denuncia|investigacion|sancion|multa|fiscalia|procuraduria|"
     r"contraloria|juzgado|tribunal|condena|imputacion|querella|fallo|litigio)\b"),
    ("Sostenibilidad y medio ambiente",
     r"\b(sostenibilidad|ambiental|medio ambiente|emisiones|carbono|reciclaje|"
     r"energia (limpia|renovable)|agua|residuos|deforestacion|huella)\b"),
    ("Responsabilidad social y comunidad",
     r"\b(donacion|solidaridad|ayuda humanitaria|voluntariado|fundacion|"
     r"responsabilidad social|comunidad(es)?|damnificado(s)?|beneficiario(s)?|obra social)\b"),
    ("Talento humano y gestión interna",
     r"\b(empleo(s)?|contratacion(es)?|nomina|trabajador(es)?|sindicato|huelga|"
     r"despido(s)?|capacitacion|bienestar laboral|clima laboral|vacante(s)?)\b"),
    ("Nombramientos y gobierno corporativo",
     r"\b(nombramiento|designacion|posesion|renuncia|dimision|junta directiva|"
     r"presidente de|gerente|ceo|relevo|sucesion|asamblea de accionistas)\b"),
    ("Reconocimientos y premios",
     r"\b(premio(s)?|galardon|reconocimiento|distincion|condecoracion|ranking|"
     r"certificacion|acreditacion|honoris causa|mejor(es)? empresa)\b"),
    ("Alianzas y convenios",
     r"\b(convenio(s)?|alianza(s)?|acuerdo(s)?|cooperacion|memorando|"
     r"colaboracion|pacto|union temporal|consorcio)\b"),
    ("Tecnología e innovación digital",
     r"\b(tecnologia|digital(izacion)?|inteligencia artificial|software|plataforma|"
     r"app|aplicacion|ciberseguridad|datos|automatizacion|transformacion digital)\b"),
    ("Salud pública y bioseguridad",
     r"\b(salud|sanitario|epidemia|pandemia|virus|gripe aviar|influenza|"
     r"vacunacion|bioseguridad|brote|contagio|hospital|eps)\b"),
    ("Infraestructura y operaciones",
     r"\b(obra(s)?|via(s)?|carretera|puerto|aeropuerto|terminal|construccion|"
     r"mantenimiento|logistica|transporte|cadena de suministro|operacion(es)?)\b"),
    ("Crisis y emergencias",
     r"\b(crisis|emergencia|desastre|sismo|terremoto|inundacion|incendio|"
     r"derrame|accidente|falla(s)?|colapso|paro|bloqueo)\b"),
    ("Educación y formación",
     r"\b(educacion|universidad|colegio|estudiante(s)?|beca(s)?|academico|"
     r"formacion|curso(s)?|diplomado|matricula|graduacion|investigacion academica)\b"),
    ("Eventos y participación institucional",
     r"\b(foro|congreso|cumbre|feria|seminario|conversatorio|taller|encuentro|"
     r"evento|jornada|panel|conferencia|rueda de prensa)\b"),
    ("Comunicación y posicionamiento de marca",
     r"\b(campana|publicidad|comunicado|vocero|marca|patrocinio|"
     r"posicionamiento|estrategia de comunicacion|imagen corporativa)\b"),
    # ── Categorías transversales: cubren clientes de banca, energía, minería,
    # transporte, cultura, migración, deporte y sector público. Se añaden al final
    # para que los patrones más específicos de arriba tengan prioridad.
    ("Política monetaria y tasas",
     r"\b(tasa(s)? de interes|tasa de usura|banco de la republica|banco central|"
     r"politica monetaria|devaluacion|revaluacion|tipo de cambio|inflacion|"
     r"emisor|encaje|liquidez)\b"),
    ("Servicios públicos y tarifas",
     r"\b(racionamiento|apagon|tarifa(s)? de|servicio publico|acueducto|"
     r"alcantarillado|energia electrica|gas natural|embalse(s)?|"
     r"factura(cion)? de servicios|suministro|desabastecimiento)\b"),
    ("Producción y explotación de recursos",
     r"\b(mina|minera|mineria|carbon|niquel|oro|petroleo|crudo|gas|refineria|"
     r"pozo(s)?|barriles|yacimiento|regalias|extraccion|explotacion minera|"
     r"produccion agricola|cosecha|cultivo(s)?|ganaderia|pesca)\b"),
    ("Movilidad y transporte",
     r"\b(vuelo(s)?|aerolinea|pasajero(s)?|peaje(s)?|trafico|metro|"
     r"transmilenio|bus(es)?|taxi(s)?|tren|ferrocarril|navegacion|"
     r"movilidad|congestion|ruta(s)? aerea)\b"),
    ("Conectividad y telecomunicaciones",
     r"\b(espectro|5g|4g|internet|banda ancha|fibra optica|telefonia|"
     r"cobertura movil|operador movil|antena(s)?|telecomunicaciones)\b"),
    ("Cultura, patrimonio y deporte",
     r"\b(teatro|museo|patrimonio|monumento|festival|carnaval|concierto|"
     r"biblioteca|artista(s)?|obra cultural|deporte|equipo|torneo|"
     r"seleccion|estadio|liga|jugador(es)?|tecnico|campeonato)\b"),
    ("Migración, derechos y grupos poblacionales",
     r"\b(migrante(s)?|migracion|refugiado(s)?|venezolano(s)?|desplazado(s)?|"
     r"victima(s)?|indigena(s)?|afro|mujer(es)?|genero|discapacidad|"
     r"derechos humanos|regularizacion)\b"),
    ("Seguridad y orden público",
     r"\b(homicidio(s)?|hurto(s)?|robo(s)?|extorsion|secuestro|atentado|"
     r"disidencia(s)?|guerrilla|banda(s)? criminal|narcotrafico|incautacion|"
     r"captura(s)?|policia|ejercito|fuerza publica|orden publico|violencia)\b"),
    ("Gestión pública y presupuesto",
     r"\b(presupuesto|regalias|contratacion publica|licitacion publica|"
     r"plan de desarrollo|ejecucion presupuestal|deficit|superavit|"
     r"gasto publico|recaudo|dian|hacienda|concejo|asamblea)\b"),
    ("Vivienda y desarrollo urbano",
     r"\b(vivienda(s)?|urbanismo|pot|plan de ordenamiento|predio(s)?|"
     r"barrio(s)?|urbanizacion|espacio publico|parque(s)?|renovacion urbana)\b"),
]


def _tema_por_taxonomia(subtema: str, texto_apoyo: str = "") -> str:
    """Deriva un Tema editorial a partir del léxico REAL del subtema/texto.
    Es el respaldo cuando el LLM no entrega un tema válido: evita 'Sin tema' y
    evita que el Tema sea una copia del Subtema.

    La CABEZA del subtema manda: 'Convenio con universidad para investigación'
    es 'Alianzas y convenios', no 'Litigios' — aunque contenga 'investigación'.
    """
    sub = unidecode(str(subtema or "").lower()).strip()
    # 1) Prioridad absoluta: las 2 primeras palabras de contenido del subtema.
    cabeza_toks = [t for t in re.findall(r"[a-z]+", sub)
                   if t not in _CONECTORES_ETIQUETA and len(t) >= 4][:2]
    if cabeza_toks:
        cabeza = " ".join(cabeza_toks)
        for tema, patron in _TAXONOMIA_TEMAS:
            if re.search(patron, cabeza, re.IGNORECASE):
                return tema
    # 2) Resto del subtema.
    if sub:
        for tema, patron in _TAXONOMIA_TEMAS:
            if re.search(patron, sub, re.IGNORECASE):
                return tema
    # 3) Texto de apoyo (título/contexto/resumen).
    base = unidecode(str(texto_apoyo or "").lower())
    if base:
        for tema, patron in _TAXONOMIA_TEMAS:
            if re.search(patron, base, re.IGNORECASE):
                return tema
    return ""


def _asegurar_tema_valido(temas, subtemas, textos=None):
    """Garantía dura de la jerarquía Tema→Subtema, fila por fila:
      1. Un Tema vacío/'Sin tema'/genérico se reemplaza por taxonomía léxica.
      2. Un Tema igual (o casi) al Subtema se sustituye por su categoría madre,
         porque un Tema que repite el Subtema no aporta jerarquía.
      3. Si nada aplica, se agrupa por el hecho del subtema con un tema derivado.
    Nunca deja una fila con Subtema y sin Tema (defecto reportado)."""
    out = []
    n = len(subtemas)
    textos = list(textos) if textos is not None else [""] * n
    if len(textos) < n:
        textos = list(textos) + [""] * (n - len(textos))
    for i in range(n):
        sub = str(subtemas[i] or "").strip()
        tema = str(temas[i] or "").strip() if i < len(temas) else ""
        apoyo = str(textos[i] or "")
        invalido = (
            not tema
            or tema.lower() in ("sin tema", "varios", "n/a", "nan", "-", "none")
            or _es_etiqueta_generica(tema)
        )
        if not invalido and _tema_es_igual_a_subtema(tema, [sub]):
            invalido = True          # Tema == Subtema: no aporta jerarquía
        if invalido:
            nuevo = _tema_por_taxonomia(sub, apoyo)
            if not nuevo:
                # Sin señal léxica en la taxonomía: se abstrae por el TIPO DE HECHO
                # que encabeza el subtema ('Restauración del teatro' -> 'Proyectos
                # y obras'). Es una categoría real, no un recorte del subtema.
                nuevo = _tema_por_cabeza_de_hecho(sub)
            if not nuevo:
                # Última red: núcleo del subtema como categoría madre (texto real).
                nucleo = _recortar_frase_completa(sub, max_palabras=3)
                nuevo = capitalizar_etiqueta(nucleo) if nucleo and not _es_etiqueta_generica(nucleo) else ""
            if nuevo:
                tema = nuevo
            else:
                tema = capitalizar_etiqueta(_recortar_frase_completa(sub, max_palabras=3))
        out.append(capitalizar_etiqueta(tema) if tema else "Sin tema")
    return out


# Abstracción por TIPO DE HECHO: mapea la cabeza del subtema a una categoría
# editorial. Independiente del sector del cliente — un 'convenio' es una alianza
# lo firme un banco, una alcaldía o una minera.
_TEMA_POR_CABEZA = [
    ({"convenio", "alianza", "acuerdo", "pacto", "cooperacion", "colaboracion",
      "memorando", "consorcio", "union"}, "Alianzas y convenios"),
    ({"inversion", "expansion", "ampliacion", "adquisicion", "compra", "fusion",
      "capitalizacion", "financiamiento", "financiacion", "credito"}, "Inversión y expansión"),
    ({"lanzamiento", "estreno", "presentacion", "apertura", "inauguracion",
      "reapertura", "activacion"}, "Lanzamientos y aperturas"),
    ({"investigacion", "denuncia", "demanda", "sancion", "multa", "condena",
      "querella", "imputacion", "fallo", "sentencia", "auditoria",
      "litigio", "proceso"}, "Litigios, investigaciones y sanciones"),
    ({"premio", "reconocimiento", "galardon", "distincion", "condecoracion",
      "certificacion", "acreditacion", "ranking", "homenaje"}, "Reconocimientos y premios"),
    ({"nombramiento", "designacion", "posesion", "renuncia", "dimision",
      "relevo", "sucesion", "eleccion"}, "Nombramientos y gobierno corporativo"),
    ({"reforma", "regulacion", "ley", "decreto", "norma", "normativa",
      "resolucion", "proyecto de ley", "aprobacion"}, "Regulación y política pública"),
    ({"obra", "construccion", "restauracion", "rehabilitacion", "remodelacion",
      "modernizacion", "renovacion", "mantenimiento", "infraestructura",
      "instalacion"}, "Proyectos y obras"),
    ({"crisis", "emergencia", "desastre", "sismo", "terremoto", "inundacion",
      "incendio", "derrame", "accidente", "colapso", "falla", "fallas",
      "racionamiento", "apagon", "paro", "bloqueo", "protesta"}, "Crisis y emergencias"),
    ({"alza", "aumento", "incremento", "reduccion", "caida", "descenso",
      "precio", "precios", "tarifa", "tarifas", "costo", "costos",
      "consumo", "demanda", "oferta", "ventas", "exportaciones",
      "importaciones", "produccion", "mercado"}, "Precios y comportamiento del mercado"),
    ({"balance", "resultados", "utilidad", "utilidades", "ingresos",
      "perdidas", "ganancias", "rentabilidad", "presupuesto", "recaudo"},
     "Resultados y desempeño financiero"),
    ({"campana", "publicidad", "patrocinio", "comunicado", "posicionamiento"},
     "Comunicación y posicionamiento de marca"),
    ({"donacion", "solidaridad", "ayuda", "apoyo", "voluntariado",
      "acompanamiento", "asistencia"}, "Responsabilidad social y comunidad"),
    ({"empleo", "contratacion", "capacitacion", "formacion", "salario",
      "nomina", "huelga", "sindicato", "despidos"}, "Talento humano y gestión interna"),
    ({"estudio", "informe", "encuesta", "analisis", "diagnostico",
      "evaluacion", "medicion"}, "Estudios e informes"),
    ({"foro", "congreso", "cumbre", "feria", "seminario", "taller",
      "encuentro", "conversatorio", "jornada", "evento", "festival"},
     "Eventos y participación institucional"),
    ({"proyecto", "programa", "plan", "iniciativa", "estrategia", "convocatoria",
      "licitacion", "adjudicacion", "contrato"}, "Proyectos y programas"),
    ({"brote", "contagio", "epidemia", "vacunacion", "virus", "salud",
      "bioseguridad", "atencion"}, "Salud pública y bioseguridad"),
    ({"cierre", "suspension", "clausura", "cancelacion", "retiro"},
     "Cierres y suspensiones"),
    ({"beca", "becas", "matricula", "graduacion", "educacion", "curriculo"},
     "Educación y formación"),
]


def _tema_por_cabeza_de_hecho(subtema: str) -> str:
    """Categoría editorial derivada del TIPO DE HECHO que encabeza el subtema.
    Agnóstico al sector: sirve igual para un banco, un hospital o una alcaldía."""
    toks = [_normaliza_token(t) for t in re.findall(r"[A-Za-zÁÉÍÓÚÑÜáéíóúñü]+", str(subtema or ""))]
    toks = [t for t in toks if t and t not in _CONECTORES_ETIQUETA]
    if not toks:
        return ""
    for tok in toks:                      # la cabeza manda, luego el resto
        for claves, tema in _TEMA_POR_CABEZA:
            if tok in claves:
                return tema
        st = _stem_es(tok)
        for claves, tema in _TEMA_POR_CABEZA:
            if any(st == _stem_es(k) for k in claves):
                return tema
    return ""

def _tema_es_igual_a_subtema(tema: str, subtemas_grupo: list) -> bool:
    if not tema or not subtemas_grupo: return False
    tn = string_norm_label(tema)
    for sub in subtemas_grupo:
        sn = string_norm_label(sub)
        if not tn or not sn: continue
        if SequenceMatcher(None, tn, sn).ratio() >= 0.80: return True
        if tn in sn or sn in tn: return True
    return False

def _generar_nombre_tema_llm(subtemas_grupo, textos_muestra, titulos_muestra, marca=""):
    subs_list = "\n".join(f"  · {s}" for s in subtemas_grupo[:8])
    palabras = []
    for t in titulos_muestra[:15]:
        for w in string_norm_label(str(t)).split():
            if len(w) > 3: palabras.append(w)
    kw = ", ".join(w for w, _ in Counter(palabras).most_common(6))
    tit_muestra = "\n".join(f"  · {t[:100]}" for t in list(dict.fromkeys(titulos_muestra))[:5])
    prompt = (
        f"Eres analista de reputación de la marca principal '{marca}'. "
        "Crea UN tema editorial preciso (2-5 palabras) que agrupe estos subtemas y describa el ámbito del hecho. "
        "Si el hecho NO está vinculado con la marca, crea el tema real de la noticia; NO lo fuerces a la marca ni a su sector.\n\n"
        "SUBTEMAS:\n" + subs_list + "\n\nTÍTULOS DE REFERENCIA:\n" + tit_muestra +
        f"\n\nKEYWORDS: {kw}\n\n"
        "REGLAS ESTRICTAS:\n"
        "  1. Conserva el asunto común que diferencia este grupo; NO uses secciones vagas de una palabra.\n"
        "  2. Debe ser más general que los subtemas, pero no abstracto: nunca copies un titular ni repitas un subtema.\n"
        "  3. NUNCA incluyas números, cantidades ni nombres propios.\n"
        "  4. 2-5 palabras, sustantivo + complemento/adjetivo.\n"
        "  5. Tildes y ñ correctas.\n\n"
        "CORRECTO: 'Regulación financiera', 'Movilidad urbana', 'Infraestructura vial', 'Salud pública territorial'\n"
        "INCORRECTO: 'Economía', 'Política', 'Actualidad', 'Destacados del sector', 'Panorama del sector', "
        "'Presencia en medios', 'Cinco congresistas con líos', 'Nuevo acuerdo'\n"
        "El tema nombra el ÁMBITO DEL HECHO (qué materia trata), NUNCA el soporte "
        "donde se publicó ni una frase de relleno sobre 'el sector'.\n\n"
        'JSON: {"tema":"..."}'
    )
    try:
        resp = call_with_retries(
            openai.ChatCompletion.create,
            model=OPENAI_MODEL_CLASIFICACION,
            messages=[{"role": "user", "content": prompt}],
            max_tokens=120,
            temperature=0.05,
            response_format={"type": "json_object"}
        )
        raw = json.loads(resp.choices[0].message.content).get("tema", "").strip().replace('"', '').replace('.', '')
        nombre = limpiar_tema(raw)
        if not _validar_estructura_tema(nombre): return None
        return nombre
    except:
        return None

def _regenerar_tema_diferente(subtemas_grupo, titulos_muestra, intento=0):
    subs_list = ", ".join(subtemas_grupo[:8])
    prompt = (
        f"Subtemas: {subs_list}\n\n"
        "Genera UNA categoría precisa (2-5 palabras), diferente a los subtemas. "
        "Conserva el asunto común; no respondas una sección vaga de una palabra como Economía, Política o Actualidad. "
        "Tildes y ñ correctas, terminar en sustantivo/adjetivo.\n"
        'JSON: {"tema":"..."}'
    )
    try:
        resp = call_with_retries(
            openai.ChatCompletion.create,
            model=OPENAI_MODEL_CLASIFICACION,
            messages=[{"role": "user", "content": prompt}],
            max_tokens=120,
            temperature=0.2 + intento * 0.1,
            response_format={"type": "json_object"}
        )
        nombre = limpiar_tema(json.loads(resp.choices[0].message.content).get("tema", "").strip().replace('"', '').replace('.', ''))
        return nombre if _validar_estructura_tema(nombre) else None
    except:
        return None

def consolidar_temas(subtemas, textos, pbar, marca=""):
    n = len(textos)
    u = _umbrales_adaptativos(n)
    pbar.progress(0.05, "Preparando temas...")
    df = pd.DataFrame({'subtema': subtemas, 'texto': textos})
    us = list(df['subtema'].unique())
    if len(us) <= 1:
        pbar.progress(1.0, "Un tema")
        return [capitalizar_etiqueta(s) for s in subtemas]

    if n <= 5 and len(us) == n:
        pbar.progress(1.0, "Corpus pequeño: temas = subtemas")
        st.info(f"Temas: **{n}** (corpus pequeño — cada noticia tiene tema propio)")
        return [capitalizar_etiqueta(s) for s in subtemas]

    pbar.progress(0.10, "Representaciones...")
    textos_por_subtema = defaultdict(list)
    for i, sub in enumerate(subtemas): textos_por_subtema[sub].append(textos[i])
    repr_enriquecidas = [_construir_representacion_grupo(sub, textos_por_subtema[sub]) for sub in us]
    pbar.progress(0.20, "Embeddings contenido...")
    emb_repr = get_embeddings_batch(repr_enriquecidas)
    emb_labels = get_embeddings_batch(us)
    ae = get_embeddings_batch(textos)
    centroids_contenido = {}
    for sub in us:
        idxs = df.index[df['subtema'] == sub].tolist()[:50]
        vecs = [ae[i] for i in idxs if ae[i] is not None]
        if vecs: centroids_contenido[sub] = np.mean(vecs, axis=0)
    pbar.progress(0.35, "Similitudes...")
    vs = [s for s in us if s in centroids_contenido]
    if len(vs) < 2:
        pbar.progress(1.0, "Sin agrupación")
        return [capitalizar_etiqueta(s) for s in subtemas]
    idx_map = {s: i for i, s in enumerate(us)}
    M_content = np.array([centroids_contenido[s] for s in vs])
    sim_content = cosine_similarity(M_content)
    has_repr = all(emb_repr[idx_map[s]] is not None for s in vs)
    has_label = all(emb_labels[idx_map[s]] is not None for s in vs)
    if has_repr and has_label:
        sim_combined = (0.50 * sim_content + 0.35 * cosine_similarity(np.array([emb_repr[idx_map[s]] for s in vs])) + 0.15 * cosine_similarity(np.array([emb_labels[idx_map[s]] for s in vs])))
    elif has_repr:
        sim_combined = (0.60 * sim_content + 0.40 * cosine_similarity(np.array([emb_repr[idx_map[s]] for s in vs])))
    else:
        sim_combined = sim_content

    pbar.progress(0.45, "Clustering temas...")
    dist_matrix = np.clip(1 - sim_combined, 0, 2)
    np.fill_diagonal(dist_matrix, 0)
    umbral_tema = u['tema']
    num_temas_max = u['num_temas_max']
    linkage_temas = 'complete' if len(vs) <= 6 else 'average'
    cl = AgglomerativeClustering(
        n_clusters=None, distance_threshold=1 - umbral_tema,
        metric='precomputed', linkage=linkage_temas
    ).fit(dist_matrix)

    clusters = defaultdict(list)
    for i, lbl in enumerate(cl.labels_): clusters[lbl].append(vs[i])
    clusters_validados = {}
    next_cluster_id = 0
    for _, subs_cluster in clusters.items():
        if len(subs_cluster) <= 1:
            clusters_validados[next_cluster_id] = subs_cluster
            next_cluster_id += 1
            continue
        dsu_tema = DSU(len(subs_cluster))
        for i in range(len(subs_cluster)):
            for j in range(i + 1, len(subs_cluster)):
                sa, sb = subs_cluster[i], subs_cluster[j]
                if _grupos_contenido_compatibles(
                    textos_por_subtema.get(sa, []),
                    textos_por_subtema.get(sb, []),
                    sa,
                    sb,
                    min_sim=max(umbral_tema, 0.82),
                    min_overlap=0.16,
                ):
                    dsu_tema.union(i, j)
        for miembros in dsu_tema.grupos(len(subs_cluster)).values():
            clusters_validados[next_cluster_id] = [subs_cluster[i] for i in miembros]
            next_cluster_id += 1
    clusters = clusters_validados
    uc = [s for s in us if s not in vs]
    mt = {}
    tc = len(clusters)
    pbar.progress(0.50, f"Nombres {tc} temas...")
    # Nombrado por cluster es INDEPENDIENTE → se paraleliza para acelerar (21 noticias).
    def _nombrar_cluster(cid, subtemas_cluster):
        titulos_cluster = []
        textos_cluster = []
        for sub in subtemas_cluster:
            for idx in df.index[df['subtema'] == sub].tolist()[:10]:
                txt = str(textos[idx])
                partes = txt.split('. ')
                if partes: titulos_cluster.append(partes[0][:120])
                textos_cluster.append(txt[:200])
        if len(subtemas_cluster) == 1:
            sub_unico = subtemas_cluster[0]
            nombre = _generar_nombre_tema_llm(subtemas_cluster, textos_cluster, titulos_cluster, marca)
            if not nombre or _tema_es_igual_a_subtema(nombre, subtemas_cluster):
                nombre = _regenerar_tema_diferente(subtemas_cluster, titulos_cluster)
            if not nombre or _tema_es_igual_a_subtema(nombre, subtemas_cluster):
                p = sub_unico.split()
                nombre = _recortar_frase_completa(" ".join(p), max_palabras=3) if len(p) > 3 else sub_unico
                if _tema_es_igual_a_subtema(nombre, subtemas_cluster): nombre = sub_unico
        else:
            nombre = _generar_nombre_tema_llm(subtemas_cluster, textos_cluster, titulos_cluster, marca)
            if not nombre or _tema_es_igual_a_subtema(nombre, subtemas_cluster):
                nombre = _regenerar_tema_diferente(subtemas_cluster, titulos_cluster)
            if not nombre or _tema_es_igual_a_subtema(nombre, subtemas_cluster):
                nombre = _regenerar_tema_diferente(subtemas_cluster, titulos_cluster, intento=1)
            if not nombre or _tema_es_igual_a_subtema(nombre, subtemas_cluster):
                all_words = []
                for sub in subtemas_cluster:
                    for w in string_norm_label(sub).split():
                        if len(w) > 3: all_words.append(w)
                nombre = capitalizar_etiqueta(" ".join(w for w, _ in Counter(all_words).most_common(2))) if all_words else subtemas_cluster[0]
        if not _frase_esta_completa(nombre):
            nombre = _recortar_frase_completa(nombre, max_palabras=4)
            if not _frase_esta_completa(nombre):
                freq = Counter(subtemas)
                nombre = _recortar_frase_completa(max(subtemas_cluster, key=lambda s: freq.get(s, 0)), max_palabras=4)
        return {sub: capitalizar_etiqueta(nombre) for sub in subtemas_cluster}

    workers = min(int(CONCURRENT_REQUESTS), tc or 1)
    if workers > 1 and tc > 1:
        futuras = {}
        with ThreadPoolExecutor(max_workers=workers) as ex:
            for cid, subs in clusters.items():
                futuras[ex.submit(_nombrar_cluster, cid, subs)] = cid
            done = 0
            for f in as_completed(futuras):
                try:
                    for sub, nombre in f.result().items(): mt[sub] = nombre
                except Exception:
                    pass
                done += 1
                pbar.progress(0.50 + 0.35 * (done / max(tc, 1)), f"Tema {done}/{tc}...")
    else:
        for k, (cid, subtemas_cluster) in enumerate(clusters.items()):
            pbar.progress(0.50 + 0.35 * (k / max(tc, 1)), f"Tema {k + 1}/{tc}...")
            for sub, nombre in _nombrar_cluster(cid, subtemas_cluster).items(): mt[sub] = nombre
    for sub in uc: mt[sub] = capitalizar_etiqueta(sub)

    pbar.progress(0.87, "Validando pertenencia mínima a temas...")
    min_tema = u['min_pertenencia_tema']
    tf_inicial = [mt.get(sub, sub) for sub in subtemas]
    tema_agrupacion: Dict[str, list] = defaultdict(list)
    for i, tema in enumerate(tf_inicial):
        if ae[i] is not None: tema_agrupacion[tema].append(ae[i])
    tema_centroids: Dict[str, np.ndarray] = {
        t: np.mean(vecs, axis=0) for t, vecs in tema_agrupacion.items() if vecs
    }
    tf_validado: List[str] = []
    n_forzadas = 0
    for i, (sub, tema_asignado) in enumerate(zip(subtemas, tf_inicial)):
        emb = ae[i]
        if emb is not None and tema_asignado in tema_centroids:
            sim = cosine_similarity(np.array(emb).reshape(1, -1), tema_centroids[tema_asignado].reshape(1, -1))[0][0]
            if sim < min_tema:
                tf_validado.append(capitalizar_etiqueta(_recortar_frase_completa(sub, max_palabras=4)))
                n_forzadas += 1
                continue
        tf_validado.append(capitalizar_etiqueta(tema_asignado))
    if n_forzadas: st.caption(f"ℹ️ {n_forzadas} noticias con baja pertenencia al tema agrupado → tema propio asignado.")

    pbar.progress(0.88, "Dedup temas...")
    tf_validado = dedup_labels(tf_validado, u['dedup_label'])

    pbar.progress(0.90, "Fusionando temas solapados...")
    mapa_fusion_temas = _fusionar_temas_contenidos(tf_validado)
    if mapa_fusion_temas:
        tf_validado = [mapa_fusion_temas.get(t, t) for t in tf_validado]

    pbar.progress(0.92, "Validando tema ≠ subtema...")
    tf_validado = _post_validar_tema_vs_subtema(tf_validado, subtemas)
    pbar.progress(0.95, "Completitud...")
    tf_validado = [capitalizar_etiqueta(_recortar_frase_completa(t) if not _frase_esta_completa(t) else t) for t in tf_validado]
    tf_validado = _unificar_tema_por_subtema(tf_validado, subtemas, textos)
    st.info(f"Temas: **{len(set(tf_validado))}** (de {len(set(subtemas))} subtemas) · Máx: {num_temas_max}")
    pbar.progress(1.0, "Temas listos")
    return tf_validado

def _fusionar_temas_contenidos(temas: List[str]) -> Dict[str, str]:
    unique = list(dict.fromkeys(temas))
    if len(unique) < 2: return {}
    normed = {t: string_norm_label(t) for t in unique}
    mapa: Dict[str, str] = {}
    for i, ta in enumerate(unique):
        for tb in unique[i + 1:]:
            na, nb = normed[ta], normed[tb]
            if not na or not nb: continue
            if na == nb or SequenceMatcher(None, na, nb).ratio() >= 0.92:
                canon = tb if len(tb) >= len(ta) else ta
                reemplazar = ta if canon == tb else tb
                mapa[reemplazar] = canon
    umbral_relajado = 0.88
    candidatos = [(t, normed[t]) for t in unique if len(t.split()) <= 3 and t not in mapa]
    if len(candidatos) >= 2:
        textos_c = [t for t, _ in candidatos]
        embs = get_embeddings_batch(textos_c)
        validos = [(textos_c[i], embs[i]) for i in range(len(textos_c)) if embs[i] is not None]
        if len(validos) >= 2:
            etqs, vecs = zip(*validos)
            sim = cosine_similarity(np.array(vecs))
            for i in range(len(etqs)):
                for j in range(i + 1, len(etqs)):
                    if sim[i][j] >= umbral_relajado:
                        ta, tb = etqs[i], etqs[j]
                        if ta in mapa or tb in mapa: continue
                        if _etiquetas_compatibles(ta, tb, min_overlap=0.60):
                            freq = Counter(temas)
                            canon = ta if freq.get(ta, 0) >= freq.get(tb, 0) else tb
                            reemplazar = tb if canon == ta else ta
                            mapa[reemplazar] = canon
    return mapa

def _post_validar_tema_vs_subtema(temas, subtemas):
    tema_a_subtemas = defaultdict(set)
    for t, s in zip(temas, subtemas): tema_a_subtemas[t].add(s)
    reemplazos = {}
    for tema, subs in tema_a_subtemas.items():
        if len(subs) == 1:
            sub_unico = list(subs)[0]
            tn = string_norm_label(tema)
            sn = string_norm_label(sub_unico)
            if tn and sn and SequenceMatcher(None, tn, sn).ratio() >= 0.80:
                nuevo = _regenerar_tema_diferente([sub_unico], [])
                if nuevo and not _tema_es_igual_a_subtema(nuevo, [sub_unico]) and _frase_esta_completa(nuevo):
                    reemplazos[tema] = capitalizar_etiqueta(nuevo)
    return [reemplazos.get(t, t) for t in temas] if reemplazos else temas

def _unificar_tema_por_subtema(temas, subtemas, textos=None):
    """Un mismo Subtema (sin importar mayúsculas) debe tener un único Tema.
    Además GARANTIZA que ninguna fila con Subtema quede sin Tema válido:
    el voto por subtema ignora temas vacíos/genéricos y, si el subtema no tiene
    ningún tema utilizable, se deriva por taxonomía léxica del propio texto."""
    vacios = {"", "nan", "n/a", "-", "sin tema", "varios", "none"}
    sub_to_temas = defaultdict(list)
    for t, s in zip(temas, subtemas):
        k = string_norm_label(s)
        if not k or k in vacios:
            continue
        sub_to_temas[k].append(t)
    sub_to_best = {}
    for k, tema_list in sub_to_temas.items():
        # Solo votan los temas realmente informativos.
        validos = [
            t for t in tema_list
            if str(t).strip().lower() not in vacios and not _es_etiqueta_generica(t)
        ]
        if validos:
            sub_to_best[k] = Counter(validos).most_common(1)[0][0]
    out = []
    for t, s in zip(temas, subtemas):
        k = string_norm_label(s)
        out.append(sub_to_best[k] if k in sub_to_best else t)
    # Red final: ninguna fila con Subtema puede quedarse sin Tema (o con Tema==Subtema).
    return _asegurar_tema_valido(out, list(subtemas), textos)

# ======================================
# Duplicados y Excel (Reglas Nuevas)
# ======================================
def _normalizar_url(url: str) -> str:
    if not url: return ""
    url = url.strip().lower()
    url = re.sub(r'^https?://', '', url)
    url = re.sub(r'^www\.', '', url)
    url = url.rstrip('/')
    return url

def detectar_duplicados_avanzado(rows, km):
    processed = deepcopy(rows)
    seen_url, seen_bcast = {}, {}
    seen_streaming: Dict[tuple, int] = {}
    tb = defaultdict(list)

    for i, row in enumerate(processed):
        if row.get("is_duplicate"): continue

        tipo    = normalizar_tipo_medio(str(row.get(km["tipodemedio"], "")))
        mencion = norm_key(row.get(km["menciones"], ""))
        medio   = norm_key(row.get(km["medio"], ""))

        streaming_url_raw = row.get(km["link_streaming"])
        if isinstance(streaming_url_raw, dict):
            streaming_url_raw = streaming_url_raw.get("url")
            
        if streaming_url_raw and mencion:
            streaming_url_norm = _normalizar_url(str(streaming_url_raw))
            if streaming_url_norm:
                sk = (streaming_url_norm, mencion)
                if sk in seen_streaming:
                    row["is_duplicate"] = True
                    row[km["idduplicada"]] = processed[seen_streaming[sk]].get(km["idnoticia"], "")
                    continue
                seen_streaming[sk] = i

        if tipo == "Internet":
            li = row.get(km["link_nota"])
            url = li.get("url") if isinstance(li, dict) else li
            if url and mencion:
                url_norm = _normalizar_url(str(url))
                k = (url_norm, mencion)
                if k in seen_url:
                    row["is_duplicate"] = True
                    row[km["idduplicada"]] = processed[seen_url[k]].get(km["idnoticia"], "")
                    continue
                seen_url[k] = i
            if medio and mencion:
                tb[(medio, mencion)].append(i)

        elif tipo in ("Radio", "Televisión"):
            hora = str(row.get(km["hora"], "")).strip()
            if mencion and medio and hora:
                k = (mencion, medio, hora)
                if k in seen_bcast:
                    row["is_duplicate"] = True
                    row[km["idduplicada"]] = processed[seen_bcast[k]].get(km["idnoticia"], "")
                else:
                    seen_bcast[k] = i

    for idxs in tb.values():
        if len(idxs) < 2: continue
        for i in range(len(idxs)):
            for j in range(i + 1, len(idxs)):
                a, b = idxs[i], idxs[j]
                if processed[a].get("is_duplicate") or processed[b].get("is_duplicate"): continue
                ta  = normalize_title_for_comparison(processed[a].get(km["titulo"]))
                tb_ = normalize_title_for_comparison(processed[b].get(km["titulo"]))
                if ta and tb_ and SequenceMatcher(None, ta, tb_).ratio() >= SIMILARITY_THRESHOLD_TITULOS:
                    if len(ta) < len(tb_):
                        processed[a]["is_duplicate"] = True
                        processed[a][km["idduplicada"]]  = processed[b].get(km["idnoticia"], "")
                    else:
                        processed[b]["is_duplicate"] = True
                        processed[b][km["idduplicada"]]  = processed[a].get(km["idnoticia"], "")

    return processed

def read_and_normalize_dossier(sheet, region_map, internet_map):
    headers = [cell.value for cell in sheet[1] if cell.value is not None]
    rows = []
    for row in sheet.iter_rows(min_row=2):
        if all(c.value is None for c in row):
            continue
        row_data = {}
        for i, h in enumerate(headers):
            if i < len(row):
                cell = row[i]
                val = cell.value
                url = cell.hyperlink.target if (cell.hyperlink and cell.hyperlink.target) else None
                if url:
                    row_data[h] = {"value": val or "Link", "url": url}
                else:
                    row_data[h] = val
        rows.append(row_data)

    df = pd.DataFrame(rows)

    tipo_medio_map = {
        'online': 'Internet', 'internet': 'Internet',
        'diario': 'Prensa',
        'am': 'Radio', 'fm': 'Radio',
        'aire': 'Televisión', 'cable': 'Televisión',
        'revista': 'Revistas', 'revistas': 'Revistas',
    }
    
    if 'Tipo de Medio' in df.columns:
        df['Tipo de Medio'] = (
            df['Tipo de Medio'].astype(str).str.lower().str.strip()
            .map(tipo_medio_map)
            .fillna(df['Tipo de Medio'].astype(str).str.strip())
        )
    else:
        df['Tipo de Medio'] = 'Otro'

    is_av = df['Tipo de Medio'].isin(['Radio', 'Televisión'])
    is_grafica = df['Tipo de Medio'].isin(['Prensa', 'Internet', 'Revistas'])
    is_internet = df['Tipo de Medio'] == 'Internet'

    if 'Medio' in df.columns:
        raw_medios_clean = df['Medio'].astype(str).str.lower().str.strip()
        df['Región'] = raw_medios_clean.map(region_map).fillna("N/A")
    else:
        df['Medio'] = 'N/A'
        df['Región'] = 'N/A'

    if 'Medio' in df.columns:
        df.loc[is_internet, 'Medio'] = (
            df.loc[is_internet, 'Medio']
            .astype(str).str.lower().str.strip()
            .map(internet_map)
            .fillna(df.loc[is_internet, 'Medio'])
        )

    df['ID Noticia'] = df.get('NoticiaId', df.get('ID Noticia', pd.Series(dtype=str)))
    df['Fecha'] = pd.to_datetime(df.get('Fecha', pd.Series(dtype=str)), dayfirst=True, errors='coerce').dt.normalize()
    df['Hora'] = df.get('Hora', pd.Series(dtype=str))
    df['Sección - Programa'] = df.get('Sección - Programa', pd.Series(dtype=str)).astype(str).apply(clean_text)
    
    titulo_col = 'Título' if 'Título' in df.columns else 'Titulo'
    df['Título'] = df.get(titulo_col, pd.Series(dtype=str)).astype(str).apply(clean_text)
    df['Autor - Conductor'] = df.get('Autor - Conductor', pd.Series(dtype=str)).astype(str).apply(clean_text)
    df['Nro. Pagina'] = df.get('Nro. Pagina', pd.Series(dtype=str))
    
    dim_col = 'Dimensioncm2' if 'Dimensioncm2' in df.columns else 'Dimensión'
    df['Dimensión'] = df.get(dim_col, pd.Series(dtype=str))
    df['Duración - Nro. Caracteres'] = df.get('Duración - Nro. Caracteres', pd.Series(dtype=str))

    df.loc[is_av, 'Dimensión'] = df.loc[is_av, 'Duración - Nro. Caracteres']
    df.loc[is_av, 'Duración - Nro. Caracteres'] = 0

    cpe_av = df.get('CPE', pd.Series([np.nan] * len(df)))
    cpe_grafica = df.get('Valor de Nota', pd.Series([np.nan] * len(df)))
    df['CPE'] = np.where(is_av, cpe_av, np.where(is_grafica, cpe_grafica, np.nan))

    df['Tier'] = df.get('Tier', pd.Series(dtype=str))
    df['Audiencia'] = df.get('Audiencia', pd.Series(dtype=str))
    df['Tono'] = df.get('Tono', pd.Series(dtype=str)).astype(str).apply(clean_text)
    df['Tema'] = df.get('Tematica', df.get('Tema', pd.Series(dtype=str))).astype(str).apply(clean_text)
    df['Temas Generales - Tema'] = df.get('Temas Generales - Tema', pd.Series(dtype=str)).astype(str).apply(clean_text)

    cuerpo_col = 'CuerpoEs' if 'CuerpoEs' in df.columns else 'Resumen - Aclaracion'
    cuerpo_cleaned = df.get(cuerpo_col, pd.Series([''] * len(df))).astype(str).apply(clean_cuerpo)

    def fmt_grafica(text):
        if not isinstance(text, str) or not text.strip():
            return text
        parrafos = [p.strip() for p in text.split('\n') if p.strip()]
        return '\n\n'.join(parrafos) if len(parrafos) > 1 else text

    df['Resumen - Aclaracion'] = np.where(is_av, cuerpo_cleaned, cuerpo_cleaned.apply(fmt_grafica))

    # ── ADICIÓN: columna con el CuerpoEs COMPLETO, sin truncar ──────────────
    # Se guarda tal cual queda cuerpo_cleaned (HTML limpio, <br> -> saltos de línea),
    # SIN pasar por corregir_texto() (que es lo que recorta/añade "..." al final).
    df['Cuerpo Completo'] = cuerpo_cleaned

    url_nota_av = df.get('URL Nota AV', df.get('Link Nota AV', pd.Series([''] * len(df))))
    url_streaming = df.get('URL (Streaming - Imagen)', pd.Series([''] * len(df)))
    
    link_nota_final = []
    for val_av, val_str, is_av_row in zip(url_nota_av, url_streaming, is_av):
        if is_av_row:
            if isinstance(val_av, dict):
                url_t = val_av.get("url", "")
                link_nota_final.append({"value": "Link", "url": url_t.replace(".com.ar", ".com.co") if url_t else None})
            else:
                url_t = str(val_av or "")
                link_nota_final.append({"value": "Link", "url": url_t.replace(".com.ar", ".com.co") if url_t else None})
        else:
            if isinstance(val_str, dict):
                link_nota_final.append(val_str)
            else:
                link_nota_final.append({"value": "Link", "url": val_str if val_str else None})
                
    df['Link Nota'] = link_nota_final

    url_nota_raw = df.get('URL Nota', pd.Series([''] * len(df)))
    link_stream_final = []
    for val_url, is_int in zip(url_nota_raw, is_internet):
        if is_int:
            if isinstance(val_url, dict):
                link_stream_final.append(val_url)
            else:
                link_stream_final.append({"value": "Link", "url": val_url if val_url else None})
        else:
            link_stream_final.append(None)
            
    df['Link (Streaming - Imagen)'] = link_stream_final

    menciones_av = df.get('Menciones - Empresa', pd.Series([''] * len(df))).fillna('').astype(str).apply(clean_text)
    menciones_grafica = df.get('Empresa rel.', pd.Series([''] * len(df))).fillna('').astype(str).apply(clean_text)
    df['Menciones - Empresa'] = np.where(is_av, menciones_av, np.where(is_grafica, menciones_grafica, menciones_av))

    return df

def generate_output_excel(rows, km):
    wb = Workbook()
    ws = wb.active
    ws.title = "Resultado"
    ORDER = [
        "ID Noticia", "Fecha", "Hora", "Medio", "Tipo de Medio",
        "Sección - Programa", "Región", "Título", "Autor - Conductor",
        "Nro. Pagina", "Dimensión", "Duración - Nro. Caracteres",
        "CPE", "Audiencia", "Tier", "Tono IA", "Tema", "Subtema",
        "Link Nota", "Resumen - Aclaracion", "Link (Streaming - Imagen)", "Menciones - Empresa",
        "ID duplicada",
        "Cuerpo Completo"   # ── ADICIÓN: columna final con el CuerpoEs completo, sin truncar ──
    ]
    NUM = {"ID Noticia", "Nro. Pagina", "Dimensión", "Duración - Nro. Caracteres", "CPE", "Tier", "Audiencia"}
    ORDER += ["Contexto analizado", "Coincidencia marca", "Origen coincidencia", "Tono", "Grupo noticia"]
    ws.append(ORDER)
    
    font_hyperlink = Font(color="000000", underline=None)
    align_left = Alignment(horizontal='left')
    font_header = Font(bold=True)
    
    for i, col_name in enumerate(ORDER, start=1):
        cell = ws.cell(row=1, column=i)
        cell.font = font_header

    col_idx_map = {name: ORDER.index(name) + 1 for name in ORDER}
        
    for row in rows:
        ctx, match, origin = _brand_audit(row.get(km.get("titulo"), ""), row.get(km.get("resumen"), ""), st.session_state.get("brand_name", ""), st.session_state.get("brand_aliases", []), row.get("Cuerpo Completo"))
        row["Contexto analizado"] = _contexto_para_excel(ctx)
        row["Coincidencia marca"], row["Origen coincidencia"] = match, origin
        tk = km.get("titulo")
        if tk and tk in row: row[tk] = clean_title_for_output(row.get(tk))
        rk = km.get("resumen")
        if rk and rk in row: row[rk] = corregir_texto(row.get(rk))
        
        out, links = [], {}
        for ci, h in enumerate(ORDER, start=1):
            dk = km.get(norm_key(h), norm_key(h))
            val = row.get(h)
            cv = None
            
            if h == 'Fecha' and pd.notna(val):
                if isinstance(val, pd.Timestamp):
                    cv = val.to_pydatetime()
                elif isinstance(val, (datetime.datetime, datetime.date)):
                    cv = val
                else:
                    cv = str(val) if val is not None else None
            elif h in NUM:
                cv = parse_numeric(val)
            elif isinstance(val, dict) and "url" in val:
                cv = val.get("value", "Link")
                if val.get("url"): links[ci] = val["url"]
            elif val is not None:
                if isinstance(val, str) and val.startswith("http"):
                    cv = "Link"
                    links[ci] = val
                else:
                    cv = str(val)
            out.append(cv)
        ws.append(out)
        
        current_row = ws.max_row
        for ci, url in links.items():
            cell = ws.cell(row=current_row, column=ci)
            cell.hyperlink = url
            cell.font = font_hyperlink
            cell.alignment = align_left
            
        date_col_idx = ORDER.index("Fecha") + 1
        date_cell = ws.cell(row=current_row, column=date_col_idx)
        if isinstance(date_cell.value, (datetime.datetime, datetime.date)):
            date_cell.number_format = 'DD/MM/YYYY'
            
        cols_millares = ["Nro. Pagina", "Dimensión", "Duración - Nro. Caracteres", "Tier", "Audiencia"]
        for col_name in cols_millares:
            col_idx = col_idx_map[col_name]
            cell = ws.cell(row=current_row, column=col_idx)
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0'

        cpe_idx = col_idx_map["CPE"]
        cpe_cell = ws.cell(row=current_row, column=cpe_idx)
        if isinstance(cpe_cell.value, (int, float)):
            cpe_cell.number_format = '$#,##0'
            
    for i, col_name in enumerate(ORDER, start=1):
        letter = ws.cell(row=1, column=i).column_letter
        if col_name in ['Título', 'Resumen - Aclaracion', 'Cuerpo Completo']:
            ws.column_dimensions[letter].width = 50
        elif col_name in ['Link Nota', 'Link (Streaming - Imagen)']:
            ws.column_dimensions[letter].width = 15
        else:
            ws.column_dimensions[letter].width = 20
            
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ======================================
# Proceso principal
# ======================================
async def run_full_process_async(df_file, bn, ba, tpkl, epkl, mode, xlsx_bytes=None, cliente="", voceros="", enable_scraping=False):
    st.session_state.update({'tokens_input': 0, 'tokens_output': 0, 'tokens_embedding': 0})
    get_embedding_cache().reset_stats()  # mantiene los embeddings cacheados entre corridas (no los limpia)
    t0 = time.time()
    
    if "API" in mode:
        try:
            openai.api_key=st.secrets["OPENAI_API_KEY"]
            openai.aiosession.set(None)
        except:
            st.error("OPENAI_API_KEY no encontrado.")
            st.stop()
            
    with st.status("Paso 1 · Carga de Configuración y Dossier", expanded=True) as s:
        region_map, internet_map = load_config_from_sheets()

        wb_in = load_workbook(df_file, data_only=True)
        df_normalized = read_and_normalize_dossier(wb_in.active, region_map, internet_map)

        medios_sin_region = sorted(set(
            df_normalized.loc[df_normalized['Región'] == 'N/A', 'Medio']
            .astype(str).str.strip()
        ) - {'', 'nan', 'None'})
        if medios_sin_region:
            st.session_state["medios_sin_mapear"] = medios_sin_region
        
        rows_expanded = []
        for idx, row_series in df_normalized.iterrows():
            menciones = [m.strip() for m in str(row_series['Menciones - Empresa']).split(';') if m.strip()]
            if not menciones:
                row_dict = row_series.to_dict()
                row_dict['Menciones - Empresa'] = ""
                row_dict['original_index'] = idx
                row_dict['expanded_index'] = len(rows_expanded)
                row_dict['is_duplicate'] = False
                rows_expanded.append(row_dict)
            else:
                for m in menciones:
                    row_dict = row_series.to_dict()
                    row_dict['Menciones - Empresa'] = m
                    row_dict['original_index'] = idx
                    row_dict['expanded_index'] = len(rows_expanded)
                    row_dict['is_duplicate'] = False
                    rows_expanded.append(row_dict)

        km = {
            "idnoticia": "ID Noticia",
            "fecha": "Fecha",
            "hora": "Hora",
            "medio": "Medio",
            "tipodemedio": "Tipo de Medio",
            "seccion_programa": "Sección - Programa",
            "region": "Región",
            "titulo": "Título",
            "autor_conductor": "Autor - Conductor",
            "nro_pagina": "Nro. Pagina",
            "dimension": "Dimensión",
            "duracion_caracteres": "Duración - Nro. Caracteres",
            "cpe": "CPE",
            "tier": "Tier",
            "audiencia": "Audiencia",
            "tono": "Tono",
            "tonoiai": "Tono IA",
            "tema": "Tema",
            "subtema": "Subtema",
            "link_nota": "Link Nota",
            "resumen": "Resumen - Aclaracion",
            "link_streaming": "Link (Streaming - Imagen)",
            "menciones": "Menciones - Empresa",
            "idduplicada": "ID duplicada"
        }
        
        rows = detectar_duplicados_avanzado(rows_expanded, km)
        for row in rows:
            if row["is_duplicate"]:
                row["Tono IA"] = "Duplicada"
                row["Tema"] = "-"
                row["Subtema"] = "-"
                
        s.update(label="✓ Paso 1 completado", state="complete")
        
    with st.status("Paso 2 · Normalización", expanded=True) as s:
        s.update(label="✓ Paso 2 · Mapeos y normalizaciones aplicados", state="complete")
        
    gc.collect()
    ta = [r for r in rows if not r.get("is_duplicate")]
    
    if ta:
        df = pd.DataFrame(ta)
        df["_txt"] = df.apply(
            lambda r: _construir_texto_basico(r, km["titulo"], km["resumen"], bn, ba),
            axis=1
        )
        with st.status("Embeddings...", expanded=True) as s:
            _ = get_embeddings_batch(df["_txt"].tolist())
            s.update(label=f"✓ {get_embedding_cache().stats()}", state="complete")
            
        with st.status("Paso 3 · Tono (Reputación)", expanded=True) as s:
            pb = st.progress(0)
            if ("PKL" in mode or tpkl) and tpkl:
                res = analizar_tono_con_pkl(
                    df["_txt"].tolist(), tpkl,
                    titulos=df[km["titulo"]], resumenes=df[km["resumen"]],
                    marca=bn, aliases=ba, progress=pb,
                    cuerpos=df['Cuerpo Completo'] if 'Cuerpo Completo' in df.columns else None,
                )
                if res is None: st.stop()
            elif "API" in mode or "Híbrido" in mode:
                res = await ClasificadorTono(bn, ba).procesar_lote_async(
                    df["_txt"], pb, df[km["resumen"]], df[km["titulo"]],
                    df['Cuerpo Completo'] if 'Cuerpo Completo' in df.columns else None,
                )
            else:
                res = [{"tono": "N/A"}] * len(ta)
            df[km["tonoiai"]] = [r["tono"] for r in res]
            s.update(label="✓ Paso 3 · Tono (Reputación)", state="complete")
            
        with st.status("Paso 4 · Clasificación", expanded=True) as s:
            pb = st.progress(0)
            if "Solo Modelos PKL" in mode:
                subtemas = ["N/A"] * len(ta)
                temas    = ["N/A"] * len(ta)
            else:
                subtemas = ClasificadorSubtema(bn, ba).procesar_lote(
                    df["_txt"], pb, df[km["resumen"]], df[km["titulo"]]
                )
                temas = consolidar_temas(subtemas, df["_txt"].tolist(), pb, bn)
            df[km["subtema"]] = subtemas
            if epkl:
                tp = analizar_temas_con_pkl(df["_txt"].tolist(), epkl)
                if tp: df[km["tema"]] = tp
            else:
                df[km["tema"]] = temas
            df[km["tema"]] = _unificar_tema_por_subtema(df[km["tema"]].tolist(), df[km["subtema"]].tolist(), df["_txt"].tolist())
            df = aplicar_consistencia_grupos(df, km["titulo"], km["resumen"],
                                             km["tonoiai"], km["tema"], km["subtema"])
            s.update(label="✓ Paso 4 · Clasificación", state="complete")
            
        rm2 = df.set_index("expanded_index").to_dict("index")
        for idx, row in enumerate(rows):
            if not row.get("is_duplicate"):
                row.update(rm2.get(row.get("expanded_index"), {}))
                
    gc.collect()
    ci = (st.session_state['tokens_input']     / 1e6) * PRICE_INPUT_1M
    co = (st.session_state['tokens_output']    / 1e6) * PRICE_OUTPUT_1M
    ce = (st.session_state['tokens_embedding'] / 1e6) * PRICE_EMBEDDING_1M
    
    st.session_state["brand_name"] = bn
    st.session_state["brand_aliases"] = ba
    with st.status("Paso 5 · Informe", expanded=True) as s:
        st.session_state["output_data"]     = generate_output_excel(rows, km)
        st.session_state["output_filename"] = f"Informe_IA_{bn.replace(' ', '_')}_{datetime.datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        st.session_state["processing_complete"] = True
        st.session_state.update({
            "brand_name": bn, "brand_aliases": ba,
            "total_rows": len(rows), "unique_rows": len(ta), "duplicates": len(rows) - len(ta),
            "process_duration": f"{time.time() - t0:.0f}s",
            "process_cost": f"${ci + co + ce:.4f} USD",
            "cache_stats": get_embedding_cache().stats()
        })
        s.update(label=f"✓ Completado · {get_embedding_cache().stats()}", state="complete")

async def run_quick_async(df, tc, sc, bn, al):
    st.session_state.update({'tokens_input': 0, 'tokens_output': 0, 'tokens_embedding': 0})
    get_embedding_cache().reset_stats()  # mantiene embeddings cacheados entre corridas
    df['_txt'] = df.apply(lambda r: _construir_texto_basico(r, tc, sc, bn, al), axis=1)
    with st.status("Embeddings...", expanded=True) as s:
        _ = get_embeddings_batch(df['_txt'].tolist())
        s.update(label=f"✓ {get_embedding_cache().stats()}", state="complete")
    with st.status("Tono", expanded=True) as s:
        pb = st.progress(0)
        res = await ClasificadorTono(bn, al).procesar_lote_async(df["_txt"], pb, df[sc].fillna(''), df[tc].fillna(''))
        df['Tono IA'] = [r["tono"] for r in res]
        audits = [_brand_audit(r.get(tc, ''), r.get(sc, ''), bn, al, r.get('Cuerpo Completo')) for _, r in df.iterrows()]
        df['Contexto analizado'] = [_contexto_para_excel(a[0]) for a in audits]
        df['Coincidencia marca'], df['Origen coincidencia'] = [a[1] for a in audits], [a[2] for a in audits]
        s.update(label="✓ Tono", state="complete")
    with st.status("Clasificación", expanded=True) as s:
        pb = st.progress(0)
        subtemas = ClasificadorSubtema(bn, al).procesar_lote(df["_txt"], pb, df[sc].fillna(''), df[tc].fillna(''))
        df['Subtema'] = subtemas
        temas = consolidar_temas(subtemas, df["_txt"].tolist(), pb, bn)
        df['Tema'] = _unificar_tema_por_subtema(temas, subtemas, df["_txt"].tolist())
        df = aplicar_consistencia_grupos(df, tc, sc)
        s.update(label="✓ Clasificación", state="complete")
    df.drop(columns=['_txt'], inplace=True)
    ci = (st.session_state['tokens_input']     / 1e6) * PRICE_INPUT_1M
    co = (st.session_state['tokens_output']    / 1e6) * PRICE_OUTPUT_1M
    ce = (st.session_state['tokens_embedding'] / 1e6) * PRICE_EMBEDDING_1M
    st.session_state['quick_cost'] = f"${ci + co + ce:.4f} USD"
    return df

def gen_quick_excel(df, original_bytes=None):
    if original_bytes:
        wb = load_workbook(io.BytesIO(original_bytes))
        ws = wb.active
        start = ws.max_column + 1
        for offset, col in enumerate([c for c in df.columns if c not in list(ws.values)[0]], start):
            ws.cell(1, offset, col)
            for i, value in enumerate(df[col].tolist(), 2): ws.cell(i, offset, value)
        out = io.BytesIO(); wb.save(out); return out.getvalue()
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as w:
        df.to_excel(w, index=False, sheet_name='Analisis')
    return buf.getvalue()

def render_quick_tab():
    st.markdown('<div class="sec-label">Análisis rápido</div>', unsafe_allow_html=True)
    if 'quick_result' in st.session_state:
        st.markdown(
            '<div class="success-banner"><div class="success-icon">✓</div>'
            '<div><div class="success-title">Completado</div>'
            '<div class="success-sub">Listo para descargar</div></div></div>',
            unsafe_allow_html=True
        )
        st.metric("Costo", st.session_state.get('quick_cost', "$0.00"))
        st.dataframe(st.session_state.quick_result.head(10), use_container_width=True)
        st.download_button(
            "Descargar",
            data=gen_quick_excel(st.session_state.quick_result, st.session_state.get('quick_bytes')),
            file_name="Analisis_Rapido_IA.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            type="primary"
        )
        if st.button("Nuevo análisis"):
            for k in ('quick_result', 'quick_df', 'quick_name', 'quick_cost'):
                if k in st.session_state:
                    del st.session_state[k]
            st.rerun()
        return
    if 'quick_df' not in st.session_state:
        st.markdown("Sube un Excel con columnas de título y resumen.")
        f = st.file_uploader("Excel", type=["xlsx"], label_visibility="collapsed", key="qu")
        if f:
            try:
                st.session_state.quick_bytes = f.getvalue()
                st.session_state.quick_df   = pd.read_excel(io.BytesIO(st.session_state.quick_bytes))
                st.session_state.quick_name = f.name
                st.rerun()
            except Exception as e:
                st.error(f"Error: {e}")
    else:
        st.success(f"**{st.session_state.quick_name}** cargado")
        with st.form("qf"):
            cols = st.session_state.quick_df.columns.tolist()
            c1, c2 = st.columns(2)
            tc = c1.selectbox("Col. título", cols, _default_text_column_index(cols, ['Título', 'Titulo', 'Titular', 'Headline'], 0))
            sc = c2.selectbox("Col. resumen", cols, _default_text_column_index(cols, ['CuerpoEs', 'Resumen - Aclaración', 'Resumen - Aclaracion', 'Resumen', 'Cuerpo', 'Descripción', 'Descripcion'], 1))
            bn  = st.text_input("Marca",       placeholder="Ej: Bancolombia")
            bat = st.text_input("Alias (;)",   placeholder="Ej: Grupo Bancolombia;Ban")
            if st.form_submit_button("Analizar", use_container_width=True, type="primary"):
                if not bn:
                    st.error("Indica la marca.")
                else:
                    try:
                        openai.api_key = st.secrets["OPENAI_API_KEY"]
                        openai.aiosession.set(None)
                    except:
                        st.error("OPENAI_API_KEY no encontrada.")
                        st.stop()
                    al = [a.strip() for a in bat.split(";") if a.strip()]
                    with st.spinner("Procesando..."):
                        st.session_state.quick_result = asyncio.run(
                            run_quick_async(st.session_state.quick_df.copy(), tc, sc, bn, al)
                        )
                    st.rerun()
        if st.button("Otro archivo"):
            for k in ('quick_df', 'quick_name', 'quick_result', 'quick_cost'):
                if k in st.session_state:
                    del st.session_state[k]
            st.rerun()


# ======================================
# EXCEL PERSONALIZADO (Mantiene formato original + 3 columnas al final)
# ======================================
async def run_custom_excel_async(file_bytes, tc, sc, bn, al, mode="API de OpenAI", tpkl=None, epkl=None):
    st.session_state.update({'tokens_input': 0, 'tokens_output': 0, 'tokens_embedding': 0})
    get_embedding_cache().reset_stats()  # mantiene los embeddings cacheados entre corridas (no los limpia)
    t0 = time.time()

    # Cargar archivo usando openpyxl para conservar estilos y formato original
    buf_in = io.BytesIO(file_bytes)
    wb = load_workbook(buf_in)
    ws = wb.active

    # Cargar DataFrame solo para extraer textos e índices
    buf_in.seek(0)
    df = pd.read_excel(buf_in)

    df['_txt'] = df.apply(
        lambda r: _construir_texto_basico(r, tc, sc, bn, al),
        axis=1
    )

    with st.status("Paso 1 · Generando Embeddings...", expanded=True) as s:
        _ = get_embeddings_batch(df['_txt'].tolist())
        s.update(label=f"✓ Embeddings listos · {get_embedding_cache().stats()}", state="complete")

    # --- PASO 2: TONO ---
    with st.status("Paso 2 · Evaluando Tono (Reputación)...", expanded=True) as s:
        pb = st.progress(0)
        if tpkl:
            # PKL de tono: predecir sobre la mención a Marca principal / Alias, no el artículo entero.
            res = analizar_tono_con_pkl(
                df["_txt"].tolist(), tpkl,
                titulos=df[tc].fillna(""), resumenes=df[sc].fillna(""),
                marca=bn, aliases=al,
                cuerpos=df['Cuerpo Completo'] if 'Cuerpo Completo' in df.columns else None,
            )
            if res is None: st.stop()
            tonos = [r["tono"] for r in res]
        elif "API" in mode or "Híbrido" in mode:
            res = await ClasificadorTono(bn, al).procesar_lote_async(
                df["_txt"], pb, df[sc].fillna(''), df[tc].fillna(''),
                df['Cuerpo Completo'] if 'Cuerpo Completo' in df.columns else None,
            )
            tonos = [r["tono"] for r in res]
        else:
            tonos = ["N/A"] * len(df)
        df['Tono IA'] = tonos
        audits = [_brand_audit(r.get(tc, ''), r.get(sc, ''), bn, al, r.get('Cuerpo Completo')) for _, r in df.iterrows()]
        df['Contexto analizado'] = [_contexto_para_excel(a[0]) for a in audits]
        df['Coincidencia marca'], df['Origen coincidencia'] = [a[1] for a in audits], [a[2] for a in audits]
        s.update(label="✓ Tono IA evaluado", state="complete")

    # --- PASO 3: SUBTEMAS Y TEMAS ---
    with st.status("Paso 3 · Clasificando Subtemas y Temas...", expanded=True) as s:
        pb = st.progress(0)
        
        # Subtemas
        if "Solo Modelos PKL" in mode:
            subtemas = ["N/A"] * len(df)
        else:
            subtemas = ClasificadorSubtema(bn, al).procesar_lote(
                df["_txt"], pb, df[sc].fillna(''), df[tc].fillna('')
            )

        # Temas
        if epkl:
            # Si se subió PKL de Temas, usar las predicciones directas del modelo
            tp = analizar_temas_con_pkl(df["_txt"].tolist(), epkl)
            if tp:
                temas = tp
            else:
                temas = ["N/A"] * len(df)
        elif "Solo Modelos PKL" in mode:
            temas = ["N/A"] * len(df)
        else:
            temas = consolidar_temas(subtemas, df["_txt"].tolist(), pb, bn)

        df['Subtema'] = subtemas
        df['Tema']    = _unificar_tema_por_subtema(temas, subtemas, df["_txt"].tolist())
        df = aplicar_consistencia_grupos(df, tc, sc)
        s.update(label="✓ Clasificación completada", state="complete")

    # Escribir las 3 columnas adicionales al final en la hoja openpyxl respetando el formato original
    max_col = ws.max_column
    col_tono    = max_col + 1
    col_tema    = max_col + 2
    col_subtema = max_col + 3
    col_contexto = max_col + 4
    col_coincidencia = max_col + 5
    col_origen = max_col + 6

    # Encabezados en negrita
    font_bold = Font(bold=True)
    ws.cell(row=1, column=col_tono, value="Tono IA").font = font_bold
    ws.cell(row=1, column=col_tema, value="Tema").font = font_bold
    ws.cell(row=1, column=col_subtema, value="Subtema").font = font_bold
    ws.cell(row=1, column=col_contexto, value="Contexto analizado").font = font_bold
    ws.cell(row=1, column=col_coincidencia, value="Coincidencia marca").font = font_bold
    ws.cell(row=1, column=col_origen, value="Origen coincidencia").font = font_bold

    # Asignar valores por fila manteniendo la coincidencia exacta
    for idx, row_data in df.iterrows():
        r = idx + 2
        ws.cell(row=r, column=col_tono, value=str(row_data['Tono IA']))
        ws.cell(row=r, column=col_tema, value=str(row_data['Tema']))
        ws.cell(row=r, column=col_subtema, value=str(row_data['Subtema']))
        ws.cell(row=r, column=col_contexto, value=str(row_data['Contexto analizado']))
        ws.cell(row=r, column=col_coincidencia, value=str(row_data['Coincidencia marca']))
        ws.cell(row=r, column=col_origen, value=str(row_data['Origen coincidencia']))

    buf_out = io.BytesIO()
    wb.save(buf_out)

    ci = (st.session_state['tokens_input']     / 1e6) * PRICE_INPUT_1M
    co = (st.session_state['tokens_output']    / 1e6) * PRICE_OUTPUT_1M
    ce = (st.session_state['tokens_embedding'] / 1e6) * PRICE_EMBEDDING_1M

    cost_str = f"${ci + co + ce:.4f} USD"
    time_str = f"{time.time() - t0:.0f}s"

    return buf_out.getvalue(), df, cost_str, time_str


def render_custom_excel_tab():
    st.markdown('<div class="sec-label">Análisis de Excel Personalizado</div>', unsafe_allow_html=True)
    st.caption("Sube cualquier archivo Excel (.xlsx). Al finalizar se descargarán los mismos datos y formato original con 3 nuevas columnas añadidas al final: **Tono IA**, **Tema** y **Subtema**.")

    if 'custom_result_bytes' in st.session_state:
        st.markdown(
            '<div class="success-banner"><div class="success-icon">✓</div>'
            '<div><div class="success-title">Análisis de Excel Finalizado</div>'
            '<div class="success-sub">Se han añadido las 3 columnas al final del Excel original manteniendo su formato.</div></div></div>',
            unsafe_allow_html=True
        )
        c1, c2 = st.columns(2)
        c1.metric("Costo estimado", st.session_state.get('custom_cost', "$0.00"))
        c2.metric("Tiempo de ejecución", st.session_state.get('custom_time', "0s"))

        if 'custom_df_preview' in st.session_state:
            st.markdown("##### Vista previa del archivo (primeras filas con columnas añadidas):")
            st.dataframe(st.session_state.custom_df_preview.head(10), use_container_width=True)

        st.download_button(
            "⬇ Descargar Excel Actualizado",
            data=st.session_state.custom_result_bytes,
            file_name=f"Analisis_{st.session_state.get('custom_filename', 'Personalizado.xlsx')}",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            type="primary"
        )

        if st.button("Nuevo análisis personalizado"):
            for k in ('custom_result_bytes', 'custom_df', 'custom_filename', 'custom_cost', 'custom_time', 'custom_df_preview'):
                if k in st.session_state:
                    del st.session_state[k]
            st.rerun()
        return

    if 'custom_df' not in st.session_state:
        f = st.file_uploader("Sube cualquier archivo Excel (.xlsx)", type=["xlsx"], key="custom_uploader")
        if f:
            try:
                bytes_data = f.getvalue()
                df_temp = pd.read_excel(io.BytesIO(bytes_data))
                st.session_state.custom_df       = df_temp
                st.session_state.custom_bytes    = bytes_data
                st.session_state.custom_filename = f.name
                st.rerun()
            except Exception as e:
                st.error(f"Error al leer el archivo Excel: {e}")
    else:
        st.success(f"📁 Archivo cargado: **{st.session_state.custom_filename}** ({len(st.session_state.custom_df)} filas)")

        cols = st.session_state.custom_df.columns.tolist()

        with st.form("custom_form"):
            st.markdown('<div class="sec-label">Selección de Columnas</div>', unsafe_allow_html=True)
            c_col1, c_col2 = st.columns(2)
            tc = c_col1.selectbox("Columna que contiene el TÍTULO", cols, index=_default_text_column_index(cols, ['Título', 'Titulo', 'Titular', 'Headline'], 0))
            sc = c_col2.selectbox("Columna que contiene el RESUMEN / CUERPO", cols, index=_default_text_column_index(cols, ['CuerpoEs', 'Resumen - Aclaración', 'Resumen - Aclaracion', 'Resumen', 'Cuerpo', 'Descripción', 'Descripcion'], 1))

            st.markdown('<div class="sec-label">Configuración del Análisis</div>', unsafe_allow_html=True)
            cl, cr = st.columns([3, 2])
            with cl:
                bn  = st.text_input("Marca principal", placeholder="Ej: Bancolombia", key="custom_bn")
                bat = st.text_input("Alias (separados por ;)", placeholder="Ej: Grupo Bancolombia;Ban", key="custom_ba")
            with cr:
                mode = st.radio(
                    "Modo de análisis",
                    ["API de OpenAI", "Híbrido (PKL + API)", "Solo Modelos PKL"],
                    index=0, key="custom_mode"
                )

            tpkl, epkl = None, None
            st.markdown('<div class="sec-label">Modelos PKL (Opcionales)</div>', unsafe_allow_html=True)
            p1, p2 = st.columns(2)
            tpkl = p1.file_uploader("Modelo Sentimiento / Tono (.pkl)", type=["pkl"], key="custom_tpkl")
            epkl = p2.file_uploader("Modelo Temas (.pkl)", type=["pkl"], key="custom_epkl")

            if st.form_submit_button("▶ Iniciar análisis personalizado", use_container_width=True, type="primary"):
                if not bn.strip():
                    st.error("Ingresa el nombre de la marca principal.")
                elif "Solo Modelos PKL" in mode and not (tpkl or epkl):
                    st.error("Seleccionaste 'Solo Modelos PKL', por favor adjunta al menos un archivo .pkl para continuar.")
                else:
                    if "API" in mode or "Híbrido" in mode:
                        try:
                            openai.api_key = st.secrets["OPENAI_API_KEY"]
                            openai.aiosession.set(None)
                        except:
                            st.error("OPENAI_API_KEY no encontrada en st.secrets.")
                            st.stop()

                    al = [a.strip() for a in bat.split(";") if a.strip()]

                    with st.spinner("Procesando Excel personalizado..."):
                        res_bytes, res_df, cost_str, time_str = asyncio.run(
                            run_custom_excel_async(
                                st.session_state.custom_bytes,
                                tc, sc, bn, al,
                                mode=mode, tpkl=tpkl, epkl=epkl
                            )
                        )

                        st.session_state.custom_result_bytes = res_bytes
                        st.session_state.custom_df_preview   = res_df
                        st.session_state.custom_cost         = cost_str
                        st.session_state.custom_time         = time_str
                        st.rerun()

        if st.button("Subir otro archivo Excel"):
            for k in ('custom_df', 'custom_bytes', 'custom_filename', 'custom_result_bytes', 'custom_cost', 'custom_time', 'custom_df_preview'):
                if k in st.session_state:
                    del st.session_state[k]
            st.rerun()


# ======================================
# Main
# ======================================
async def run_sentiment_only_async(df, title_col, summary_col, brand, aliases, pkl_file=None):
    details = [extraer_contexto_marca_detallado(r.get(title_col, ''), r.get(summary_col, ''), brand, aliases, r.get('Cuerpo Completo')) for _, r in df.iterrows()]
    df = df.copy()
    idx = [i for i, d in enumerate(details) if d['contexto']]
    results = [{'tono':'Neutro','confianza':'Alta','justificacion':'La marca no aparece en el título ni en el resumen.'} for _ in details]
    if idx:
        pb = st.progress(0)
        if pkl_file:
            raw = analizar_tono_con_pkl([details[i]['contexto'] for i in idx], pkl_file)
        else:
            raw = await ClasificadorTono(brand, aliases).procesar_lote_async(pd.Series([details[i]['contexto'] for i in idx]), pb, pd.Series([df.iloc[i][summary_col] for i in idx]), pd.Series([df.iloc[i][title_col] for i in idx]))
        for i, r in zip(idx, raw or []): results[i].update(r)
    df['Tono IA'] = [r.get('tono','Neutro') for r in results]
    df['Confianza Tono'] = [r.get('confianza','Media') for r in results]
    df['Marca encontrada'] = [d['marca_encontrada'] for d in details]
    df['Contexto analizado'] = [_contexto_para_excel(d['contexto']) for d in details]
    df['Coincidencia marca'] = [d['coincidencia'] for d in details]
    df['Origen coincidencia'] = [d['origen'] for d in details]
    return df

def render_sentiment_tab():
    st.markdown('<div class="sec-label">Sentimiento por Marca</div>', unsafe_allow_html=True)
    st.caption('Analiza exclusivamente el impacto reputacional de la marca encontrada en título y resumen.')
    f = st.file_uploader('Sube un Excel (.xlsx)', type=['xlsx'], key='sentiment_uploader')
    if not f: return
    try:
        df = pd.read_excel(io.BytesIO(f.getvalue())); cols = df.columns.tolist()
        with st.form('sentiment_form'):
            tc = st.selectbox('Columna de título', cols, _default_text_column_index(cols, ['Título', 'Titulo', 'Titular', 'Headline'], 0))
            sc = st.selectbox('Columna de resumen / aclaración', cols, _default_text_column_index(cols, ['CuerpoEs', 'Resumen - Aclaración', 'Resumen - Aclaracion', 'Resumen', 'Cuerpo', 'Descripción', 'Descripcion'], 1))
            brand = st.text_input('Marca principal'); alias_text = st.text_input('Alias separados por ;')
            pkl = st.file_uploader('Modelo PKL opcional', type=['pkl'], key='sentiment_pkl')
            submit = st.form_submit_button('Analizar sentimiento', type='primary', use_container_width=True)
        if submit:
            if not brand.strip(): st.error('Ingresa la marca principal.')
            else:
                if not pkl: openai.api_key = st.secrets['OPENAI_API_KEY']
                aliases = [a.strip() for a in alias_text.split(';') if a.strip()]
                with st.spinner('Analizando menciones de la marca...'):
                    result = asyncio.run(run_sentiment_only_async(df, tc, sc, brand.strip(), aliases, pkl))
                st.dataframe(result.head(20), use_container_width=True)
                out = io.BytesIO()
                with pd.ExcelWriter(out, engine='openpyxl') as w: result.to_excel(w, index=False, sheet_name='Sentimiento')
                output_name = f"sentimiento_{_safe_filename_part(brand)}.xlsx"
                st.download_button('Descargar Excel de Sentimiento', out.getvalue(), output_name, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', use_container_width=True, type='primary')
    except Exception as e: st.error(f'Error durante el análisis: {e}')

def main():
    load_custom_css()
    if not check_password(): return

    st.markdown("""
    <div class="app-header">
        <div class="app-header-icon">◈</div>
        <div class="app-header-text">
            <div class="app-header-title">Análisis de Noticias - API</div>
            <div class="app-header-version">v18.5 · 😼 Realizado por Johnathan Cortés 🕵️‍♂️ </div>
        </div>
        <div class="app-header-badge">IA</div>
    </div>""", unsafe_allow_html=True)

    tab1, tab2, tab3, tab4 = st.tabs(["Análisis Completo", "Análisis Rápido", "Excel Personalizado", "Sentimiento"])

    with tab1:
        if not st.session_state.get("processing_complete", False):
            col_cfg1, col_cfg2 = st.columns([4, 1])
            with col_cfg1:
                st.markdown(
                    '<span class="config-badge">⚙ Configuración: Google Sheets (Regiones / Internet)</span>',
                    unsafe_allow_html=True
                )
            with col_cfg2:
                if st.button("↻ Refrescar config", use_container_width=True):
                    refresh_config_cache()
                    st.success("Config recargada")

            st.markdown('<div class="sec-label">Configuración</div>', unsafe_allow_html=True)
            cl, cr = st.columns([3, 2])
            with cl:
                bn  = st.text_input("Marca principal", placeholder="Ej: Bancolombia", key="bn")
                bat = st.text_input("Alias (separados por ;)", placeholder="Ej: Grupo Bancolombia;Ban", key="ba")
            with cr:
                mode = st.radio(
                    "Modo de análisis",
                    ["API de OpenAI", "Híbrido (PKL + API)", "Solo Modelos PKL"],
                    index=0, key="mode"
                )

            tpkl, epkl = None, None
            if "PKL" in mode:
                st.markdown('<div class="sec-label">Modelos PKL</div>', unsafe_allow_html=True)
                p1, p2 = st.columns(2)
                tpkl = p1.file_uploader(
                    "Modelo de Sentimiento (.pkl)", type=["pkl"], key="tpkl",
                    help="Pipeline sklearn para clasificar tono: -1/0/1 o Negativo/Neutro/Positivo"
                )
                epkl = p2.file_uploader(
                    "Modelo de Temas (.pkl)", type=["pkl"], key="epkl",
                    help="Pipeline sklearn para clasificar temas"
                )

            with st.form("main_form"):
                st.markdown('<div class="sec-label">Archivo de entrada</div>', unsafe_allow_html=True)
                st.markdown("""
                <div class="upload-zone" style="grid-template-columns:1fr">
                    <div class="upload-zone-card">
                        <div class="upload-zone-icon uz-dossier">📋</div>
                        <div class="upload-zone-text">
                            <div class="upload-zone-title">Dossier</div>
                            <div class="upload-zone-desc">Sube las noticias en el nuevo formato .xlsx a analizar</div>
                        </div>
                    </div>
                </div>""", unsafe_allow_html=True)
                f1 = st.file_uploader("Dossier", type=["xlsx"], label_visibility="collapsed", key="f1")

                st.markdown(
                    f'<div class="cluster-info">'
                    f'<b>Parámetros base</b> · Sub={UMBRAL_SUBTEMA} · Tema={UMBRAL_TEMA} · Máx={NUM_TEMAS_MAX} '
                    f'· FusInter={UMBRAL_FUSION_INTERGRUPO} · FusSem={UMBRAL_FUSION_SUBTEMAS} '
                    f'· Dedup={UMBRAL_DEDUP_LABEL} · MinSub={UMBRAL_MIN_PERTENENCIA_SUBTEMA} '
                    f'· MinTema={UMBRAL_MIN_PERTENENCIA_TEMA} · MaxGrupo={MAX_GRUPO_ETIQUETA} · '
                    f'<b>Coherencia={UMBRAL_COHERENCIA_ETIQUETA}</b> · '
                    f'<b>SimMin={SIM_MINIMA_AGRUPACION_SUBTEMA}</b> (adaptativos según n)'
                    f'</div>',
                    unsafe_allow_html=True
                )

                if st.form_submit_button("▶ Iniciar análisis", use_container_width=True, type="primary"):
                    if not all([f1, bn.strip()]):
                        st.error("Completa todos los campos.")
                    else:
                        al = [a.strip() for a in bat.split(";") if a.strip()]
                        cur_mode = st.session_state.get("mode", "API de OpenAI")
                        cur_tpkl = st.session_state.get("tpkl")
                        cur_epkl = st.session_state.get("epkl")
                        asyncio.run(run_full_process_async(f1, bn, al, cur_tpkl, cur_epkl, cur_mode,
                                                         xlsx_bytes=None, cliente="", voceros="",
                                                         enable_scraping=False))
                        st.rerun()
        else:
            total = st.session_state.total_rows
            uniq  = st.session_state.unique_rows
            dups  = st.session_state.duplicates
            dur   = st.session_state.process_duration
            cost  = st.session_state.get("process_cost", "$0.00")
            st.markdown(
                '<div class="success-banner"><div class="success-icon">✓</div>'
                '<div><div class="success-title">Análisis completado</div>'
                '<div class="success-sub">Informe listo para descargar</div></div></div>',
                unsafe_allow_html=True
            )

            medios_sin_mapear = st.session_state.get("medios_sin_mapear")
            if medios_sin_mapear:
                st.warning(
                    "⚠️ Los siguientes medios no tienen región asignada en el Sheets de "
                    f"'Regiones' (quedaron como N/A): {', '.join(medios_sin_mapear)}. "
                    "Agrégalos en el Google Sheets para que se mapeen automáticamente la próxima vez."
                )

            st.markdown(f"""
            <div class="metrics-grid">
              <div class="metric-card m-total"><div class="metric-val" style="color:var(--text)">{total}</div><div class="metric-lbl">Total</div></div>
              <div class="metric-card m-unique"><div class="metric-val" style="color:var(--green)">{uniq}</div><div class="metric-lbl">Únicas</div></div>
              <div class="metric-card m-dup"><div class="metric-val" style="color:var(--amber)">{dups}</div><div class="metric-lbl">Duplicados</div></div>
              <div class="metric-card m-time"><div class="metric-val" style="color:var(--blue)">{dur}</div><div class="metric-lbl">Tiempo</div></div>
              <div class="metric-card m-cost"><div class="metric-val" style="color:var(--accent)">{cost}</div><div class="metric-lbl">Costo</div></div>
            </div>""", unsafe_allow_html=True)
            if 'cache_stats' in st.session_state: st.caption(f"📊 {st.session_state['cache_stats']}")
            c1, c2 = st.columns(2)
            c1.download_button(
                "⬇ Descargar informe",
                data=st.session_state.output_data,
                file_name=st.session_state.output_filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                type="primary"
            )
            if c2.button("Nuevo análisis", use_container_width=True):
                pwd = st.session_state.get("password_correct")
                st.session_state.clear()
                st.session_state.password_correct = pwd
                st.rerun()

    with tab2:
        render_quick_tab()

    with tab3:
        render_custom_excel_tab()

    with tab4:
        render_sentiment_tab()

    st.markdown(
        '<div class="footer">v18.2 · Análisis de Noticias con IA · Johnathan Cortés ©</div>',
        unsafe_allow_html=True
    )

if __name__ == "__main__":
    main()

