# ======================================
# Estilo de hipervínculos Link Nota / Link (Streaming)
# ======================================
import io
import unittest
from zipfile import ZipFile
from xml.etree import ElementTree as ET

from openpyxl import load_workbook

from pipeline import BASE_OUTPUT_COLUMNS, KEY_MAP, PLAIN_HYPERLINK_COLUMNS, generate_output_excel

NS_MAIN = {"m": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}


def _sample_rows():
    return [
        {
            "ID Noticia": 101,
            "Título": "Nota de prueba",
            "Link Nota": {"value": "Link", "url": "https://example.com/nota"},
            "Link (Streaming - Imagen)": {"value": "Link", "url": "https://example.com/stream"},
            "resumen corto": "texto",
        }
    ]


def _font_is_black(font) -> bool:
    color = getattr(font, "color", None)
    if color is None:
        return True
    rgb = getattr(color, "rgb", None)
    if rgb is None:
        theme = getattr(color, "theme", None)
        indexed = getattr(color, "indexed", None)
        if theme is None and indexed is None:
            return True
        return False
    hex_rgb = str(rgb).upper().replace(" ", "")
    return hex_rgb.endswith("000000")


def _font_not_underlined(font) -> bool:
    underline = getattr(font, "underline", None)
    return underline in (None, False, "none", "None")


class LinkExportStyleTests(unittest.TestCase):
    def test_link_nota_and_streaming_are_black_without_underline(self):
        data = generate_output_excel(_sample_rows(), KEY_MAP)
        wb = load_workbook(io.BytesIO(data))
        ws = wb["Resultado"]
        headers = [cell.value for cell in ws[1]]

        for col_name in ("Link Nota", "Link (Streaming - Imagen)"):
            self.assertIn(col_name, PLAIN_HYPERLINK_COLUMNS)
            self.assertIn(col_name, headers)
            col_idx = headers.index(col_name) + 1
            cell = ws.cell(row=2, column=col_idx)
            self.assertEqual(cell.value, "Link")
            self.assertIsNotNone(cell.hyperlink, f"{col_name} debe conservar el hipervínculo")
            target = cell.hyperlink.target or cell.hyperlink.ref
            self.assertTrue(str(target).startswith("https://example.com/"))
            self.assertTrue(_font_is_black(cell.font), f"{col_name} debe ser texto negro, color={cell.font.color}")
            self.assertTrue(
                _font_not_underlined(cell.font),
                f"{col_name} no debe ir subrayado, underline={cell.font.underline}",
            )

        title_idx = headers.index("Título") + 1
        title_cell = ws.cell(row=2, column=title_idx)
        self.assertEqual(title_cell.value, "Nota de prueba")
        self.assertIsNone(title_cell.hyperlink)

    def test_xlsx_xml_keeps_hyperlink_relationship_with_plain_font(self):
        data = generate_output_excel(_sample_rows(), KEY_MAP)
        with ZipFile(io.BytesIO(data)) as zf:
            sheet = zf.read("xl/worksheets/sheet1.xml")
            styles = zf.read("xl/styles.xml")
            rels = zf.read("xl/worksheets/_rels/sheet1.xml.rels")

        sheet_tree = ET.fromstring(sheet)
        hyperlinks = [
            el for el in sheet_tree.iter() if el.tag.endswith("hyperlink")
        ]
        self.assertGreaterEqual(len(hyperlinks), 2)

        rel_tree = ET.fromstring(rels)
        external = [
            rel.attrib.get("Target")
            for rel in rel_tree
            if rel.attrib.get("TargetMode") == "External"
        ]
        self.assertIn("https://example.com/nota", external)
        self.assertIn("https://example.com/stream", external)

        styles_tree = ET.fromstring(styles)
        fonts = styles_tree.findall("m:fonts/m:font", NS_MAIN)
        plain_black = False
        for font in fonts:
            color = font.find("m:color", NS_MAIN)
            underline = font.find("m:u", NS_MAIN)
            rgb = (color.attrib.get("rgb") if color is not None else "") or ""
            if rgb.upper().endswith("000000") and underline is None:
                plain_black = True
                break
        self.assertTrue(plain_black, "Debe existir una fuente negra sin subrayado en styles.xml")

    def test_plain_hyperlink_columns_are_only_the_two_requested(self):
        self.assertEqual(
            PLAIN_HYPERLINK_COLUMNS,
            frozenset({"Link Nota", "Link (Streaming - Imagen)"}),
        )
        for col in PLAIN_HYPERLINK_COLUMNS:
            self.assertIn(col, BASE_OUTPUT_COLUMNS)


if __name__ == "__main__":
    unittest.main()
